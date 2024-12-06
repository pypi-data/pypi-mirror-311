#!/usr/bin/env python3
# from case_manage.lib.excel import Excel
from typing import Dict, List, Optional
from pathlib import Path
import os
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment
from openpyxl.chart import BarChart, Reference, Series, PieChart
from openpyxl.styles import PatternFill, GradientFill, Color
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import DataPoint
# from openpyxl.styles.colors import ColorChoice
import multiprocessing

class NormalOutput:
    __slots__ = (
        '_output_path',
        '_myxl',
        '_sheet_name_list',
        '_row',
        '_column',
        '_mode',
    )
    _output_path: Path
    _myxl: Optional
    _sheet_name_list: List
    _row: int
    _column: int
    _mode: bool

    def __init__(self, path: Path, mode: bool):
        self._output_path = path
        self._mode = mode
        self._sheet_name_list = []
        self._row = 0
        self._column = 0

    # def __enter__(self):
    #     self._myxl = Excel(self._output_path, self._mode)
    #     return self

    # def __exit__(self, exc_type, exc_val, exc_tb):
    #     self._myxl.close()
    #     if len(self._sheet_name_list) == 0:
    #         os.remove(self._output_path)

    # def add_new_sheet(self, name: str):
    #     print(f"in NormalOutput")
    #     self._sheet_name_list.append(name)
    #     self._myxl.add_sheet(name)
    #     self._row = 0
    #     self._column = 0

    def get_sheet_id_by_case_name(self, name: str) -> int:
        id = -1
        for case_name in self._sheet_name_list:
            id += 1
            if case_name == name:
                return id
        return id

    # def _render_list(self, data: Optional, rank: int) -> (str, int, int):
    #     render_value = ""
    #     max_len = 10
    #     max_number_of_items = 0
        
    #     if isinstance(data, list):
    #         render_value += "["
    #         max_number_of_subitems = 0
    #         for subvalue in enumerate(data, start = 1):
    #             (subdata, length, number_of_subitems) = self._render_list(subvalue[1], rank+1)
    #             subdata = "None" if subdata == "" else subdata
    #             render_value += str(subdata)
    #             if subvalue[0] != len(data):
    #                 render_value += "\n"
    #             max_len = length if length > max_len else max_len
    #             max_number_of_subitems = number_of_subitems if number_of_subitems > max_number_of_subitems else max_number_of_subitems
    #             max_number_of_items += 1
    #         max_number_of_items = max_number_of_subitems if max_number_of_subitems > max_number_of_items else max_number_of_items
    #         # if rank == 0:
    #         #     render_value += "." * (subvalue[0] - 1)
    #         render_value += "]"
    #         return (render_value, max_len, max_number_of_items)
    #     else:
    #         max_len = len(data) if len(data) > max_len else max_len
    #         return (data, max_len, max_number_of_items)

    # def data_output(self, target_data: Dict, reference_data: Dict, matches: List) -> None:
    #     myxl = self._myxl
    #     column = 0
    #     base_info_flag = 0
    #     if matches == [0,0]:
    #         matches = [[0,0]]
    #         base_info_flag = 1
    #     # print(f"matches is {matches}")

    #     for match in enumerate(matches, start = 0):
    #         row = match[0]*3
    #         column = self._column
    #         # 
    #         for field1, field2 in zip(target_data.keys(), reference_data.keys()):
    #             # print(f"field is {field1}")
    #             option = 1
    #             max_str_len = len(field1)
    #             if match[0] == 0:
    #                 myxl.writetoxl(row,column,field1,0)
    #             # print(f"field1 is {field1}, tar data is {target_data[field1]}, ref data is {reference_data[field1]}, match is {match[1]}"  )
    #             if (not target_data[field1] == None) and (not base_info_flag):
    #                 value1 = target_data[field1][match[1][0]]
    #             else:
    #                 value1 = target_data[field1]
    #             # print(f"tar data is {target_data[field2]}, ref data is {reference_data[field2]}, match is {match[1]}"  )
    #             if (not reference_data[field2] == None) and (not base_info_flag):
    #                 value2 = reference_data[field2][match[1][1]]
    #             else:
    #                 value2 = reference_data[field2]

    #             if isinstance(value1, list):
    #                 render_value = ""
    #                 for subvalue in value1:
    #                     render_value += str(subvalue) + "\n"
    #                     max_str_len = len(str(subvalue)) if len(str(subvalue)) > max_str_len else max_str_len
    #                 myxl.writetoxl(row+1 ,column,render_value,option)
    #             else:
    #                 max_str_len = len(str(value1)) if len(str(value1)) > max_str_len else max_str_len
    #                 myxl.writetoxl(row+1,column,str(value1),option)
                
    #             option = 3
    #             if isinstance(value2, list):
    #                 render_value = ""
    #                 for subvalue in value2:
    #                     render_value += str(subvalue) + "\n"
    #                     max_str_len = len(str(subvalue)) if len(str(subvalue)) > max_str_len else max_str_len
    #                 myxl.writetoxl(row+2,column,render_value,option)
    #             else:
    #                 max_str_len = len(str(value2)) if len(str(value2)) > max_str_len else max_str_len
    #                 myxl.writetoxl(row+2,column,str(value2),option)
    #             if max_str_len > 10:
    #                 myxl.set_column(column,int(max_str_len+3))

    #             column += 1
    #     self._column = column

    # def data_output_single(self, target_data: Dict, cell_number:int = 0) -> None:
    #     myxl = self._myxl
    #     column = 0
    #     base_flag = 0 if cell_number == 0 else 1
    #     cell_number = 1 if cell_number == 0 else cell_number
    #     for id in range(cell_number):
    #         row = id
    #         column = self._column
    #         # 
    #         for field1 in target_data.keys():
    #             # print(f"field is {field1}")
    #             option = int(id % 2) + 5
    #             max_str_len = len(field1)
    #             if id == 0:
    #                 myxl.writetoxl(row,column,field1,0)
    #             # if 'case_name' in field1:
    #             #     print(f"field1 is {field1}, tar data is {target_data[field1]},  cell_number is {cell_number} id is {id}"  )
    #             if (not target_data[field1] == None) and (base_flag):
    #                 value1 = target_data[field1][id]
    #             else:
    #                 value1 = target_data[field1]

    #             if isinstance(value1, list):
    #                 render_value = ""
    #                 for subvalue in value1:
    #                     render_value += str(subvalue) + "\n"
    #                     max_str_len = len(str(subvalue)) if len(str(subvalue)) > max_str_len else max_str_len
    #                 myxl.writetoxl(row+1 ,column,render_value,option)
    #             else:
    #                 max_str_len = len(str(value1)) if len(str(value1)) > max_str_len else max_str_len
    #                 myxl.writetoxl(row+1,column,str(value1),option)
                
    #             if max_str_len > 10:
    #                 myxl.set_column(column,int(max_str_len+3))

    #             column += 1
    #     self._column = column

    # def update_summary_sheet(self, data):
    #     pass

    # def sorted_summary(self, field: str):
    #     pass

# header_in_find_list_out = ['_target_name','_function_name','_py_path_list','_x86_path_list','_hw_path_list','_team_info']
# class FindListOutput(NormalOutput):

#     def __enter__(self):
#         self._myxl = Excel(self._output_path, 1)
#         return self

#     def __exit__(self, exc_type, exc_val, exc_tb):
#         for id in range(len(header_in_find_list_out)):
#             self._myxl.writetoxl(0, id, header_in_find_list_out[id], 0)
#         self._myxl.close()

#     def _get_cloumn_info(self, attr: str) -> int:
#         return {
#             id for id in range(len(header_in_find_list_out)) if attr == header_in_find_list_out[id]
#         }

#     def _get_member_attr(slef, element: Optional) -> Optional:
#         return {
#             attr for attr in dir(element) if not attr.startswith('__')
#         }

#     def data_output(self, output_data: Optional) -> None:
#         max_str_lens = [10 for i in range(len(header_in_find_list_out))]

#         from collections import defaultdict

#         team_freqs = defaultdict(int)
#         site_freqs = {
#                       'WRO':0,
#                       'HZ':0,
#                       'SH':0,
#                       'Ulm':0,
#                       'StP':0,
#                     #   'ML':0,
#                       'Espoo':0,
#                     #   'ES':0,
#                       'LN':0,
#                     #   'LAN':0
#                        'others':0
#                     }

#         for element in enumerate(output_data, start = 1):
#             attrs = self._get_member_attr(element[1])
#             for attr in attrs:
#                 column = self._get_cloumn_info(str(attr)).pop()
#                 value = getattr(element[1], attr)
#                 if 'team_info' in attr:
#                     team_value = value[0].split(",")
#                     if isinstance(team_value, list):
#                         for subvalue in team_value:
#                             subvalue = re.sub(r'"|\n', "", subvalue)
#                             team_freqs[subvalue] += 1
#                     else:
#                         team_value = re.sub(r'"|\n', "", team_value)
#                         team_freqs[team_value] += 1

#                     find_key = 0
#                     for key in site_freqs.keys():
#                         # print(f'key is {key} and team_value is {team_value}')
#                         team_value = re.sub(r'LAN', 'LN', str(team_value))
#                         team_value = re.sub(r'ES', 'Espoo', str(team_value))
#                         team_value = re.sub(r'Shanghai', 'SH', str(team_value))
#                         if key in team_value:
#                             site_freqs[key] += 1
#                             find_key = 1
#                     if find_key == 0:
#                         site_freqs['others'] += 1
#                 (data, max_len, number_of_item) = self._render_list(value, 0)
#                 max_str_lens[column] = max_len if max_len > max_str_lens[column] else max_str_lens[column]
#                 self._myxl.writetoxl(element[0], column, str(data), bool(number_of_item > 1))
        
#         for max_str_len in enumerate(max_str_lens, start=0):
#             self._myxl.set_column(max_str_len[0], max_str_len[1] + 3)

#         team_freqs = dict(sorted(team_freqs.items(), key=lambda item: item[1], reverse=True))

#         print(f'start print team freqs {team_freqs}')
#         print(f'start print site freqs {site_freqs}')

#         self._myxl.paint_chart('Team_Freqs', 'Team_Name', 'Number', team_freqs, 'A2', len(header_in_find_list_out) + 1)
#         self._myxl.paint_chart('site_Freqs', 'Site_Name', 'Number', site_freqs, 'B2', len(header_in_find_list_out) + 3)

#     def data_output_additional(self, fields: List[str], data: Optional) -> None:
#         column = len(header_in_find_list_out)
#         for field, field_data in zip(fields, data):
#             self._myxl.writetoxl(0, column, field, 0)
#             row = 1
#             for sub_data in field_data:
#                 self._myxl.writetoxl(row, column, str(sub_data), 0)
#                 row += 1
#             column += 1


# class SimilarOutput(NormalOutput):
#     __slots__ = (
#         # '_output_path',
#         # '_myxl',
#         # '_sheet_name_list',
#         # '_row',
#         # '_column',
#         '_combine_list_output',
#         '_cover_list_output',
#         '_diff_channel_list_output',
#     )
#     # _output_path: Path
#     # _myxl: Optional
#     # _sheet_name_list: List
#     # _row: int
#     # _column: int
#     _combine_list_output : List
#     _cover_list_output : List
#     _diff_channel_list_output : List

#     def __init__(self, path: Path, mode: bool):
#         super().__init__(path, mode)
#         # self._output_path = path
        
#         # self._sheet_name_list = []
#         # self._row = 0
#         # self._column = 0
#         self._cover_list_output = []
#         self._combine_list_output = []
#         self._diff_channel_list_output = []

#     def __enter__(self):
#         # self._myxl = Excel(self._output_path)
#         super().__enter__()
#         self._myxl.add_sheet("summary_cover")
#         for id in range(len(header_in_find_list_out)):
#             self._myxl.writetoxl(0, id, header_in_find_list_out[id], 0)
#         self._myxl.add_sheet("summary_combine")
#         for id in range(len(header_in_find_list_out)):
#             self._myxl.writetoxl(0, id, header_in_find_list_out[id], 0)
#         self._myxl.add_sheet("summary_diff_channel_combine")
#         for id in range(len(header_in_find_list_out)):
#             self._myxl.writetoxl(0, id, header_in_find_list_out[id], 0)
#         # self._sheet.set_column('A:AL',40)

#         return self

#     def _get_cloumn_info(self, attr: str) -> int:
#             return {
#             id for id in range(len(header_in_find_list_out)) if attr == header_in_find_list_out[id]
#         }

#     def _get_member_attr(slef, element: Optional) -> Optional:
#         return {
#             attr for attr in dir(element) if not attr.startswith('__')
#         }

#     def _data_output_for_summary(self, data: Optional, sheetid : int) -> None:
#         max_str_lens = [10 for i in range(len(header_in_find_list_out))]
#         # print(f"start summary in sheet {sheetid}")
#         for element in enumerate(data, start = 1):
#             attrs = self._get_member_attr(element[1][0])
#             for attr in attrs:
#                 column = self._get_cloumn_info(str(attr)).pop()
#                 value = getattr(element[1][0], attr)
#                 (data, max_len, number_of_item) = self._render_list(value, 0)
#                 max_str_lens[column] = max_len if max_len > max_str_lens[column] else max_str_lens[column]
#                 if column == 0:
#                     detail_sheetid = element[1][1]
#                     # print(f"detail sheeid is {detail_sheetid}, element is {element[1][1]}")
#                     link = r'=HYPERLINK("#'+self._sheet_name_list[detail_sheetid]+'!d1' + '\", \"' + data + '")'
#                     if (int ((element[0] - 1) / 2 )) % 2 == 0:
#                         self._myxl.writetoxl_by_given_id(element[0], column, link, sheetid, 0)
#                     else:
#                         self._myxl.writetoxl_by_given_id(element[0], column, link, sheetid, 1)
#                 else:
#                     if (int ((element[0] - 1) / 2 )) % 2 == 0:
#                         self._myxl.writetoxl_by_given_id(element[0], column, str(data), sheetid, 2)
#                     else:
#                         self._myxl.writetoxl_by_given_id(element[0], column, str(data), sheetid, 3)
#         for max_str_len in enumerate(max_str_lens, start=0):
#             self._myxl.set_column_by_given_id(max_str_len[0], max_str_len[1] + 3, sheetid)

#     def __exit__(self, exc_type, exc_val, exc_tb):

#         # row = 0
#         # max_str_lens = [0 for i in range(6)]
#         # for combin_output in self._combine_list_output:
#         #     column = 0 
#         #     # print (f"details is {combin_output['details']} and self._sheet_name_list is {self._sheet_name_list}")

#         #     for field in combin_output:
#         #         if row == 0:
#         #             max_str_lens[column] = len(field) if max_str_lens[column] < len(field) else max_str_lens[column]
#         #             self._myxl.writetoxl_by_given_id(0, column, field, 0 , 1)
#         #         if field == 'details':
#         #             link = r'=HYPERLINK("#'+self._sheet_name_list[combin_output['details']]+'!d1' + '\", \"' + combin_output['case_name'] + '")'
#         #             value = link
#         #             self._myxl.writetoxl_by_given_id(row + 1, column, value, 0 , 0)
#         #         else:
#         #             value = combin_output[field]
#         #             if isinstance(value, list):
#         #                 render_value = ""
#         #                 str_len = 0
#         #                 for subvalue in value:
#         #                     render_value += str(subvalue) + "\n"
#         #                     str_len = len(str(subvalue)) if len(str(subvalue)) > str_len else str_len
#         #                 value = render_value
#         #             else:
#         #                 str_len = len(str(value))
#         #             if (field == 'x86_list' or field == 'hw_list') and len(combin_output[field]) > 1:
#         #                 self._myxl.writetoxl_by_given_id(row + 1, column, str(value), 0 , 2)
#         #             else:
#         #                 self._myxl.writetoxl_by_given_id(row + 1, column, str(value), 0 , 1)
#         #         max_str_lens[column] = str_len if max_str_lens[column] < str_len else max_str_lens[column]
#         #         column += 1
#         #     row += 1

#         # for max_str_len in enumerate(max_str_lens, start=0):
#         #     # print (f"max_str_len is {max_str_len}")
#         #     self._myxl.set_column_by_given_id(max_str_len[0],max_str_len[1]+10,0)
#         # self._myxl.set_column_by_given_id(5,max_str_lens[0]+10,0)

#         self._data_output_for_summary(self._cover_list_output, 0)
#         self._data_output_for_summary(self._combine_list_output, 1)
#         self._data_output_for_summary(self._diff_channel_list_output, 2)

#         self._myxl.close()
#         if len(self._sheet_name_list) == 0:
#             os.remove(self._output_path)

#     def update_summary_cover(self, data):
#         self._cover_list_output.append(data)

#     def update_summary_combine(self, data):
#         self._combine_list_output.append(data)

#     def update_summary_diff_channel(self, data):
#         self._diff_channel_list_output.append(data)

#     def sorted_summary(self, field):
#         self._combine_list_output = sorted(self._combine_list_output, key=lambda k: (k[field]), reverse=False)

#     # def add_new_sheet(self, name: str):
#     #     self._sheet_name_list.append(name)
#     #     self._myxl.add_sheet(name)
#     #     self._row = 0
#     #     self._column = 0

#     # def get_sheet_id_by_case_name(self, name: str) -> int:
#     #     id = -1
#     #     for case_name in self._sheet_name_list:
#     #         id += 1
#     #         if case_name == name:
#     #             return id
#     #     return id

#     # def data_output(self, target_data: Dict, compare_data: Dict, compare_result_list: List, comprared_id_list: List, compare_cell_number: int):
#     #     pass
#         # It needs to be designed by the customer
        
class PrintCaseInfoOutput(NormalOutput):
    __slots__ = (
        '_case_list_output',
    )
    _case_list_output : List

    def __init__(self, path: Path, mode: bool):
        super().__init__(path, mode)

        self._case_list_output = []

    def __enter__(self):

        super().__enter__()
        self._myxl.add_sheet("summary_case_basic_info")
        for id in range(len(header_in_find_list_out)):
            self._myxl.writetoxl(0, id, header_in_find_list_out[id], 0)

        return self

    def _get_cloumn_info(self, attr: str) -> int:
            return {
            id for id in range(len(header_in_find_list_out)) if attr == header_in_find_list_out[id]
        }

    def _get_member_attr(slef, element: Optional) -> Optional:
        return {
            attr for attr in dir(element) if not attr.startswith('__')
        }

    def _data_output_for_summary(self, data: Optional, sheetid : int) -> None:
        max_str_lens = [10 for i in range(len(header_in_find_list_out))]
        # print(f"start summary in sheet {sheetid}")
        for element in enumerate(data, start = 1):
            attrs = self._get_member_attr(element[1][0])
            for attr in attrs:
                column = self._get_cloumn_info(str(attr)).pop()
                value = getattr(element[1][0], attr)
                (data, max_len, number_of_item) = self._render_list(value, 0)
                max_str_lens[column] = max_len if max_len > max_str_lens[column] else max_str_lens[column]
                if column == 0:
                    detail_sheetid = element[1][1]
                    # print(f"detail sheeid is {detail_sheetid}, element is {element[1][1]}")
                    link = r'=HYPERLINK("#'+self._sheet_name_list[detail_sheetid]+'!d1' + '\", \"' + data + '")'
                    if (int ((element[0] - 1) / 1 )) % 2 == 0:
                        self._myxl.writetoxl_by_given_id(element[0], column, link, sheetid, 0)
                    else:
                        self._myxl.writetoxl_by_given_id(element[0], column, link, sheetid, 1)
                else:
                    if (int ((element[0] - 1) / 1 )) % 2 == 0:
                        self._myxl.writetoxl_by_given_id(element[0], column, str(data), sheetid, 2)
                    else:
                        self._myxl.writetoxl_by_given_id(element[0], column, str(data), sheetid, 3)
        for max_str_len in enumerate(max_str_lens, start=0):
            self._myxl.set_column_by_given_id(max_str_len[0], max_str_len[1] + 3, sheetid)

    def __exit__(self, exc_type, exc_val, exc_tb):

        self._data_output_for_summary(self._case_list_output, 0)

        self._myxl.close()
        if len(self._sheet_name_list) == 0:
            os.remove(self._output_path)

    def update_summary_case(self, data):
        self._case_list_output.append(data)
    
def get_max_len(column_data, column, max_column_len):
    for value in column_data:
        new_length = len(str(value))
        old_length = int(max_column_len)
        max_column_len = new_length if new_length > old_length else old_length
    return (column, max_column_len)
    

class FindListPandasOutput(NormalOutput):
    
    def __enter__(self):
        self._myxl = pd.ExcelWriter(self._output_path, engine='openpyxl')
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        # self._myxl.save()
        self._myxl.close()

    def add_new_sheet(self, name: str):
        # print(f"in FindListPandasOutput")
        self._sheet_name_list.append(name)

    def data_output(self, output_data: Optional, start_row: int, channels :List = []) -> None:

        init_len = [10 for i in range(len(output_data.columns))]
        # max_column_len = 0
        max_column_len = pd.Series(init_len, index=output_data.columns[:])
        # print(f"max_column_len is {max_column_len}")
        my_sheet_name=self._sheet_name_list[-1]
        # for rowIndex, row in output_data.iterrows():
        #     for columnIndex, value in row.items():
        #         if isinstance(value, list) and value and 'summary' in my_sheet_name:
        #             temp_value = ''
        #             for subvalue in value:
        #                 temp_value +=  str(subvalue) + '\n'
                    
        #             new_length = len(str(subvalue))
        #             output_data.loc[rowIndex, columnIndex] = temp_value
        #             # print(f"columnIndex is {columnIndex} max_column_len is {max_column_len[columnIndex]} ,new length is {new_length}")
        #         else:
        #             new_length = len(str(value))
        #         # print(f"columnIndex is {columnIndex} max_column_len is {max_column_len[columnIndex]}")
        #         old_length = int(max_column_len[columnIndex])
        #         max_column_len[columnIndex] = new_length if new_length > old_length else old_length
        num_processes = os.cpu_count()
        # print(f"The system has {num_processes} CPUs/cores in write excel.")
        if __name__ == "__main__":
            with multiprocessing.Pool(processes=num_processes) as pool:
                args = [(column_data, column, max_column_len[column]) for column_data, column in zip(output_data.values.T, output_data.columns)]
                results = pool.starmap(get_max_len, args)
                # for index, value in output_data[column].items():
                #     new_length = len(str(value))
                #     old_length = int(max_column_len[column])
                #     max_column_len[column] = new_length if new_length > old_length else old_length
            for column, max_len in results:
                max_column_len[column] = max_len

        output_data.to_excel(self._myxl, sheet_name=my_sheet_name, startrow=start_row)
        
        worksheet = self._myxl.book[my_sheet_name]

        painter_col = worksheet.max_column
        channel_painter = {}
        pie_col = worksheet.max_column
        # set a gradient fill color for the range of cells from column B to column E in row 1
        gradient_fill = GradientFill(type='linear', degree=0, stop=[Color('FFFF0000'), Color('FF800080')])
        gradient_fill_1 = GradientFill(type='linear', degree=0, stop=[Color('FF00FF00'), Color('FF0000FF')])
        red_fill = PatternFill(start_color='FF0000', end_color='FF2020', fill_type='lightTrellis')
        red_font = Font(color='FF0000')
        for id, column_cells in enumerate(worksheet.columns):
            column_letter = get_column_letter(id + 1)
            length = max_column_len[column_cells[start_row].value] + 3 if id > 0 else 10
            column = worksheet.column_dimensions[column_letter]
            column.width = length
            column.auto_size = True

            if "summary" in self._output_path and 'short' not in self._output_path and column_cells[0].value:
                if 'repeat_number' in column_cells[0].value:
                    painter_col = id+1
                    # print(f"column_cells[0].value is {column_cells[0].value} id is {id}")
                    for channel in channels:
                        if channel in column_cells[0].value:
                            channel_painter.update({channel: id+1})
                            break
                elif 'repeat_level' == column_cells[0].value:
                    pie_col = id+2
                    for row_id, row in enumerate(column_cells):
                        # print(f'row is {row.value}')
                        if 'over 90' in row.value:
                            # col_letter = worksheet.cell(row=row_id+1, column=id+1).column_letter
                            # print(f'row is {row.value}')
                            # worksheet.column_dimensions[col_letter].fill = red_fill
                            worksheet.cell(row=row_id+1, column=id+1).fill = gradient_fill
                            # worksheet.cell(row=row_id+1, column=id-15).fill = gradient_fill_1
                else:
                    pass
                
        # with multiprocessing.Pool(processes=24) as pool:
        #     # results = pool.map(align_data, [0,worksheet.columns, channels, worksheet, max_column_len, start_row, channel_painter])
        #     aligned_columns = pool.map(align_data, worksheet.columns)
        # worksheet.columns = aligned_columns
   
        if "summary" in self._output_path and 'short' not in self._output_path:
            chart = BarChart()
            # Fill the basic information like chart title,..
            chart.title = my_sheet_name + "RepeatView"
            chart.y_axis.title = 'Repeat Number'
            chart.x_axis.title = my_sheet_name
            # Now we will create a reference to the data and append the data to the chart.
            # data = Reference(worksheet, min_row=2, min_col=worksheet.max_column, max_col=1, max_row=2)
            # chart.add_data (data, from_rows = True, titles_from_data = True )

            # first_column = Reference(worksheet, min_col=1, min_row=2, max_col=1, max_row=worksheet.max_row)
            # painter_col1 = column_index_from_string("number")
            # print(f"number col is {painter_col1}")
            if channel_painter:
                # print(f"channel_painter is {channel_painter}")
                for channel in channels:
                    # # Define the data range for the series using the Reference class
                    # values = Reference(worksheet, min_col=channel_painter[channel], min_row=2, max_row=worksheet.max_row)
                    
                    # # Check if any values in the series are equal to zero
                    # if all(val == 0 for val in values):
                    #     continue  # Skip the series if all values are zero
                    
                    # # Create a new reference object excluding any values that are zero
                    # non_zero_values = [val if val != 0 else None for val in values]
                    # non_zero_range = Reference(worksheet, min_col=channel_painter[channel], min_row=2, max_row=worksheet.max_row, values=non_zero_values)
                    
                    # # Create a series object for the chart
                    # series = Series(non_zero_range, title=channel)

                    values = Reference(worksheet, min_col=channel_painter[channel], min_row=2, max_row=worksheet.max_row)
                    series = Series(values, title=channel)
                    chart.series.append(series)
            else:
                last_column = Reference(worksheet, min_col=painter_col, min_row=2, max_col=painter_col, max_row=worksheet.max_row)
                chart.add_data(last_column)
            # Finally, Add the chart to the sheet and save the file.
            worksheet.add_chart(chart,"A6")

            if "cell_user" in my_sheet_name:
                first_sheet_row = worksheet.max_row
                pie = PieChart()
                # Get unique age values
                df_unique = output_data.drop_duplicates(subset='repeat_level', keep='first')[['repeat_level', 'repeat_level_counts']]

                df_unique.to_excel(self._myxl, sheet_name=my_sheet_name, startrow=start_row + first_sheet_row, startcol= pie_col+2)
                # Create a list of data series
                # series = []
                # labels_list = []
                # data_list = []

                    # series.append(Series(values=[df_unique.iloc[i]['repeat_level_percentage']],
                    #                     xvalues=[df_unique.iloc[i]['repeat_level']]))

                labels = Reference(worksheet, min_col=pie_col+4, min_row=2+first_sheet_row, max_row=df_unique.shape[0]+1+first_sheet_row)
                data = Reference(worksheet, min_col=pie_col+5, min_row=1+first_sheet_row, max_row=df_unique.shape[0]+1+first_sheet_row)
                    # labels_list.append(labels)
                    # data_list.append(data)

                    # pie.append(Series(data, title= labels))

                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)
                pie.title = "repeat_level_percentage"

                # Define the colors for the gradient
                # colors = ['FF0000', 'FF00FF', '0000FF', '00FFFF']

                # # Assign each color to a different data point in the chart
                # # for idx, pt in enumerate(pie.series[0].points):
                # #     pt.format.fill.color.rgb = colors[idx]

                # # # create list of color stops for gradient
                # gradient_stops = [DataPoint(idx=i, fill=gradient_fill) for i in range(len(data) - 1)]

                # # # apply gradient to chart
                # pie.series[0].data_points = gradient_stops
                # pie.width = 10
                # pie.height = 5
                worksheet.add_chart(pie, "G6")


# def align_data(id, column_cells, channels, worksheet, max_column_len, start_row, channel_painter):
#     column_letter = get_column_letter(id + 1)
#     if column_cells[0].value and 'repeat_number' in column_cells[0].value:
#         print(f"column_cells[0].value is {column_cells[0].value}")
#         for channel in channels:
#             if channel in column_cells[0].value:
#                 channel_painter.update({channel: id+1})
#                 break

#     length = max_column_len[column_cells[start_row].value] + 3 if id > 0 else 10
#     column = worksheet.column_dimensions[column_letter]
#     column.width = length
#     column.auto_size = True

#     for cell in column_cells:
#         alignment_obj = cell.alignment.copy(horizontal='center', vertical='center', wrap_text=True)
#         cell.alignment = alignment_obj

def align_data(args):
    column_cells = args
    # column_letter = get_column_letter(id + 1)
    # if column_cells[0].value and 'repeat_number' in column_cells[0].value:
    #     print(f"column_cells[0].value is {column_cells[0].value} id is {id}")
        # for channel in channels:
        #     if channel in column_cells[0].value:
        #         channel_painter.update({channel: id+1})
        #         break

    # length = max_column_len[column_cells[start_row].value] + 3 if id > 0 else 10
    # column = worksheet.column_dimensions[column_letter]
    # column.width = length
    # column.auto_size = True
    aligned_cells = []
    for cell in column_cells:
        alignment_obj = cell.alignment.copy(horizontal='center', vertical='center', wrap_text=True)
        cell.alignment = alignment_obj
        aligned_cells.append(cell)

    # return aligned_cells
        

class PrintCaseInfoOutputByPanda(FindListPandasOutput):
    __slots__ = (
        '_case_list_output',
    )
    _case_list_output : pd.DataFrame

    def __init__(self, path: Path, mode: bool):
        super().__init__(path, mode)

    def __exit__(self, exc_type, exc_val, exc_tb):
        self._myxl.save()
        # self._myxl.close()
        if len(self._sheet_name_list) == 0:
            os.remove(self._output_path)

    def update_summary_case(self, data):
        if self._case_list_output.size:
            data.index = [r'=HYPERLINK("#'+"print_1"+'!A3' + '\", \"' + "HipTD_Thor_CAP_FDD_24LTE_10MHz_4RX_128PUSCH_UE_1200PRB_11000RRC.json" + '")']
            frames = [self._case_list_output, data]
            self._case_list_output = pd.concat(frames)
        else:
            
            self._case_list_output = data
