import os
import sys

import pandas as pd
import numpy as np
import uuid

import secrets
from datetime import datetime
import calendar
from openpyxl import load_workbook
from openpyxl.styles import *
from openpyxl.formatting.rule import *
from openpyxl.styles.differential import *
from openpyxl.utils.cell import *


from openpyxl.utils import range_boundaries
from fnschool import *
from fnschool.canteen import *
from fnschool.language import *
from fnschool.canteen.food import *
from fnschool.canteen.bill import *
from fnschool.canteen.path import *
from fnschool.canteen.config import *


class WorkBook:
    def __init__(self, bill):
        self.bill = bill
        self.check_sheet_name = "清点表"
        self.unit_sheet_name = "计量单位表"
        self.warehousing_sheet_name = "入库单"
        self.unwarehousing_sheet_name = "未入库明细表"
        self.consuming_sum_sheet_name = "出库汇总表"
        self.consuming_sheet_name = "出库单"
        self.inventory_sheet_name = "食材盘存表"
        self.food_sheet0_name = "材料台账母表"
        self.pre_consuming_sheet0_name = "简易出库母表"
        self.base_class_sheet_name = "大类表"
        self.pre_consuming_sheet_name_prefix = "出库表"
        self.purchase_sum_sheet_name = "入库、未入库汇总表"
        self.cover_sheet_name = "六大类总封面"
        self.pre_consuming_sheet0_name = "出库计划表"
        self._recounts = None
        self.food_name_col_names = ["商品名称", "食材名称", "商品名", "食材名"]
        self.purchase_sheet_names = [
            "食堂购入表",
            "客户商品销售报表",
            "客户送货明细报表",
        ]
        self.negligible_col_names = [
            "忽略",
            "不计",
            "非入库",
            "可忽略",
            "非盘点",
        ]
        self.residue_col_names = [
            "上季结余",
            "是剩余",
            "是结余",
            "上年结余",
            "剩余",
            "结余",
        ]
        self.org_col_names = ["客户名称"]
        self.total_price_col_names = ["金额", "折后金额", "总价", "折前金额"]
        self.count_col_names = [
            "数量",
            "记账数量",
            "订货数量",
            "订货总量",
            "订货总数量",
        ]
        self.unit_name_col_names = ["单位", "订货单位"]
        self.xdate_col_names = ["送货日期", "送货时间"]
        self.warehousing_form_index_offset = 0
        self.inventory_form_index_offset = 1
        self.bill_workbook = None
        self.pre_consuming_workbook0 = None
        self._main_spreadsheet_path = None
        self._check_df = None
        self._unit_name_list = None
        self._unit_df = None
        self._negligible_class_list = None
        self._base_class_df = None
        self.purchase_workbook_fdpath = None
        self.pre_consuming_sheet_col_index_offset = 5
        self.pre_consuming_sheet_row_index_offset = 3
        self.spreadsheet_ext_names = ["xlsx"]
        self.cell_alignment0 = Alignment(horizontal="center", vertical="center")
        self.cell_side0 = Side(border_style="thin")
        self.cell_border0 = Border(
            top=self.cell_side0,
            left=self.cell_side0,
            right=self.cell_side0,
            bottom=self.cell_side0,
        )
        self.purchase_sheets_properties = []
        self.excluded_purchase_sheets = []

    def get_recounts(self):
        if self._recounts is None:
            self._recounts = get_food_recounts_config()
        return self._recounts

    @property
    def profile(self):
        return self.bill.profile

    @property
    def food(self):
        return self.bill.food

    def get_pre_consuming_workbook0(self):
        if not self.pre_consuming_workbook0:
            self.pre_consuming_workbook0 = load_workbook(pre_consuming0_fpath)
        return self.pre_consuming_workbook0

    def get_pre_consuming_workbook_fpath(self, time_node=None):
        time_node = time_node or self.bill.get_time_node()
        t0, t1 = time_node
        ext = pre_consuming0_fpath.as_posix().split(".")[-1]
        consuming_fpath = (
            pre_consuming0_fpath.as_posix()[: -len(ext) - 1]
            + "."
            + t0.strftime("%Y%m%d")
            + t1.strftime("%Y%m%d")
            + "."
            + ext
        )
        consuming_fpath = (
            self.get_profile_copy_data_dpath()
            / consuming_fpath.split(os.sep)[-1]
        )
        return consuming_fpath

    def get_pre_consuming_workbook(self, time_node=None, new_foods=None):
        pre_consuming_fpath = self.get_pre_consuming_workbook_fpath()
        if not pre_consuming_fpath.exists():
            print_warning(_("Workbook {0} doesn't exist."))
            return None
        workbook = load_workbook(pre_consuming_fpath.as_posix())
        return workbook

    def get_pre_consuming_sheet_of_time_node(self, new_foods=None):
        pre_consuming_fpath = self.get_pre_consuming_workbook_fpath()
        if not pre_consuming_fpath.exists():
            shutil.copy(pre_consuming0_fpath, pre_consuming_fpath)
            print_info(
                _("Spreadsheet {0} was copied to {1} .").format(
                    pre_consuming0_fpath, pre_consuming_fpath
                )
            )
            print_warning(
                _(
                    "Please design the consumptions of spreadsheet '{0}' ."
                ).format(pre_consuming_fpath)
            )
            self.design_pre_consuming_sheet(new_foods=new_foods)
        wb = load_workbook(pre_consuming_fpath.as_posix())
        sheet = wb[self.pre_consuming_sheet0_name]
        return [wb, sheet]

    def get_conver_sheet(self):
        return self.get_bill_sheet(self.cover_sheet_name)

    def get_purchase_sum_sheet(self):
        return self.get_bill_sheet(self.purchase_sum_sheet_name)

    def get_consuming_sum_sheet(self):
        return self.get_bill_sheet(self.consuming_sum_sheet_name)

    def get_days_of_pre_consuming_sheet(self, name):
        t0, t1 = self.get_time_node_of_pre_consuming_sheet(name)
        return (t1 - t0).days + 1

    def get_time_node_of_pre_consuming_sheet(self, name):
        pcsheet = self.get_bill_sheet(name)
        t0 = pcsheet.cell(
            1, self.pre_consuming_sheet_col_index_offset
        ).value.split(".")
        t0 = datetime(int(t0[0]), int(t0[1]), int(t0[2]))
        time_node = [t0, None]
        for col_index in range(
            self.pre_consuming_sheet_col_index_offset, pcsheet.max_column
        ):
            cell_value = pcsheet.cell(1, col_index).value
            if not cell_value:
                t1 = pcsheet.cell(1, col_index - 1).value.split(".")
                t1 = datetime(int(t1[0]), int(t1[1]), int(t1[2]))
                time_node[1] = t1
                return time_node

        return None

    def get_base_class_names(self, include_negligible_name=False):
        df = self.get_base_class_df()
        names = df.index.tolist()
        return names[:-1]

    def get_base_class_names_include_negligible_name(self):
        return self.get_base_class_names(include_negligible_name=True)

    def get_base_class_df(self):
        return self.get_base_class_df_from_spreadsheet()

    def get_base_class_df_from_spreadsheet(self, sheet_name="大类表"):
        if not self._base_class_df is None:
            return self._base_class_df

        base_class_df = pd.read_excel(
            self.bill.workbook.get_main_spreadsheet_path(),
            sheet_name=sheet_name,
        )
        self._base_class_df = base_class_df.T
        return self._base_class_df

    def get_negligible_class_list(self):
        if not self._negligible_class_list is None:
            return self._negligible_class_list
        self._negligible_class_list = (
            self.get_base_class_df().loc["非入库类"].to_list()
        )
        return self._negligible_class_list

    def get_unit_name_list(self):
        if not self._unit_name_list is None:
            return self._unit_name_list
        unit_df = self.get_unit_df()
        self._unit_name_list = unit_df["Name"].tolist()
        return self._unit_name_list

    def get_sheet_names(self):
        wb = self.get_bill_workbook()
        return wb.sheetnames

    def includes_sheet(self, sheet):
        if isinstance(sheet, str):
            name = sheet
        else:
            name = sheet.title
        names = self.get_sheet_names()
        if name.endswith("*"):
            name = name.replace("*", "")
            return any(_name.startswith(name) for _name in names)
        return name in names

    def get_consuming_n_by_time_node(self, time_node):
        time_nodes = self.bill.get_time_nodes()
        time_nodes = [t for t in time_nodes if t[0].month == time_node.month]
        n_index = 0
        for t0, t1 in time_nodes:
            time_range = [
                t0 + timedelta(days=i) for i in range((t1 - t0).days + 1)
            ]
            if time_node in time_range:
                n_index += time_range.index(time_node) + 1
            else:
                n_index += len(time_range)
        return n_index

    def format_inventory_sheet(self):
        isheet = self.get_inventory_sheet()
        self.unmerge_cells_of_sheet(isheet)

        for row in isheet.iter_rows(
            min_row=1, max_row=isheet.max_row, min_col=1, max_col=9
        ):
            isheet.row_dimensions[row[0].row].height = 14.25

            if row[8].value and "原因" in str(row[8].value).replace(" ", ""):
                isheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row + 1,
                    start_column=9,
                    end_column=9,
                )

            if row[6].value and str(row[6].value).replace(" ", "") == "差额栏":
                isheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=7,
                    end_column=8,
                )

            if row[4].value and str(row[4].value).replace(" ", "") == "盘点栏":
                isheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=5,
                    end_column=6,
                )

            if row[2].value and str(row[2].value).replace(" ", "") == "账面栏":
                isheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=3,
                    end_column=4,
                )

            if row[0].value and row[0].value.replace(" ", "") == "食材名称":
                isheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row + 1,
                    start_column=1,
                    end_column=1,
                )

            if row[0].value and (
                "备注" in row[0].value.replace(" ", "")
                or "审核人" in row[0].value.replace(" ", "")
            ):
                isheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=9,
                )

            if row[0].value and row[0].value.replace(" ", "") == "食材盘存表":
                isheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=9,
                )
                isheet.row_dimensions[row[0].row].height = 22.5

                isheet.merge_cells(
                    start_row=row[0].row + 1,
                    end_row=row[0].row + 1,
                    start_column=1,
                    end_column=9,
                )

    def update_inventory_sheet(self):
        isheet = self.get_inventory_sheet()
        tnfoods = self.food.get_residue_foods(self.bill.month)
        form_indexes = self.get_inventory_form_indexes()

        for form_index_n in range(0, len(form_indexes)):
            form_index = form_indexes[form_index_n]
            form_index0, form_index1 = form_index
            food_index0 = form_index0 + 3
            food_index1 = form_index1 - 1
            for row in isheet.iter_rows(
                min_row=food_index0,
                max_row=food_index1,
                min_col=1,
                max_col=9,
            ):
                for cell in row:
                    cell.value = ""

        for i, (tn, _foods) in enumerate(tnfoods):
            form_indexes_n = i
            t0, t1 = tn
            form_index = form_indexes[form_indexes_n]
            form_i0, form_i1 = form_index
            fentry_i0 = form_i0 + 3
            fentry_i1 = form_i1 - 1

            self.unmerge_cells_of_sheet(isheet)

            isheet.cell(
                form_i0,
                1,
                f"     "
                + f"学校名称：{self.bill.profile.org_name}"
                + f"                "
                + f"{t1.year} 年 {t1.month} 月 {t1.day} 日"
                + f"              ",
            )

            for row in isheet.iter_rows(
                min_row=fentry_i0,
                max_row=fentry_i1,
                min_col=1,
                max_col=9,
            ):
                for cell in row:
                    cell.value = ""
                    cell.alignment = self.cell_alignment0
                    cell.border = self.cell_border0

            for findex, food in enumerate(_foods):
                row_index = fentry_i0 + findex
                if (
                    isheet.cell(row_index + 1, 1).value.replace(" ", "")
                    == "合计"
                ):
                    isheet.insert_rows(row_index + 1, 1)
                isheet.cell(row_index, 1, food.name)
                isheet.cell(row_index, 2, food.unit_name)
                isheet.cell(row_index, 3, food.get_remainder_by_time(tn[1]))
                isheet.cell(
                    row_index,
                    4,
                    food.get_remainder_by_time(tn[1]) * food.unit_price,
                )
                isheet.cell(row_index, 5, food.get_remainder_by_time(tn[1]))
                isheet.cell(
                    row_index,
                    6,
                    food.get_remainder_by_time(tn[1]) * food.unit_price,
                )

        self.format_inventory_sheet()

        wb = self.get_bill_workbook()
        wb.active = isheet
        print_info(_("Sheet '%s' was updated.") % (self.inventory_sheet_name))

    def update_check_sheet_by_time_node_m1(self):
        cksheet = self.get_check_sheet()
        rfoods = self.food.get_foods_from_pre_consuming_sheet_m1()
        time_node = self.bill.time_node
        t0, t1 = time_node
        rfoods = [f for f in rfoods if f.get_remainder() > 0.0]

        for food in rfoods:
            food_exists = False
            for row in cksheet.iter_rows(
                min_row=2, max_row=cksheet.max_row + 1, min_col=1, max_col=2
            ):
                if row[0].value == food.fid and row[1].value == t1.strftime(
                    "%Y%m%d"
                ):
                    food_exists = True
                    break
            if food_exists:
                continue
            cksheet.insert_rows(2, 1)
            cksheet.cell(2, 1, food.fid)
            cksheet.cell(2, 2, t1.strftime("%Y%m%d"))
            cksheet.cell(2, 3, food.name)
            cksheet.cell(2, 4, food.get_remainder())
            cksheet.cell(2, 5, food.get_remainder() * food.unit_price)
            cksheet.cell(2, 6, "Y")

        wb = self.get_bill_workbook()
        wb.active = cksheet
        print_info(_("Sheet '%s' was updated.") % (self.check_sheet_name))

    def unmerge_cells_of_sheet(self, sheet):
        if isinstance(sheet, str):
            sheet = self.get_bill_sheet(sheet)
        merged_ranges = list(sheet.merged_cells.ranges)
        for cell_group in merged_ranges:
            sheet.unmerge_cells(str(cell_group))

    def update_cover_sheet(self):
        time_nodes = self.bill.get_time_nodes()
        t0, t1 = time_nodes[-1]
        cvsheet = self.get_conver_sheet()
        cvsheet.cell(
            1,
            1,
            self.bill.profile.org_name
            + f"{t1.year}年{t1.month}月份食堂食品采购统计表",
        )
        foods = [
            f
            for f in self.food.get_foods()
            if (not f.is_residue and f.xdate.month == self.bill.month)
        ]
        wfoods = [f for f in foods if not f.is_negligible]
        uwfoods = [f for f in foods if f.is_negligible]
        total_price = 0.0
        for row in cvsheet.iter_rows(
            min_row=3, max_row=9, min_col=1, max_col=3
        ):
            class_name = row[0].value.replace(" ", "")
            _total_price = 0.0
            for f in foods:
                if f.class_name == class_name:
                    _total_price += f.count * f.unit_price
            cvsheet.cell(row[0].row, 2, _total_price)

            total_price += _total_price
        cvsheet.cell(10, 2, total_price)

        w_seasoning_total_price = sum(
            [f.count * f.unit_price for f in wfoods if ("调味" in f.class_name)]
        )
        unw_seasoning_total_price = sum(
            [
                f.count * f.unit_price
                for f in uwfoods
                if ("调味" in f.class_name)
            ]
        )

        cvsheet.cell(
            8,
            3,
            f"入库：{w_seasoning_total_price:.2f}元；"
            + f"未入库：{unw_seasoning_total_price:.2f}元",
        )

        if self.bill.is_changsheng and self.bill.is_xuelan:
            self.update_cover_sheet_for_cangsheng_xuelan(
                cvsheet, foods, wfoods, uwfoods, total_price
            )

        wb = self.get_bill_workbook()
        wb.active = cvsheet

        print_info(_("Sheet '%s' was updated.") % self.cover_sheet_name)

    def get_food_form_index(self, sheet):
        indexes = self.get_food_form_indexes(sheet)
        _index_range = indexes[self.bill.month - 1]
        return _index_range

    def get_food_form_indexes(self, sheet):
        indexes = []
        for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row, min_col=1, max_col=14
        ):
            if row[0].value and "材料名称" in str(row[0].value).replace(
                " ", ""
            ):
                indexes.append([row[0].row + 3, None])

            if row[2].value and "合计" in str(row[2].value).replace(" ", ""):
                indexes[-1][1] = row[0].row + 1
        return indexes

    def get_residual_foods_by_month_m1(self):
        time_nodes = self.bill.get_time_nodes()
        t0, t1 = time_nodes[-1]
        t1_mm1 = datetime(t1.year, t1.month, 1) + timedelta(days=-1)
        time_nodes_mm1 = [
            t
            for t in time_nodes
            if self.bill.times_are_same_year_month(t[0], t1_mm1)
        ]

        foods = self.food.get_food_list_from_check_sheet()

        if len(time_nodes_mm1) < 1:
            t1_mm1 = t0 + timedelta(days=-1)
        else:
            time_nodes_mm1 = sorted(time_nodes_mm1, key=lambda t: t[1])
            t1_mm1 = time_nodes_mm1[-1][1]

        foods = [f for f in foods if (f.is_residue and f.xdate == t1_mm1)]
        return foods

    def get_food_sheet(self, name):
        sheet = None
        _, t1 = self.bill.get_time_nodes()[-1]
        if self.includes_sheet(name):
            sheet = self.get_bill_sheet(name)
        else:
            wb = self.get_bill_workbook()
            sheet = wb.copy_worksheet(self.get_food_sheet0())
            sheet.title = name
            for row in sheet.iter_rows(
                min_row=1, max_row=sheet.max_row, min_col=1, max_col=14
            ):
                if row[0].value and "材料名称" in str(row[0].value).replace(
                    " ", ""
                ):
                    row[0].value = (
                        f"材料名称：{name}" + f"（{self.food.unit_name}）"
                    )
                    sheet.cell(row[0].row + 1, 1, f"{t1.year}年")
        return sheet

    def format_food_sheet(self, sheet):
        if isinstance(sheet, str):
            sheet = self.get_food_sheet(sheet)
        self.unmerge_cells_of_sheet(sheet)
        for row in sheet.iter_rows(
            min_row=1,
            max_row=sheet.max_row,
            min_col=1,
            max_col=14,
        ):
            sheet.row_dimensions[row[0].row].height = 15.75
            if row[0].value and "入库、出库台账" in str(row[0].value):
                sheet.row_dimensions[row[0].row].height = 27
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=13,
                )

            if row[0].value and "年" in str(row[0].value):
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=2,
                )

            if row[3].value and "入库" in str(row[3].value).replace(
                " ", ""
            ).replace("　", ""):
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=4,
                    end_column=6,
                )

            if row[6].value and "出库" in str(row[6].value).replace(
                " ", ""
            ).replace("　", ""):
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=7,
                    end_column=9,
                )

            if row[9].value and "库存" in str(row[9].value).replace(
                " ", ""
            ).replace("　", ""):
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=10,
                    end_column=12,
                )

            if row[12].value and "编号" in str(row[12].value).replace(" ", ""):
                sheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row + 1,
                    start_column=13,
                    end_column=13,
                )

    def update_food_sheets(self):
        time_nodes = sorted(
            [
                tn
                for tn in self.bill.get_time_nodes()
                if tn[1].month == self.bill.month
            ]
        )
        time_nodes_mm1 = sorted(
            [
                tn
                for tn in self.bill.get_time_nodes()
                if tn[1].month == self.bill.month - 1
            ]
        )
        t0, t1 = time_nodes[-1]
        cfoods = [
            f
            for f in self.food.get_foods()
            if (
                (
                    f.xdate.month == self.bill.month
                    or (
                        self.bill.month
                        in [d.month for d, c in f.consuming_list]
                    )
                )
                and not f.is_negligible
            )
        ]
        food_names = list(set([f.name for f in cfoods]))
        wb = self.get_bill_workbook()
        tn0_dm1 = (
            (time_nodes[0][0] + timedelta(days=-1))
            if len(time_nodes_mm1) < 1
            else (time_nodes_mm1[-1][1])
        )

        rfoods = [
            f
            for f in self.food.get_foods()
            if (
                f.get_remainder_by_time(tn0_dm1) > 0
                and not f.is_negligible
                and f.xdate.month < self.bill.month
            )
        ]

        food_names = list(set([f.name for f in rfoods] + food_names))

        sheet = None
        for food_name in food_names:
            sheet = self.get_food_sheet(food_name)
            form_index_range = self.get_food_form_index(sheet)
            index_start, index_end = form_index_range

            for row_index in range(index_start, index_end - 1):
                for col_index in range(1, 14):
                    sheet.cell(row_index, col_index).value = ""
            row_index = index_start
            col_index = 1

            _rfoods = [f for f in rfoods if f.name == food_name]
            _cfoods = [f for f in cfoods if f.name == food_name]

            self.unmerge_cells_of_sheet(sheet)

            sheet.cell(index_start - 2, 1, f"{t1.year}年")

            if len(_rfoods) > 0:
                for _row_index in range(
                    index_start, index_start + len(_rfoods)
                ):
                    food = _rfoods[_row_index - index_start]
                    sheet.cell(
                        _row_index,
                        3,
                        ("上年结转" if t1.month == 1 else "上月结转"),
                    )
                    sheet.cell(row_index, 10, food.count)
                    sheet.cell(row_index, 11, food.unit_price)
                    sheet.cell(row_index, 12, food.count * food.unit_price)
                    row_index += 1
            else:
                sheet.cell(
                    row_index,
                    3,
                    ("上年结转" if t1.month == 1 else "上月结转"),
                )

                row_index += 1

            _cdates = []
            for food in _cfoods:
                if len(food.consuming_list) > 0:
                    _cdates += [d for d, c in food.consuming_list]
                _cdates.append(food.xdate)
            _cdates = [d for d in _cdates if d.month == self.bill.month]
            _cdates = sorted(list(set(_cdates)))

            consuming_n = 1
            warehousing_n = 1
            for cdate in _cdates:
                for food in _cfoods:

                    if food.xdate == cdate:
                        sheet.cell(row_index, 1, cdate.month)
                        sheet.cell(row_index, 2, cdate.day)
                        sheet.cell(row_index, 4, food.count)
                        sheet.cell(row_index, 5, food.unit_price)
                        sheet.cell(row_index, 6, food.count * food.unit_price)
                        sheet.cell(row_index, 9, "")
                        sheet.cell(row_index, 10, food.count)
                        sheet.cell(row_index, 11, food.unit_price)
                        sheet.cell(row_index, 12, food.count * food.unit_price)
                        sheet.cell(
                            row_index,
                            13,
                            f"R{cdate.month:0>2}{warehousing_n:0>2}",
                        )
                        warehousing_n += 1

                        if "合计" in str(sheet.cell(row_index + 1, 3).value):
                            sheet.insert_rows(row_index + 1, 1)

                        row_index += 1

                    if cdate in [d for d, __ in food.consuming_list]:
                        _count = [
                            c for d, c in food.consuming_list if d == cdate
                        ][0]
                        _remainder = food.count - sum(
                            [c for d, c in food.consuming_list if d <= cdate]
                        )
                        sheet.cell(row_index, 1, cdate.month)
                        sheet.cell(row_index, 2, cdate.day)
                        sheet.cell(row_index, 6, "")
                        sheet.cell(row_index, 7, _count)
                        sheet.cell(row_index, 8, food.unit_price)
                        sheet.cell(row_index, 9, _count * food.unit_price)
                        sheet.cell(row_index, 10, _remainder)
                        sheet.cell(row_index, 11, food.unit_price)
                        sheet.cell(row_index, 12, _remainder * food.unit_price)
                        sheet.cell(
                            row_index,
                            13,
                            f"C{cdate.month:0>2}{consuming_n:0>2}",
                        )
                        consuming_n += 1

                        if "合计" in str(sheet.cell(row_index + 1, 3).value):
                            sheet.insert_rows(row_index + 1, 1)

                        row_index += 1

            self.format_food_sheet(sheet)
            print_info(_("Sheet '%s' was updated.") % sheet.title)

        wb.active = sheet

        _food_names = list(set([f.name for f in self.food.get_foods()]))
        for name in _food_names:
            if self.includes_sheet(name):
                sheet = self.get_bill_sheet(name)
                sheet.sheet_properties.tabColor = "0" * 8

        print_info(_("All food sheets have their tab colors reset."))

        for name in food_names:
            sheet = self.get_food_sheet(name)
            sheet.sheet_properties.tabColor = secrets.token_hex(4)

        print_info(
            _("Food sheets [{0}] have their tab colors recolor.").format(
                " ".join(food_names)
            )
        )
        print_info(
            _("Food sheets [{0}] have been updated.").format(
                " ".join(food_names)
            )
        )

    def get_food_form_day_index(self, sheet):
        indexes = []
        for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row, min_col=1, max_col=14
        ):
            if (
                row[2].value
                and "结转" in str(row[2].value).replace(" ", "")
                and not sheet.cell(row[0].row + 1, 3).value
            ):
                indexes.append([row[0].row + 1, None])

            if row[2].value and "合计" in str(row[2].value).replace(" ", ""):
                indexes[-1][1] = row[0].row - 1
        index_range = indexes[self.bill.month - 1]
        return index_range

    def update_cover_sheet_for_cangsheng_xuelan(
        self, cvsheet, foods, wfoods, uwfoods, total_price
    ):
        egg_milk_total_price = sum(
            [f.count * f.unit_price for f in foods if "蛋奶" in f.class_name]
        )
        xl_milk_total_price = sum(
            [f.count * f.unit_price for f in foods if "雪兰" in f.name]
        )

        cvsheet.cell(
            5,
            3,
            f"昌盛：{(egg_milk_total_price - xl_milk_total_price):.2f}元；"
            + f"雪兰：{xl_milk_total_price:.2f}元",
        )

        cvsheet.cell(
            10,
            3,
            f"昌盛：{total_price-xl_milk_total_price:.2f}元；"
            + f"雪兰：{xl_milk_total_price:.2f}元",
        )

    def get_changsheng_purchase_properties(self, fpath):
        if not fpath.split(".")[-1] in self.spreadsheet_ext_names:
            return None
        chwb_fpath = fpath
        ef = pd.ExcelFile(chwb_fpath)
        sheet_names = ef.sheet_names
        ef.close()
        for sheet_name in sheet_names:
            if sheet_name in self.purchase_sheet_names:
                chwb0 = load_workbook(chwb_fpath, read_only=True)
                cssheet0 = chwb0[sheet_name]
                cs_dates = []

                header_names = [
                    str(cssheet0.cell(1, ci).value)
                    for ci in range(1, cssheet0.max_column + 1)
                    if cssheet0.cell(1, ci).value
                ]

                xdate_col_index = 0
                org_name_col_index = 0
                for hn in header_names:
                    if hn in self.xdate_col_names:
                        xdate_col_index = header_names.index(hn) + 1
                    elif hn in self.org_col_names:
                        org_name_col_index = header_names.index(hn) + 1
                if xdate_col_index == 0:
                    print_error(
                        _("Got no check date column index of {0} .").format(
                            chwb_fpath + "-->" + sheet_name
                        )
                    )
                if org_name_col_index == 0:
                    print_error(
                        _("Got no organization name column of {0}").format(
                            chwb_fpath + "-->" + sheet_name
                        )
                    )

                org_names = [
                    str(cssheet0.cell(ri, org_name_col_index).value)
                    for ri in range(1, cssheet0.max_row)
                ]
                if not self.bill.profile.org_name in org_names:
                    chwb0.close()
                    return None
                cs_dates = [
                    str(cssheet0.cell(ri, xdate_col_index).value)
                    for ri in range(1, cssheet0.max_row)
                ]
                cs_dates = sorted(
                    list(
                        set(
                            [
                                datetime.strptime(d, "%Y-%m-%d")
                                for d in cs_dates
                                if re.search(r"\d{4}-\d{2}-\d{2}", d)
                            ]
                        )
                    )
                )
                chwb0.close()
                return (sheet_name, cs_dates)

        return None

    def get_changsheng_sheet_name(self, wb):
        sheet_names = []
        if isinstance(wb, str):
            ef = pd.ExcelFile(wb)
            sheet_names = ef.sheet_names
            ef.close()
        else:
            sheet_names = wb.sheetnames
        for sn in self.purchase_sheet_names:
            if sn in sheet_names:
                return sn
        return None

    def update_sheets(self):
        foods = self.food.get_foods()
        self.update_inventory_sheet()
        self.update_consuming_sheet()
        self.update_warehousing_sheet()
        self.update_unwarehousing_sheet()
        self.update_consuming_sum_sheet()
        self.update_purchase_sum_sheet()
        self.update_cover_sheet()
        self.update_food_sheets()
        self.save_updated_workbooks()
        print_info(_("Update completely!"))

    def save_updated_workbooks(self):
        spreadsheet0_fpath = self.get_main_spreadsheet_path()
        random_spreadsheet_fpath = self.get_random_bill_workbook_copy_fpath()
        print_info(
            _(
                "Do you want to save all updated data "
                + 'to "{0}"? or just save it as a '
                + 'copy to "{1}". (YyNn)'
            ).format(spreadsheet0_fpath, random_spreadsheet_fpath)
        )
        print_warning(
            _(
                'If you save updated data to "{0}", '
                + "data of food sheets will be saved "
                + "for every month."
            ).format(spreadsheet0_fpath)
        )
        _input = input0()
        if len(_input) > 0 and _input in "Yy":
            self.save_bill_workbook(wb_fpath=spreadsheet0_fpath)
            print_info(
                _(
                    "You can fill in the monthly missing data "
                    + "to food sheets, they will be saved "
                    + "for next updating."
                )
            )
        else:
            self.copy_bill_workbook(wb_fpath=random_spreadsheet_fpath)
            spreadsheet0_fpath = random_spreadsheet_fpath

        open_path(spreadsheet0_fpath)
        print_info(_("Updated data was saved."))

    def get_changsheng_properties_by_dir(self, fdpath=None):
        fd_path = self.purchase_workbook_fd_path or fdpath
        properties = []
        if not Path(fd_path).is_dir():
            return None

        for _file in os.listdir(fd_path):
            if _file.split(".")[-1] in self.spreadsheet_ext_names:
                chwb_fpath = (Path(fd_path) / _file).as_posix()
                sheet_name = None
                ptime = None
                if chwb_fpath in [
                    p[0] for p in self.purchase_sheets_properties
                ]:
                    pinfo = [
                        p
                        for p in self.purchase_sheets_properties
                        if p[0] == chwb_fpath
                    ][0]
                    sheet_name, ptimes = pinfo[1:]
                elif chwb_fpath in self.excluded_purchase_sheets:
                    continue
                else:
                    print_info(_("Spreadsheet %s is being tested.") % _file)
                    pinfo = self.get_changsheng_purchase_properties(chwb_fpath)
                    if not pinfo:
                        self.excluded_purchase_sheets.append(chwb_fpath)
                        continue
                    sheet_name, ptimes = pinfo
                    if not chwb_fpath in [
                        p[0] for p in self.purchase_sheets_properties
                    ]:
                        self.purchase_sheets_properties.append(
                            [chwb_fpath, sheet_name, ptimes]
                        )

                if ptimes:
                    print_info(
                        _(
                            "The food purchasing times of preadsheet {0} is {1} ."
                        ).format(
                            _file,
                            " | ".join(
                                [ptime.strftime("%Y.%m.%d") for ptime in ptimes]
                            ),
                        )
                    )
                    properties.append([_file, sheet_name, chwb_fpath, ptimes])

        return properties if properties else None

    def get_pre_consuming_workbook_fpaths(self):
        fpaths = []
        for time_node in self.bill.get_time_nodes():
            t0, t1 = time_node
            fpath = Path(self.get_profile_copy_data_dpath()) / (
                "consuming-"
                + t0.strftime("%Y%m%d")
                + "-"
                + t1.strftime("%Y%m%d")
                + "."
                + self.spreadsheet_ext_names[0]
            )

            if not fpath.exists():
                shutil.copy(pre_consuming0_fpath, fpath)
                print(
                    _(
                        "Workbook '{0}' doesn't exist, "
                        + "workbook '{1}' was copied to '{2}' ."
                    ).format(fpath, pre_consuming0_fpath, fpath)
                )
            fpaths.append([time_node, fpath])

        return fpaths

    def get_changsheng_col_indexes(self, sheet):
        workbook_fpath, sheet = sheet
        food_name_index = [_("Food name index"), -1]
        food_count_index = [_("Food count index"), -1]
        food_total_price_index = [_("Food total price index"), -1]
        food_unit_index = [_("Food unit index"), -1]
        food_xdate_index = [_("Food check date index"), -1]
        food_neglect_mark_index = [
            _("Food 'negligible' mark index"),
            -1,
        ]
        food_residue_mark_index = [_("Food 'residue' mark index"), -1]
        food_org_name_index = [_("Food purchaser name index"), -1]

        header_names = [
            str(sheet.cell(1, ci).value)
            for ci in range(1, sheet.max_column + 1)
        ]

        for _col_index, cell_value in enumerate(header_names):
            if cell_value in ["商品名称"]:
                food_name_index[1] = _col_index
                pass
            elif cell_value in self.unit_name_col_names:
                food_unit_index[1] = _col_index
                pass
            elif cell_value in self.count_col_names:
                food_count_index[1] = _col_index
                pass
            elif cell_value in self.total_price_col_names:
                food_total_price_index[1] = _col_index
                pass
            elif cell_value in self.xdate_col_names:
                food_xdate_index[1] = _col_index
                pass
            elif cell_value in self.negligible_col_names:
                food_neglect_mark_index[1] = _col_index
                pass
            elif cell_value in self.residue_col_names:
                food_residue_mark_index[1] = _col_index
                pass
            elif cell_value in self.org_col_names:
                food_org_name_index[1] = _col_index
                pass

        indexes = self.clean_supplier_col_indexes(
            workbook_fpath,
            [
                food_name_index,
                food_count_index,
                food_total_price_index,
                food_unit_index,
                food_xdate_index,
                food_neglect_mark_index,
                food_org_name_index,
            ],
        ) + [food_residue_mark_index[1]]
        return indexes

    def read_consumptions_from_pre_consuming_workbooks(self):
        foods = self.bill._foods
        col_index_offset = self.pre_consuming_sheet_col_index_offset
        row_index_offset = self.pre_consuming_sheet_row_index_offset
        for time_node, fpath in self.get_pre_consuming_workbook_fpaths():
            t0, t1 = time_node
            wb = load_workbook(fpath)
            sheet = wb[self.pre_consuming_sheet0_name]
            ckt0, ckt1 = self.bill.get_check_times(time_node)
            _foods = [
                f
                for f in foods
                if (f.xdate <= ckt1 and f.remainder > 0 and not f.is_negligible)
            ]
            if len(_foods) < 1:
                print_warning(
                    _(
                        "There is not food of time node %s ,skip workbook designing."
                    )
                    % (t0.strftime("%Y.%m.%d") + t1.strftime("%Y.%m.%d"))
                )
                continue

            if not sheet.cell(1, col_index_offset).value:
                for i, t in enumerate(
                    [
                        t0 + timedelta(days=a)
                        for a in range(0, (t1 - t0).days + 1)
                    ]
                ):
                    sheet.cell(1, col_index_offset + i, t.strftime("%Y.%m.%d"))
                for i, _f in enumerate(_foods):
                    row_index = i + row_index_offset
                    sheet.cell(
                        row_index,
                        1,
                        _f.name + (_f.residue_mark if _f.xdate < ckt0 else ""),
                    )
                    sheet.cell(row_index, 2, _f.remainder)
                    sheet.cell(row_index, 4, _f.unit_price)
                wb.save(fpath)
                wb.close()
                print_info(
                    _(
                        "Workbook '{0}' was updated, please design the "
                        + "daily foods consumption and press ANY key "
                        + "to continue."
                    ).format(fpath)
                )
                open_path(fpath)
                input0()
                wb = load_workbook(fpath)
                sheet = wb[self.pre_consuming_sheet0_name]
                print_info(_("Workbook %s was read.") % fpath)

            for i, _f in enumerate(_foods):
                row_index = row_index_offset + i
                for col_index in range(col_index_offset, sheet.max_column + 1):
                    cell_value = sheet.cell(row_index, col_index).value
                    if cell_value:
                        _date = datetime.strptime(
                            sheet.cell(1, col_index).value, "%Y.%m.%d"
                        )
                        _count = float(cell_value)
                        _f.consume(_date, _count)

            print_info(_("Read consumption from '{0}' .").format(fpath))

    def clean_supplier_col_indexes(self, workbook_fpath, indexes):
        global _
        for i, [name, cn_index] in enumerate(indexes):
            if cn_index < 0:
                error_msg = _(
                    "Unable to find {0} from {1}, "
                    + "You can input it (1 base) directly or "
                    + "give feedback to the maintainers "
                    + "--> {2} ."
                ).format(name, workbook_fpath, get_new_issue_url())
                print_error(error_msg)
                for __ in range(3):
                    cn_index = input0()
                    if cn_index.isnumeric():
                        cn_index = int(cn_index) - 1
                        indexes[i] = [name, cn_index]
                        break
                    else:
                        print("Unexpected value was got.")
        indexes = [i for __, i in indexes]
        return indexes

    def clean_quotation_marks(self, value):
        value = value.replace("‘", "").replace("’", "").replace("'", "")
        return value

    def read_changsheng_foods(self, dpath=None):
        global Food
        dpath = self.purchase_workbook_fdpath or dpath
        dpath0 = (Path.home() / "Downloads").as_posix()
        if not dpath:
            print_info(
                _(
                    "Please enter the 'purchase list file path' of "
                    + "spreadsheet Changsheng provided, "
                    + "or enter the directory path and then {app_name} will "
                    + "read all spreadsheets."
                    + " (default: '{dpath0}')"
                ).format(app_name=app_name, dpath0=dpath0)
            )
            dpath = input0()

        if dpath.replace(" ", "") == "":
            dpath = dpath0
        if dpath.startswith("~"):
            dpath = Path.home().as_posix() + dpath[1:]
        if not Path(dpath).exists():
            print_error(_("File or directory '%s' doesn't exist.") % (dpath))
            return None

            self.purchase_workbook_fdpath = dpath
        print_info(_("Entered directory: %s") % dpath)

        for file in os.listdir(dpath):
            if file.split(".")[-1] in self.spreadsheet_ext_names:
                wb_fpath = (Path(dpath) / file).as_posix()
                if wb_fpath in self.excluded_purchase_sheets:
                    continue
                wb = load_workbook(wb_fpath, read_only=True)
                sheetnames = [
                    n for n in wb.sheetnames if n in self.purchase_sheet_names
                ]
                if len(sheetnames) < 1:
                    if not wb_fpath in self.excluded_purchase_sheets:
                        self.excluded_purchase_sheets.append(wb_fpath)
                    continue
                sheet = wb[sheetnames[0]]

                header_row_index = 1
                header_names = [
                    str(sheet.cell(header_row_index, ci).value)
                    for ci in range(1, sheet.max_column + 1)
                ]

                _org_name_col_index = -1

                for hn in header_names:
                    if hn in self.org_col_names:
                        _org_name_col_index = header_names.index(hn)

                if _org_name_col_index < 0:
                    print_warning(
                        _(
                            "{app_name} desn't pick the index of organization name"
                            + " column, input it (0 base) or '{wb_fpath}' will be ignored."
                        ).format(app_name=app_name, wb_fpath=wb_fpath)
                    )
                    open_path(wb_fpath)
                    _org_name_col_index = input0()
                    if not _org_name_col_index.isnumeric():
                        continue
                    _org_name_col_index = int(_org_name_col_index)

                _org_names = list(
                    set(
                        [
                            str(sheet.cell(ri, _org_name_col_index + 1).value)
                            for ri in range(
                                header_row_index + 1, sheet.max_row + 1
                            )
                            if sheet.cell(ri, _org_name_col_index + 1).value
                        ]
                    )
                )

                if not any(
                    [o == self.bill.profile.org_name for o in _org_names]
                ):
                    print_warning(
                        (
                            _('Organization name read from {0} are "{1}",')
                            if len(_org_names) > 1
                            else _('Organization name read from {0} is "{1}", ')
                        ).format(wb_fpath, " | ".join(_org_names))
                        + _('but organization name of {0} is "{1}".').format(
                            f"{self.profile.label}({self.profile.name})",
                            self.profile.org_name,
                        )
                    )
                    print_error(
                        _("'{wb_fpath}' was discarded.").format(
                            wb_fpath=wb_fpath
                        )
                    )
                    if not wb_fpath in self.excluded_purchase_sheets:
                        self.excluded_purchase_sheets.append(wb_fpath)
                    continue

                if not any(
                    [(h in self.negligible_col_names) for h in header_names]
                ):
                    print_info(
                        _(
                            "It seems you didn't set the 'negligible' mark "
                            + "for workbook '{0}' , update this workbook "
                            + "and press ANY key to continue. (Add the "
                            + "'negligible' column name even though there "
                            + "is no negligible foods)"
                        ).format(wb_fpath)
                    )
                    print_info(
                        _(
                            "The column names of 'negligible' mark are following:"
                        )
                        + "\n\t"
                        + " | ".join(self.negligible_col_names)
                    )
                    open_path(wb_fpath)
                    input0()
                    wb = load_workbook(wb_fpath, read_only=True)
                    sheet = wb[sheetnames[0]]

                print_info(_("Spreadsheet '%s' was used."))

                (
                    food_name_index,
                    food_count_index,
                    food_total_price_index,
                    food_unit_index,
                    food_xdate_index,
                    food_neglect_mark_index,
                    food_org_name_index,
                    food_residue_mark_index,
                ) = self.get_changsheng_col_indexes([wb_fpath, sheet])

                row_index = 1
                col_index = 1
                for row in sheet.iter_rows(
                    min_row=2,
                    max_row=sheet.max_row,
                    min_col=1,
                    max_col=sheet.max_column,
                ):
                    if row[food_name_index].value:
                        xdate = self.clean_quotation_marks(
                            row[food_xdate_index].value
                        )

                        xdate = (
                            datetime.strptime(xdate, "%Y-%m-%d")
                            if "-" in xdate
                            else (
                                datetime.strptime(xdate, "%Y/%m/%d")
                                if "/" in xdate
                                else datetime.strptime(xdate, "%Y%d%m")
                            )
                        )
                        org_name = row[food_org_name_index].value

                        if org_name != self.bill.profile.org_name:
                            if not wb_fpath in self.excluded_purchase_sheets:
                                self.excluded_purchase_sheets.append(wb_fpath)
                            continue

                        name = row[food_name_index].value
                        count = row[food_count_index].value
                        if isinstance(count, str):
                            count = float(self.clean_quotation_marks(count))
                        unit = row[food_unit_index].value
                        total_price = row[food_total_price_index].value
                        if isinstance(total_price, str):
                            total_price = float(
                                self.clean_quotation_marks(total_price)
                            )

                        is_negligible = (
                            not row[food_neglect_mark_index].value is None
                            if food_neglect_mark_index > 0
                            else False
                        )
                        is_residue = (
                            not row[food_residue_mark_index].value is None
                            if food_residue_mark_index > 0
                            else False
                        )

                        count = self.food.clean_count(name, count, unit)
                        unit = self.food.clean_unit_name(name)

                        _food = Food(
                            self.bill,
                            name=name,
                            xdate=xdate,
                            count=count,
                            is_residue=is_residue,
                            total_price=total_price,
                            is_negligible=is_negligible,
                            unit_name=unit,
                        )

                        if not _food in self.bill._foods:
                            self.bill._foods.append(_food)

        return self.bill._foods

    def read_changsheng_foods_by_time_node(self, fd_path=None, time_node=None):
        global Food
        fd_path = self.purchase_workbook_fd_path or fd_path
        time_node = time_node or self.bill.time_node
        t0, t1 = time_node
        seeking_dpath0 = (Path.home() / "Downloads").as_posix()
        if not fd_path:
            print_info(
                _(
                    "Please enter the 'purchase list file path' of "
                    + "spreadsheet Changsheng provided, "
                    + "or enter the directory path and then {app_name} will "
                    + "read all spreadsheets."
                    + " (default: '{seeking_dpath0}')"
                ).format(app_name=app_name, seeking_dpath0=seeking_dpath0)
            )
            fd_path = input0()

        if fd_path.replace(" ", "") == "":
            fd_path = seeking_dpath0
            self.purchase_workbook_fd_path = fd_path
        if fd_path.startswith("~"):
            fd_path = Path.home().as_posix() + fd_path[1:]
        if not Path(fd_path).exists():
            print_error(_("File or directory '%s' doesn't exist.") % (fd_path))
            return None

        chwb = None
        cssheet = None
        ck_t0, ck_t1 = self.bill.get_check_times_of_time_node()
        if Path(fd_path).is_dir():
            print_info(_("Entered directory: %s") % fd_path)
            csproperties = self.get_changsheng_properties_by_dir(fd_path)
            if not csproperties:
                return None
            for csproperty in csproperties:
                file_name, sheet_name, file_path, purchase_times = csproperty
                for ptime in purchase_times:
                    if ck_t0 <= ptime <= ck_t1:
                        fd_path = file_path
                        cssheet = sheet_name
                        break
                if cssheet:
                    break
            if not cssheet:
                return None

        else:
            cssheet = self.get_changsheng_sheet_name(fd_path)

        chwb = load_workbook(fd_path, read_only=True)
        cssheet = chwb[cssheet]

        food_name_index = 0
        food_count_index = 0
        food_total_price_index = 0
        food_unit_index = 0
        food_xdate_index = 0
        food_neglect_mark_index = 0
        food_residue_mark_index = 0
        food_org_name_index = 0

        for _col_index, cell_value in enumerate(
            [
                str(cssheet.cell(1, ci).value)
                for ci in range(1, cssheet.max_column + 1)
            ]
        ):
            if cell_value in ["商品名称"]:
                food_name_index = _col_index
            elif cell_value in ["单位", "订货单位"]:
                food_unit_index = _col_index
            elif cell_value in ["数量", "记账数量"]:
                food_count_index = _col_index
            elif cell_value in ["金额", "折前金额"]:
                food_total_price_index = _col_index
            elif cell_value in self.xdate_col_names:
                food_xdate_index = _col_index
            elif cell_value in self.negligible_col_names:
                food_neglect_mark_index = _col_index
            elif cell_value in self.residue_col_names:
                food_residue_mark_index = _col_index
            elif cell_value in self.org_col_names:
                food_org_name_index = _col_index

        csfoods = []
        is_residue = False

        row_index = 1
        col_index = 1
        for row in cssheet.iter_rows(
            min_row=2,
            max_row=cssheet.max_row,
            min_col=1,
            max_col=cssheet.max_column,
        ):
            if row[food_name_index].value:
                xdate = row[food_xdate_index].value
                xdate = (
                    datetime.strptime(xdate, "%Y-%m-%d")
                    if "-" in xdate
                    else datetime.strptime(xdate, "%Y%d%m")
                )
                if not (ck_t0 <= xdate <= ck_t1):
                    continue
                org_name = row[food_org_name_index].value
                if org_name != self.bill.profile.org_name:
                    continue

                name = row[food_name_index].value
                count = row[food_count_index].value
                unit = row[food_unit_index].value
                total_price = row[food_total_price_index].value
                is_negligible = (
                    not row[food_neglect_mark_index].value is None
                    if food_neglect_mark_index > 0
                    else False
                )
                is_residue = (
                    not row[food_residue_mark_index].value is None
                    if food_residue_mark_index > 0
                    else False
                )

                csfoods.append(
                    Food(
                        self.bill,
                        name=name,
                        xdate=xdate,
                        count=count,
                        is_residue=is_residue,
                        total_price=total_price,
                        is_negligible=is_negligible,
                        unit_name=unit,
                    )
                )

        chwb.close()
        if len(csfoods) < 1:
            print_warning(
                _("Got no purchased foods of time node %s .").format(
                    t0.strftime("%Y.%m.%d") + "->" + t1.strftime("%Y.%m.%d")
                )
            )
        self.update_foods_consuming(csfoods)
        return csfoods

    def update_foods_consuming(self, foods):
        t0, t1 = self.bill.get_time_node()
        foods = [f for f in foods if not f.is_negligible]

        if foods is None or len(foods) < 1:
            print_warning(
                _(
                    "It seems the residue foods from last "
                    + "time node are going to be consumed. "
                    + "The new foods of this time node ({0})"
                    + "aren't purchased."
                ).format(
                    self.bill.time_node[0].strftime("%Y.%m.%d")
                    + "->"
                    + self.bill.time_node[1].strftime("%Y.%m.%d")
                )
            )

        residue_foods = self.food.get_residue_foods_of_time_node()
        if residue_foods:
            foods += residue_foods
        if foods is None:
            return
        wb, sheet = self.get_pre_consuming_sheet_of_time_node(foods)

        for i, f in enumerate(foods):
            r = i + self.pre_consuming_sheet_row_index_offset
            for c in range(
                self.pre_consuming_sheet_col_index_offset, sheet.max_column + 1
            ):
                cell_value = sheet.cell(r, c).value
                if cell_value:
                    cdate = datetime.strptime(
                        sheet.cell(1, c).value, "%Y.%m.%d"
                    )
                    ccount = float(cell_value)
                    f.consume(cdate, ccount)

    def get_inventory_form_index_of_time_node(self):
        indexes = self.get_inventory_form_indexes()
        tn_index = (
            self.bill.get_time_node_index() + self.inventory_form_index_offset
        )
        indexes_len = len(indexes)
        _index = None
        if indexes_len <= tn_index:
            return None
        _index = indexes[tn_index]
        return _index

    def update_inventory_sheet_of_time_node(self, fd_path=None, time_node=None):
        fd_path = self.purchase_workbook_fd_path or fd_path
        time_node = time_node or self.bill.time_node
        t0, t1 = time_node
        isheet = self.get_inventory_sheet()
        foods = self.food.time_node_residue_foods
        (
            iform_index0,
            iform_index1,
        ) = self.get_inventory_form_index_of_time_node()

        if not foods:
            print_warning(_("There is no residue foods."))
            return isheet

        for food in foods:
            isheet.cell(iform_index0, 1, name)
            isheet.cell(iform_index0, 2, self.food.unit_name)
            isheet.cell(iform_index0, 3, count)
            isheet.cell(iform_index0, 4, total_price)
            isheet.cell(iform_index0, 5, count)
            isheet.cell(iform_index0, 6, total_price)
            if (
                isheet.cell(iform_index0 + 1, 1).value
                and isheet.cell(iform_index0 + 1, 1).value.replace(" ", "")
                == "合计"
            ):
                isheet.insert_rows(iform_index0 + 1, 1)
                r_total_price += total_price
                iform_index0 += 1
                isheet.cell(isht_form_index1, 4, r_total_price)
                isheet.cell(isht_form_index1, 6, r_total_price)

        print_info(_("Sheet '%s' was updated.") % self.inventory_sheet_name)

        wb = self.get_bill_workbook()
        wb.active = isheet
        print_info(_("Sheet '%s' was updated.") % self.check_sheet_name)
        return isheet

    def update_purchase_sum_sheet(self):
        time_nodes = [
            tn
            for tn in self.bill.get_time_nodes()
            if tn[1].month == self.bill.month
        ]
        t0, t1 = time_nodes[-1]
        pssheet = self.get_purchase_sum_sheet()

        pssheet.cell(
            2,
            1,
            f"编制单位：{self.bill.profile.org_name}"
            + f"        "
            + f"单位：元"
            + f"         "
            + f"{t1.year}年{t1.month}月{t1.day}日",
        )
        pssheet.cell(
            20,
            1,
            f"编制单位：{self.bill.profile.org_name}"
            + f"        "
            + f"单位：元"
            + f"         "
            + f"{t1.year}年{t1.month}月{t1.day}日",
        )
        foods = [
            f
            for f in self.food.get_foods()
            if (not f.is_residue and f.xdate.month == self.bill.month)
        ]
        wfoods = [f for f in foods if not f.is_negligible]
        uwfoods = [f for f in foods if f.is_negligible]
        total_price = 0.0
        for row in pssheet.iter_rows(
            min_row=4, max_row=10, min_col=1, max_col=3
        ):
            class_name = row[0].value.replace(" ", "")
            _total_price = 0.0
            for food in wfoods:
                if food.class_name == class_name:
                    _total_price += food.count * food.unit_price
            pssheet.cell(row[0].row, 2, _total_price)
            total_price += _total_price
        total_price_cn = self.bill.convert_num_to_cnmoney_chars(total_price)
        pssheet.cell(
            11, 1, f"总金额（大写)：{total_price_cn}    ¥{total_price:.2f}"
        )
        pssheet.cell(12, 1, f"经办人：{self.bill.profile.name}  ")

        total_price = sum([f.count * f.unit_price for f in uwfoods])
        total_price_cn = self.bill.convert_num_to_cnmoney_chars(total_price)
        pssheet.cell(27, 2, total_price)
        pssheet.cell(
            29, 1, f"总金额（大写)：{total_price_cn}    ¥{total_price:.2f}"
        )

        pssheet.cell(30, 1, f"经办人：{self.bill.profile.name}  ")

        wb = self.get_bill_workbook()
        wb.active = pssheet

        print_info(_("Sheet '%s' was updated.") % self.purchase_sum_sheet_name)

    def update_consuming_sum_sheet(self):
        cssheet = self.get_consuming_sum_sheet()
        time_nodes = sorted(
            [
                tn
                for tn in self.bill.get_time_nodes()
                if tn[1].month == self.bill.month
            ]
        )
        t0, t1 = time_nodes[-1]
        foods = [
            f
            for f in self.food.get_foods()
            if (self.bill.month in [d.month for d, c in f.consuming_list])
        ]

        total_price = 0.0
        for row in cssheet.iter_rows(
            min_row=4, max_row=10, min_col=1, max_col=3
        ):
            class_name = row[0].value.replace(" ", "")
            _total_price = 0.0
            for food in foods:
                if food.class_name == class_name:
                    _total_price += sum(
                        [
                            _count * food.unit_price
                            for _date, _count in food.consuming_list
                            if _date.month == self.bill.month
                        ]
                    )
            total_price += _total_price
            cssheet.cell(row[0].row, 2, _total_price)
            cssheet.cell(row[0].row, 2).number_format = numbers.FORMAT_NUMBER_00

        total_price_cn = self.bill.convert_num_to_cnmoney_chars(total_price)
        cssheet.cell(
            2,
            1,
            f"编制单位：{self.bill.profile.org_name}       "
            + f"单位：元         "
            + f"{t1.year}年{t1.month}月{t1.day}日",
        )
        cssheet.cell(
            11,
            1,
            (f"总金额（大写)：{total_price_cn}    " + f"¥{total_price:.2f}"),
        )
        cssheet.cell(12, 1, f"经办人：{self.bill.profile.name}  ")

        wb = self.get_bill_workbook()
        wb.active = cssheet

        print_info(_("Sheet '%s' was updated.") % self.consuming_sum_sheet_name)

    def update_consuming_sheet(self):
        foods = self.food.get_foods()
        csheet = self.get_consuming_sheet()
        form_indexes = self.get_consuming_form_indexes()

        time_nodes = self.bill.get_time_nodes()
        days = []
        class_names = self.food.get_class_names()
        for t0, t1 in time_nodes:
            days += [
                t0 + timedelta(days=i) for i in range(0, (t1 - t0).days + 1)
            ]
        print_info(
            _("Consuming days:")
            + " "
            + " ".join([d.strftime("%Y.%m.%d") for d in days])
        )

        merged_ranges = list(csheet.merged_cells.ranges)
        for cell_group in merged_ranges:
            csheet.unmerge_cells(str(cell_group))

        max_day_index = 0
        for day_index in range(0, len(days)):
            max_day_index = day_index + 1
            day = days[day_index]
            form_index = form_indexes[day_index]
            form_index0, form_index1 = form_index
            food_index0 = form_index0 + 2
            food_index1 = form_index1 - 1
            food_index_len = food_index1 - food_index0 + 1
            tfoods = [
                food
                for food in foods
                if day in [_date for _date, _count in food.consuming_list]
            ]
            tfoods_classes = [f.class_name for f in tfoods]

            classes_without_food = [
                _name for _name in class_names if not _name in tfoods_classes
            ]

            tfoods_len = len(tfoods)
            consuming_n = day_index + 1
            csheet.cell(form_index0, 2, self.bill.profile.org_name)
            csheet.cell(
                form_index0,
                4,
                f"{day.year}年 {day.month} 月 {day.day} 日  " + f"单位：元",
            )

            csheet.cell(
                form_index0,
                7,
                f"编号：C{day.month:0>2}{consuming_n:0>2}",
            )

            csheet.cell(
                form_index1 + 1,
                1,
                (
                    "   "
                    + "审核人："
                    + "        "
                    + "经办人："
                    + "　    "
                    + "过称人："
                    + self.bill.profile.name
                    + "      "
                    + "仓管人："
                    + " 　"
                ),
            )

            row_difference = (
                tfoods_len + len(classes_without_food) - food_index_len
            )

            if row_difference > 0:
                csheet.insert_rows(food_index0 + 1, row_difference)
                form_indexes = self.get_consuming_form_indexes()
                form_index1 += row_difference
                food_index1 += row_difference
                row_difference = 0

            for row in csheet.iter_rows(
                min_row=food_index0,
                max_row=food_index1,
                min_col=1,
                max_col=8,
            ):
                for cell in row:
                    cell.value = ""
                    cell.alignment = self.cell_alignment0
                    cell.border = self.cell_border0

            fentry_index = food_index0

            for class_name in class_names:
                class_foods = [
                    food for food in tfoods if (food.class_name == class_name)
                ]

                fentry_index_start = fentry_index
                if len(class_foods) < 1:
                    csheet.cell(fentry_index_start, 1, class_name)
                    fentry_index = fentry_index_start + 1
                    continue

                class_consuming_count = 0.0
                for food in class_foods:
                    for _date, _count in food.consuming_list:
                        if _date == day:
                            class_consuming_count += _count * food.unit_price

                class_foods_len = len(class_foods)
                if class_name == class_names[0] and row_difference < 0:
                    class_foods_len += abs(row_difference) - len(
                        classes_without_food
                    )
                fentry_index_end = fentry_index_start - 1 + class_foods_len

                csheet.cell(fentry_index_start, 1, class_name)
                csheet.cell(fentry_index_start, 7, class_consuming_count)
                csheet.cell(fentry_index_start, 7).number_format = (
                    numbers.FORMAT_NUMBER_00
                )

                for findex, food in enumerate(class_foods):
                    consuming_count = [
                        _count
                        for _date, _count in food.consuming_list
                        if _date == day
                    ][0]
                    frow_index = fentry_index_start + findex
                    csheet.cell(frow_index, 2, food.name)
                    csheet.cell(frow_index, 3, food.unit_name)
                    csheet.cell(frow_index, 4, consuming_count)
                    csheet.cell(frow_index, 5, food.unit_price)
                    csheet.cell(
                        frow_index, 6, consuming_count * food.unit_price
                    )
                    csheet.cell(frow_index, 4).number_format = (
                        numbers.FORMAT_NUMBER
                    )
                    csheet.cell(frow_index, 5).number_format = (
                        numbers.FORMAT_NUMBER_00
                    )
                    csheet.cell(frow_index, 6).number_format = (
                        numbers.FORMAT_NUMBER_00
                    )

                fentry_index = fentry_index_end + 1

            tfoods_total_price = 0.0
            for food in tfoods:
                for _date, _count in food.consuming_list:
                    if _date == day:
                        tfoods_total_price += _count * food.unit_price
            csheet.cell(form_index1, 6, tfoods_total_price)
            csheet.cell(form_index1, 7, tfoods_total_price)

        if len(form_indexes) > max_day_index:
            for time_index in range(max_day_index, len(form_indexes)):
                form_index0, form_index1 = form_indexes[time_index]
                food_index0, food_index1 = (
                    form_index0 + 2,
                    form_index1 - 1,
                )
                for row in csheet.iter_rows(
                    min_row=food_index0,
                    max_row=food_index1,
                    min_col=2,
                    max_col=7,
                ):
                    for cell in row:
                        cell.value = ""

        self.format_consuming_sheet()

        wb = self.get_bill_workbook()
        wb.active = csheet
        print_info(_("Sheet '%s' was updated.") % self.consuming_sheet_name)

    def design_pre_consuming_sheet(self, new_foods=None):
        t0, t1 = self.bill.time_node
        wb_fpath = self.get_pre_consuming_workbook_fpath()
        wb = load_workbook(wb_fpath)
        sheet = wb[self.pre_consuming_sheet0_name]

        foods = new_foods
        foods = [f for f in foods if not f.is_negligible]
        row_index_offset = self.pre_consuming_sheet_row_index_offset
        col_index_offset = self.pre_consuming_sheet_col_index_offset
        rc_index = 0
        days_difference = (t1 - t0).days

        for day in range(0, days_difference + 1):
            time_header = (t0 + timedelta(days=day)).strftime("%Y.%m.%d")
            cell = sheet.cell(1, rc_index + col_index_offset)
            cell.value = time_header
            cell.number_format = numbers.FORMAT_TEXT
            rc_index += 1

        rc_index = 0
        for row in sheet.iter_rows(
            max_col=5,
            min_row=row_index_offset,
            max_row=row_index_offset + len(foods),
        ):
            if rc_index > len(foods) - 1:
                break
            food = foods[rc_index]
            row[0].value = food.get_name_with_residue_mark()
            row[1].value = food.count
            row[3].value = food.unit_price
            rc_index += 1

        wb.active = sheet
        wb.save(wb_fpath)
        wb.close()
        print_warning(
            _(
                "Sheet '{0}' was updated.\n"
                + "Press any key to continue when you have "
                + "completed the foods allocation."
            ).format(sheet.title)
        )
        print_info(
            _("Ok! I have updated spreadsheet '{0}'. (Press any key)").format(
                wb_fpath
            )
        )
        open_path(wb_fpath)
        input0()

    def set_warehousing_form_index_offset(self, offset=0):
        self.warehousing_form_index_offset = offset

    def set_inventory_form_index_offset(self, offset=0):
        self.inventory_form_index_offset = offset

    @property
    def food_list(self):
        return self.bill.get_food_list()

    def get_warehousing_sheet(self):
        return self.get_bill_sheet(self.warehousing_sheet_name)

    def get_unwarehousing_sheet(self):
        return self.get_bill_sheet(self.unwarehousing_sheet_name)

    def get_consuming_sheet(self):
        return self.get_bill_sheet(self.consuming_sheet_name)

    def get_inventory_sheet(self):
        return self.get_bill_sheet(self.inventory_sheet_name)

    def get_food_sheet0(self):
        return self.get_bill_sheet(self.food_sheet0_name)

    def get_consuming_form_indexes(self):
        csheet = self.get_consuming_sheet()
        indexes = []
        row_index = 1
        for row in csheet.iter_rows(max_row=csheet.max_row + 1, max_col=9):
            if row[0].value:
                if row[0].value.replace(" ", "") == "出库单":
                    indexes.append([row_index + 1, 0])
                if row[0].value.replace(" ", "") == "合计":
                    indexes[-1][1] = row_index
            row_index += 1

        if len(indexes) > 0:
            return indexes

        return none

    def get_unwarehousing_form_indexes(self):
        unwsheet = self.get_unwarehousing_sheet()
        indexes = []
        row_index = 1
        for row in unwsheet.iter_rows(max_row=unwsheet.max_row + 1, max_col=7):
            if row[0].value and "未入库明细表" in row[0].value.replace(" ", ""):
                indexes.append([row_index + 1, 0])

            if row[1].value and "合计" in row[1].value.replace(" ", ""):
                indexes[-1][1] = row_index

            row_index += 1

        if len(indexes) > 0:
            return indexes

        return None

    def get_inventory_form_indexes(self):
        isheet = self.get_inventory_sheet()
        indexes = []
        row_index = 1
        for row in isheet.iter_rows(max_row=isheet.max_row + 1, max_col=8):
            if row[0].value:
                if row[0].value.replace(" ", "") == "食材盘存表":
                    indexes.append([row_index + 1, 0])
                if row[0].value.replace(" ", "") == "合计":
                    indexes[-1][1] = row_index
            row_index += 1

        if len(indexes) > 0:
            return indexes

        return None

    def get_warehousing_form_indexes(self):
        wsheet = self.get_warehousing_sheet()
        indexes = []
        row_index = 1
        for row in wsheet.iter_rows(max_row=wsheet.max_row + 1, max_col=8):
            if row[0].value:
                if row[0].value.replace(" ", "") == "入库单":
                    indexes.append([row_index + 1, 0])
                if row[0].value.replace(" ", "") == "合计":
                    indexes[-1][1] = row_index
            row_index += 1

        if len(indexes) > 0:
            return indexes

        return None

    def format_consuming_sheet(self):
        csheet = self.get_consuming_sheet()
        merged_ranges = list(csheet.merged_cells.ranges)
        for cell_group in merged_ranges:
            csheet.unmerge_cells(str(cell_group))

        for row in csheet.iter_rows(
            min_row=1, max_row=csheet.max_row, min_col=1, max_col=8
        ):
            if row[0].value and row[0].value.replace(" ", "") == "出库单":
                csheet.row_dimensions[row[0].row].height = 21
                csheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=8,
                )
                csheet.merge_cells(
                    start_row=row[0].row + 1,
                    end_row=row[0].row + 1,
                    start_column=4,
                    end_column=6,
                )
                csheet.merge_cells(
                    start_row=row[0].row + 1,
                    end_row=row[0].row + 1,
                    start_column=7,
                    end_column=8,
                )

            if row[0].value and row[0].value.replace(" ", "").endswith("类"):
                row[6].number_format = numbers.FORMAT_NUMBER_00
                for _row in csheet.iter_rows(
                    min_row=row[0].row + 1,
                    max_row=csheet.max_row + 1,
                    min_col=1,
                    max_col=1,
                ):
                    if _row[0].value and (
                        _row[0].value.replace(" ", "").endswith("类")
                        or _row[0].value.replace(" ", "") == "合计"
                    ):
                        csheet.merge_cells(
                            start_row=row[0].row,
                            end_row=_row[0].row - 1,
                            start_column=1,
                            end_column=1,
                        )
                        csheet.merge_cells(
                            start_row=row[0].row,
                            end_row=_row[0].row - 1,
                            start_column=7,
                            end_column=7,
                        )
                        break

            if row[0].value and "审核人" in row[0].value.replace(" ", ""):
                csheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=8,
                )

        wb = self.get_bill_workbook()
        wb.active = csheet

        print_info(_("Sheet '%s' was formatted.") % self.consuming_sheet_name)

    def format_warehousing_sheet(self):
        wsheet = self.get_warehousing_sheet()
        merged_ranges = list(wsheet.merged_cells.ranges)
        for cell_group in merged_ranges:
            wsheet.unmerge_cells(str(cell_group))

        for row in wsheet.iter_rows(
            min_row=1, max_row=wsheet.max_row, min_col=1, max_col=8
        ):
            if row[0].value and row[0].value.replace(" ", "") == "入库单":
                wsheet.row_dimensions[row[0].row].height = 21
                wsheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=8,
                )
                wsheet.merge_cells(
                    start_row=row[0].row + 1,
                    end_row=row[0].row + 1,
                    start_column=4,
                    end_column=6,
                )
                wsheet.merge_cells(
                    start_row=row[0].row + 1,
                    end_row=row[0].row + 1,
                    start_column=7,
                    end_column=8,
                )

            if row[0].value and row[0].value.replace(" ", "").endswith("类"):
                row[6].number_format = numbers.FORMAT_NUMBER_00
                for _row in wsheet.iter_rows(
                    min_row=row[0].row + 1,
                    max_row=wsheet.max_row + 1,
                    min_col=1,
                    max_col=1,
                ):
                    if _row[0].value and (
                        _row[0].value.replace(" ", "").endswith("类")
                        or _row[0].value.replace(" ", "") == "合计"
                    ):
                        wsheet.merge_cells(
                            start_row=row[0].row,
                            end_row=_row[0].row - 1,
                            start_column=1,
                            end_column=1,
                        )
                        wsheet.merge_cells(
                            start_row=row[0].row,
                            end_row=_row[0].row - 1,
                            start_column=7,
                            end_column=7,
                        )
                        break

            if row[0].value and "审核人" in row[0].value.replace(" ", ""):
                wsheet.merge_cells(
                    start_row=row[0].row,
                    end_row=row[0].row,
                    start_column=1,
                    end_column=8,
                )

        wb = self.get_bill_workbook()
        wb.active = wsheet

        print_info(_("Sheet '%s' was formatted.") % self.warehousing_sheet_name)

    def update_unwarehousing_sheet(self):
        unwsheet = self.get_unwarehousing_sheet()
        form_indexes = self.get_unwarehousing_form_indexes()
        time_nodes = [
            tn
            for tn in self.bill.get_time_nodes()
            if tn[1].month == self.bill.month
        ]

        t0, t1 = time_nodes[-1]
        foods = [
            f
            for f in self.food.get_foods()
            if (f.is_negligible and f.xdate.month == self.bill.month)
        ]
        foods = sorted(foods, key=lambda f: f.xdate)
        row_indexes = []
        for form_index in form_indexes:
            form_index0, form_index1 = form_index
            unwsheet.cell(
                form_index0, 1, f" 学校名称：{self.bill.profile.org_name}"
            )
            unwsheet.cell(
                form_index0,
                4,
                f"        "
                + f"{t1.year} 年 {t1.month} 月 "
                + f"{t1.day} 日"
                + f"               ",
            )
            row_index_start = form_index0 + 2
            row_index_end = form_index1 - 1
            row_indexes += list(range(row_index_start, row_index_end + 1))

        for row_index in row_indexes:
            for col_index in range(1, 7 + 1):
                unwsheet.cell(row_index, col_index, "")

        total_price = 0.0
        use_forms = False

        for _index, row_index in enumerate(row_indexes):
            food = foods[_index]
            total_price += food.total_price
            unwsheet.cell(row_index, 1, food.xdate.strftime("%Y.%m.%d"))
            unwsheet.cell(row_index, 2, food.name)
            unwsheet.cell(row_index, 3, food.unit_name)
            unwsheet.cell(row_index, 4, food.count)
            unwsheet.cell(row_index, 5, food.unit_price)
            unwsheet.cell(row_index, 6, food.total_price)
            unwsheet.cell(row_index, 5).number_format = numbers.FORMAT_NUMBER_00
            unwsheet.cell(row_index, 6).number_format = numbers.FORMAT_NUMBER_00
            if (
                str(unwsheet.cell(row_index + 1, 2).value)
                .replace(" ", "")
                .endswith("合计")
                and len(foods) - 1 > _index
            ):
                unwsheet.cell(row_index + 1, 2, "合计")
                unwsheet.cell(row_index + 1, 6, total_price)
                use_forms = True

            if len(foods) - 1 == _index:
                for row in unwsheet.iter_rows(
                    min_row=row_index,
                    max_row=unwsheet.max_row,
                    min_col=1,
                    max_col=7,
                ):
                    if row[2].value and str(row[2].value).replace(
                        " ", ""
                    ).endswith("合计"):
                        row[2].value = "总合计" if use_forms else "合计"
                        row[6].value = total_price
                        break
                break

        print_info(_("Sheet '%s' was updated.") % self.unwarehousing_sheet_name)

    def update_warehousing_sheet(self):
        wsheet = self.get_warehousing_sheet()
        foods = [
            f
            for f in self.food.get_foods()
            if (
                not f.is_residue
                and not f.is_negligible
                and f.xdate.month == self.bill.month
            )
        ]
        form_indexes = self.get_warehousing_form_indexes()
        class_names = self.food.get_class_names()
        time_nodes = self.bill.get_time_nodes()

        self.unmerge_cells_of_sheet(wsheet)

        for form_index0, form_index1 in form_indexes:
            food_index0 = form_index0 + 2
            food_index1 = form_index1 - 1
            for row in wsheet.iter_rows(
                min_row=food_index0,
                max_row=food_index1,
                min_col=1,
                max_col=8,
            ):
                for cell in row:
                    cell.value = ""

        w_times = sorted(
            list(
                set(
                    [
                        food.xdate
                        for food in foods
                        if food.xdate.month == self.bill.month
                    ]
                )
            )
        )

        max_time_index = 0
        for windex, w_time in enumerate(w_times):
            time_point = w_time
            max_time_index = windex + 1
            form_index0, form_index1 = form_indexes[windex]
            food_index0 = form_index0 + 2
            food_index1 = form_index1 - 1
            entry_index = food_index0
            warehousing_n = windex + 1

            wfoods = [f for f in foods if (f.xdate == time_point)]
            foods_class_names = [f.class_name for f in wfoods]
            class_names_without_food = [
                _name for _name in class_names if not _name in foods_class_names
            ]
            row_difference = (
                len(wfoods)
                + len(class_names_without_food)
                - (food_index1 - food_index0 + 1)
            )

            if row_difference > 0:
                wsheet.insert_rows(food_index0 + 1, row_difference)
                for row in wsheet.iter_rows(
                    min_row=food_index0 + 1,
                    max_row=food_index0 + 1 + row_difference,
                    min_col=1,
                    max_col=8,
                ):
                    for cell in row:
                        cell.alignment = self.cell_alignment0
                        self.border = self.cell_border0

                form_indexes = self.get_warehousing_form_indexes()
                form_index1 += row_difference
                food_index1 = form_index1 - 1
                row_difference = 0

            for row in wsheet.iter_rows(
                min_row=food_index0,
                max_row=food_index1,
                min_col=1,
                max_col=8,
            ):
                for cell in row:
                    cell.value = ""
                    cell.alignment = self.cell_alignment0
                    cell.border = self.cell_border0

            wsheet.cell(form_index0, 2, self.bill.profile.org_name)
            wsheet.cell(
                form_index0,
                4,
                f"{time_point.year}年 {time_point.month} 月 "
                + f"{time_point.day} 日  单位：元",
            )
            wsheet.cell(
                form_index0,
                7,
                f"编号：R{time_point.month:0>2}{warehousing_n:0>2}",
            )

            wsheet.cell(
                form_index1 + 1,
                1,
                (
                    "   "
                    + "审核人："
                    + "        "
                    + "经办人："
                    + self.bill.profile.name
                    + " 　    "
                    + "过称人："
                    + "        "
                    + "仓管人："
                    + " 　"
                ),
            )

            for class_name in class_names:
                cfoods = [f for f in wfoods if f.class_name == class_name]
                cfoods_total_price = sum([f.total_price for f in cfoods])

                wsheet.cell(entry_index, 1, class_name)
                wsheet.cell(entry_index, 7, cfoods_total_price)

                if len(cfoods) < 1:
                    entry_index += 1
                    continue

                for cfindex, cfood in enumerate(cfoods):
                    cfood_row_index = entry_index + cfindex
                    wsheet.cell(cfood_row_index, 2, cfood.name)
                    wsheet.cell(cfood_row_index, 3, cfood.unit_name)
                    wsheet.cell(cfood_row_index, 4, cfood.count)
                    wsheet.cell(cfood_row_index, 5, cfood.unit_price)
                    wsheet.cell(cfood_row_index, 6, cfood.total_price)
                    wsheet.cell(cfood_row_index, 5).number_format = (
                        numbers.FORMAT_NUMBER_00
                    )
                    wsheet.cell(cfood_row_index, 6).number_format = (
                        numbers.FORMAT_NUMBER_00
                    )

                entry_index_end = entry_index + len(cfoods) - 1

                if class_name == class_names[0] and row_difference < 0:
                    entry_index_end = (
                        entry_index_end
                        + abs(row_difference)
                        - len(class_names_without_food)
                    )

                entry_index = entry_index_end + 1
            foods_total_price = sum([f.total_price for f in wfoods])
            wsheet.cell(form_index1, 6, foods_total_price)
            wsheet.cell(form_index1, 7, foods_total_price)

        if len(form_indexes) > max_time_index:
            for time_index in range(max_time_index, len(form_indexes)):
                form_index0, form_index1 = form_indexes[time_index]
                food_index0, food_index1 = (
                    form_index0 + 2,
                    form_index1 - 1,
                )

                for row in wsheet.iter_rows(
                    min_row=food_index0,
                    max_row=food_index1,
                    min_col=2,
                    max_col=7,
                ):
                    for cell in row:
                        cell.value = ""

        self.format_warehousing_sheet()
        wb = self.get_bill_workbook()
        wb.active = wsheet

        print_info(_("Sheet '%s' was updated.") % (self.warehousing_sheet_name))

    def get_unit_df(self):
        if not self._unit_df is None:
            return self._unit_df
        self._unit_df = pd.read_excel(
            self.get_main_spreadsheet_path(),
            sheet_name=self.unit_sheet_name,
            names=["Name", "Unit"],
        )
        return self._unit_df

    def get_check_df_from_spreadsheet(self):
        check_df = pd.DataFrame(self.get_check_sheet().values)
        check_df.columns = check_df.iloc[0]
        check_df = check_df[1:]
        check_df = check_df.dropna(axis=0, how="all")
        if check_df.shape[0] == 0:
            return None

        return check_df

    def get_check_df(self):
        if not self._check_df is None:
            return self._check_df
        self._check_df = self.get_check_df_from_spreadsheet()
        return self._check_df

    def clear_check_df(self):
        self._check_df = None

    def get_unit_sheet(self):
        return self.get_bill_sheet(self.unit_sheet_name)

    def get_entry_row_index_of_unit_sheet(self):
        unit_sheet = self.get_unit_sheet()
        row_index = 1
        for row in unit_sheet.iter_rows(
            min_row=1, max_row=unit_sheet.max_row + 1, max_col=1
        ):
            cell = row[0]
            if cell.value == None:
                return row_index
            row_index += 1

        return None

    def add_food_names_to_unit_sheet(self, names):
        unit_sheet = self.get_unit_sheet()
        entry_row_index = self.get_entry_row_index_of_unit_sheet()
        for index, name in enumerate(names):
            unit_sheet.cell(entry_row_index + index, 1, name)

    def get_main_spreadsheet_path(self):
        return (
            self._main_spreadsheet_path
            or self.get_main_spreadsheet_path_of_profile()
        )

    def get_main_spreadsheet_path_of_profile(self):
        s_fpath = bill0_fpath
        ps_fpath = user_data_dir / self.profile.label / "workbook.xlsx"
        if not ps_fpath.parent.exists():
            os.makedirs(ps_fpath.parent)
            print_info(_("Directory '%s' has been made.") % (ps_fpath.parent))

        if not ps_fpath.exists():
            shutil.copy(s_fpath, ps_fpath)
            print_info(
                _("Workbook '{0}' was copied to '{1}'.").format(
                    s_fpath, ps_fpath
                )
            )
        print_info(_("Workbook '%s' has been used.") % ps_fpath)
        return ps_fpath

    def get_main_spreadsheet0_path(self):
        _path = bill0_fpath
        return _path

    def set_main_spreadsheet_path(self, file_path=None):
        if not Path(file_path).exists():
            shutil.copy(self.get_main_spreadsheet0_path(), file_path)
        self._main_spreadsheet_path = file_path

    def get_bill_workbook(self):
        if self.bill_workbook:
            return self.bill_workbook
        self.bill_workbook = load_workbook(self.get_main_spreadsheet_path())
        return self.bill_workbook

    def clear_workbook(self):
        self.bill_workbook = None

    def save_bill_workbook(self, wb_fpath=None, info=None):
        info = info or _("Saving workbook. . .")
        print_info(info)
        wb_fpath = wb_fpath or self.get_main_spreadsheet_path()
        wb = self.get_bill_workbook()
        wb.save(wb_fpath)
        print_info(
            _("Workbook '%s' was saved.") % self.get_main_spreadsheet_path()
        )

    def print_dir_was_created_info(self, dir_path):
        print_info(_("Directory %s was created.") % (dir_path))

    def get_profile_data_dpath(self):
        dpath = user_data_dir / self.bill.profile.label
        if not dpath.exists():
            os.makedirs(dpath)
            self.print_dir_was_created_info(dpath)
        return dpath

    def get_profile_copy_data_dpath(self):
        dpath = self.get_profile_data_dpath()
        dpath = dpath / "copy"
        if not dpath.exists():
            os.makedirs(dpath)
            self.print_dir_was_created_info(dpath)
        return dpath

    def get_random_bill_workbook_copy_fpath(self):
        file_path = self.get_main_spreadsheet_path()
        file_path = file_path.as_posix()
        file_path = file_path.split(os.sep)[-1]
        file_path = file_path[:-5] + f".{uuid.uuid4()}.xlsx"
        file_path = self.get_profile_copy_data_dpath() / file_path

        return file_path

    def copy_bill_workbook(self, wb_fpath=None):
        print_info(_("Copying workbook..."))
        file_path = wb_fpath or self.get_random_bill_workbook_copy_fpath()
        wb = self.get_bill_workbook()
        wb.save(file_path)
        print_info(_("Workbook '%s' was saved.") % file_path)
        return file_path

    def get_bill_sheet(self, name):
        wb = self.get_bill_workbook()
        return wb[name]

    def get_check_sheet(self):
        return self.get_bill_sheet(self.check_sheet_name)


# The end.
