# # 这是转角双层石墨烯的连续模型和紧束缚模型的代码整理，里面包含了一些功能实现，如吸收谱，拉曼光谱等。
import functools
import numpy as np
import cmath
import multiprocessing
import os.path
import time
from multiprocessing import Pool
from os.path import expanduser
from typing import Union, Literal, NoReturn
import scipy.signal

import numpy as np
from numpy import (
    exp,
    sin,
    array,
    conj,
    linspace,
    real,
    arange,
    dot,
    save,
    zeros,
    imag,
    ones,
    average,
    sqrt,
    pi,
    cos,
    kron,
    argsort,
    block,
    eye,
    diag,
    arcsin,
    arccos,
)

# basic parameters
from numpy.linalg import norm, eig
from scipy.special import struve, yn
from scipy import interpolate
import cv2 as cv
import shutil
import openpyxl
import csv
import cv2

import matplotlib

matplotlib.use("agg")
# matplotlib.rcParams['font.family'] = ['arial']
import matplotlib.pyplot as plt
import matplotlib.animation as animation
from matplotlib.axes import Axes
from matplotlib.figure import Figure

plt.rc("font", family="Times New Roman")  # change the font of plot
plt.rcParams["mathtext.fontset"] = "stix"  # change the font of math
from matplotlib.lines import Line2D

import plotly.graph_objs as go
from scipy.linalg import expm, block_diag
from cmath import log

import plotly.express as px
from plotly.subplots import make_subplots

from time import perf_counter
import matplotlib.patches as mpatches
import matplotlib.transforms as mtransforms
from mpl_toolkits.axes_grid1.inset_locator import inset_axes
import matplotlib.ticker as tck
import itertools
import sys
import pandas as pd
from mpl_toolkits.mplot3d import Axes3D
import matplotlib.ticker as ticker
from scipy.optimize import curve_fit
from scipy.optimize import least_squares

# from dash import Dash, html, dcc, callback, Output, Input, dash_table, State, ctx
# from dash.exceptions import PreventUpdate
import dash

import statsmodels.api as sm

from scipy.signal import savgol_filter

import json


from public.error import *

# # 这部分为在计算当中需要用到的常数
from public.consts import *

global_fig_format = ".pdf"
global_title_font_size = 16
global_label_font_size = 14
global_tick_font_size = 12
global_legend_font_size = 13
data_file_dir = expanduser("~/code") + os.sep + "Data" + os.sep


class PlotMethod:
    root_dir = "Plots/"

    def __init__(
        self,
        x_in,
        y_in,
        x_label="",
        y_label="",
        x_lim=None,
        y_lim=None,
        title=None,
        save_name="fig",
        legend=None,
        save_dir="tmp",
        text=None,
        text_xy=None,
        existing_fig_ax=None,
        line_type=None,
        ax_return=False,
        fig_size=(7, 5),
        hold_on=False,
        marker_type=None,
        colors: list = None,
    ) -> None:
        self.x = x_in
        self.y = y_in

        self.list_len = len(
            [ele for ele in self.x if isinstance(ele, (list, np.ndarray))]
        )
        self.scatter_len = len([ele for ele in self.x if isinstance(ele, (float, int))])
        if marker_type is None:
            marker_type = ["."] * self.scatter_len
        if line_type is None:
            line_type = ["-"] * self.list_len

        self.xlabel = x_label
        self.ylabel = y_label
        self.xlim = x_lim
        self.ylim = y_lim
        self.fig_size = fig_size
        if title is None:
            self.title = "title"
        else:
            self.title = title
        self.legend = legend

        self.line_type = line_type
        self.marker_type = marker_type

        self._save_dir = data_file_dir + self.root_dir + save_dir
        if not os.path.exists(self._save_dir):
            os.makedirs(self._save_dir)

        self._save_name = save_name
        self.save_pdf_path = self._save_dir + "/{}.pdf".format(self._save_name)
        self.save_png_path = self._save_dir + "/{}.png".format(self._save_name)

        self.ax_return = ax_return
        self.hold_on = hold_on

        self.text = text
        self.text_xy = text_xy

        if colors is None:
            default_colors = [
                "#1f77b4",
                "#ff7f0e",
                "#2ca02c",
                "#d62728",
                "#9467bd",  # 使用颜色编码定义颜色
                "#8c564b",
                "#e377c2",
                "#7f7f7f",
                "#bcbd22",
                "#17becf",
            ]
            self.colors = default_colors[: len(self.x)]
        else:
            self.colors = colors

        if existing_fig_ax is None:
            self.fig, self.ax = plt.subplots(figsize=self.fig_size, dpi=330)
        else:
            self.fig, self.ax = existing_fig_ax

    def __del__(self) -> NoReturn:
        plt.close(self.fig)

    @property
    def save_dir(self):
        return self._save_dir

    @save_dir.setter
    def save_dir(self, save_dir: str):
        self._save_dir = data_file_dir + self.root_dir + save_dir
        if not os.path.exists(self._save_dir):
            os.makedirs(self._save_dir)

    @property
    def save_name(self):
        return self._save_name

    @save_name.setter
    def save_name(self, save_name: str):
        self._save_name = save_name
        self.save_pdf_path = self._save_dir + "/{}.pdf".format(self._save_name)
        self.save_png_path = self._save_dir + "/{}.png".format(self._save_name)
        return

    def new_fig_ax(self):
        self.fig, self.ax = plt.subplots(figsize=self.fig_size)

    def fig_and_ax(self):
        """Return Figure and axe object

        Returns:
            fig, ax
        """

        return self.fig, self.ax

    def set_ax(self, ax: Axes = None):
        if ax is None:
            ax = self.ax
        ax.set_aspect("auto")
        ax.set_xlabel(self.xlabel, fontsize=global_label_font_size)
        ax.set_ylabel(self.ylabel, fontsize=global_label_font_size)
        ax.set_title(self.title, fontsize=global_title_font_size)
        ax.tick_params(axis="x", labelsize=global_tick_font_size)
        ax.tick_params(axis="y", labelsize=global_tick_font_size)
        if self.xlim is None:
            ax.set_xlim(ax.get_xlim())
        else:
            ax.set_xlim(self.xlim)
        if self.ylim is None:
            ax.set_ylim(ax.get_ylim())
        else:
            ax.set_ylim(self.ylim)

        return

    def set_legends(self):
        if len(self.legend) == len(self.x):
            handles, labels = self.ax.get_legend_handles_labels()
            self.ax.legend(handles, labels, fontsize=global_legend_font_size)
        else:
            self.ax.legend(self.legend, fontsize=global_legend_font_size)

    def save_fig(self, fig=None, pdf_path=None, png_path=None):
        """
        Used to save the figure
        """
        if fig is None:
            fig = self.fig
        if (pdf_path is None) or (png_path is None):
            fig.savefig(
                self.save_pdf_path,
                dpi=330,
                facecolor="w",
                bbox_inches="tight",
                pad_inches=0.1,
            )
            fig.savefig(
                self.save_png_path,
                dpi=330,
                facecolor="w",
                bbox_inches="tight",
                pad_inches=0.1,
            )
        else:
            fig.savefig(
                pdf_path, dpi=330, facecolor="w", bbox_inches="tight", pad_inches=0.1
            )
            fig.savefig(
                png_path, dpi=330, facecolor="w", bbox_inches="tight", pad_inches=0.1
            )

        return

    def post_op(self, fig=None, ax=None):
        if fig is None:
            fig = self.fig
        if ax is None:
            ax = self.ax
        if self.ax_return:
            return fig, ax
        elif not self.hold_on:
            plt.close(fig)

    def line_plot(self):
        print("Plotting line...")
        fig, ax = self.fig_and_ax()
        if not (self.legend is None):
            ax.plot(self.x, self.y, label=self.legend[0])
        else:
            ax.plot(self.x, self.y)

        self.set_ax(ax)
        self.set_legends()
        self.save_fig(fig)

        return self.post_op(fig, ax)

    def multiple_line_one_plot(self, opacity=None, text_annotate=False):
        list_len = len([ele for ele in self.x if isinstance(ele, (np.ndarray, list))])
        scatter_len = len([ele for ele in self.x if isinstance(ele, (float, int))])
        if self.line_type is None:
            self.line_type = ["-"] * list_len
        if self.marker_type is None:
            self.marker_type = ["."] * scatter_len
        line_i = 0
        if opacity is None:
            opacity = [1] * list_len
        marker_i = 0
        fig, ax = self.fig_and_ax()
        for i in range(len(self.x)):
            if isinstance(self.x[i], (float, int)):  # Plot as scatter
                ax.scatter(
                    self.x[i],
                    self.y[i],
                    marker=self.marker_type[marker_i],
                    label=self.legend[i],
                )
                marker_i += 1
            elif isinstance(self.x[i], (np.ndarray, list)):  # Plot as line
                if len(self.legend) == len(self.x):
                    ax.plot(
                        self.x[i],
                        self.y[i],
                        self.line_type[line_i],
                        alpha=opacity[line_i],
                        label=self.legend[i],
                        color=self.colors[i],
                    )
                    line_i += 1
                else:
                    ax.plot(
                        self.x[i],
                        self.y[i],
                        self.line_type[line_i],
                        alpha=opacity[line_i],
                        color=self.colors[i],
                    )
                    line_i += 1

        if len(self.legend) == len(self.x):
            handles, labels = ax.get_legend_handles_labels()
            ax.legend(handles, labels)
        else:
            ax.legend(self.legend)

        self.set_ax(ax)

        if text_annotate:
            if self.text is None:
                raise TypeError("The input text and text_xy should not be None!")

            ax.text(self.text_xy[0], self.text_xy[1], self.text, transform=ax.transAxes)

        self.save_fig(fig)

        return self.post_op(fig, ax)

    def multiple_line_multiple_plot(self):
        """
        Save different Plot in different files
        """

        for ele_i in range(len(self.x)):
            save_pdf_path = self._save_dir + "/{}_{}.pdf".format(
                self._save_name, ele_i + 1
            )
            save_png_path = self._save_dir + "/{}_{}.png".format(
                self._save_name, ele_i + 1
            )
            fig, ax = self.fig_and_ax()
            ax.plot(self.x[ele_i], self.y[ele_i])

            self.set_ax(ax)

            self.save_fig(fig, pdf_path=save_pdf_path, png_path=save_png_path)

            plt.close(fig)

        return

    def scatter_plot(self):
        """
        Plot x and y as scatters
        """
        fig, ax = self.fig_and_ax()
        ax.scatter(self.x, self.y, marker=".")

        self.set_ax(ax)

        self.save_fig(fig)

        return self.post_op(fig, ax)

    def multiple_scatter_one_plot(self):
        if self.legend is None:
            raise TypeError("You should input legends to label each component")
        fig, ax = self.fig_and_ax()

        for ele_i in range(len(self.x)):
            ax.scatter(self.x[ele_i], self.y[ele_i], label=self.legend[ele_i])

        if len(self.legend) == len(self.x):
            handles, labels = ax.get_legend_handles_labels()
            ax.legend(handles, labels)
        else:
            ax.legend(self.legend)

        self.set_ax(ax)

        self.save_fig(fig)

        return self.post_op(fig, ax)

    @staticmethod
    def add_right_cax(ax, pad, width):  # ax is the axe of figure
        axpos = ax.get_position()

        caxpos = mtransforms.Bbox.from_extents(
            axpos.x1 + pad, axpos.y0, axpos.x1 + pad + width, axpos.y1
        )

        cax = ax.figure.add_axes(caxpos)

        return cax

    def center_of_curve(self):
        """
        Get the center of the 2D-curve numerically.

        #   Return
        xbar, ybar
        """
        diff_y = np.diff(self.y)
        diff_x = np.diff(self.x)

        der_y = diff_y / diff_x

        ds_factor = sqrt(1 + der_y**2)

        x_center_kernel = self.x[:-1] * ds_factor
        y_center_kernel = self.y[:-1] * ds_factor

        integral_s = np.trapz(ds_factor, self.x[:-1])

        integral_x = np.trapz(x_center_kernel, self.x[:-1])
        integral_y = np.trapz(y_center_kernel, self.x[:-1])

        x_bar = integral_x / integral_s
        y_bar = integral_y / integral_s

        return x_bar, y_bar

    def h_lines(self, color_list: list[str]):
        fig, ax = self.fig_and_ax()
        if len(np.array(self.y).shape) > 1:
            for ele_i in np.arange(len(self.y)):
                ax.hlines(
                    self.y[ele_i],
                    ax.get_xlim()[0],
                    ax.get_xlim()[1],
                    label=self.legend[ele_i],
                    linestyles="dashed",
                    colors=color_list[ele_i],
                )
        else:
            ax.hlines(
                self.y,
                ax.get_xlim()[0],
                ax.get_xlim()[1],
                label=self.legend[0],
                linestyles="dashed",
                colors=color_list[0],
            )
        if len(self.legend) == len(self.y):
            handles, labels = ax.get_legend_handles_labels()
            ax.legend(handles, labels)
        else:
            ax.legend(self.legend)
        self.set_ax(ax)
        self.save_fig(fig)

        return self.post_op(fig, ax)


class PlotExistingAxMethod(PlotMethod):
    def __init__(
        self,
        x_in=None,
        y_in=None,
        x_label="",
        y_label="",
        x_lim=None,
        y_lim=None,
        title="",
        save_name="fig",
        legend=None,
        save_dir="",
        text=None,
        text_xy=None,
        existing_fig_ax: tuple[Figure, Axes] = None,
        line_type=None,
        ax_return=False,
        color="black",
    ) -> None:
        super().__init__(
            x_in,
            y_in,
            x_label,
            y_label,
            x_lim,
            y_lim,
            title,
            save_name,
            legend,
            save_dir,
            text,
            text_xy,
            existing_fig_ax,
            line_type,
            ax_return,
        )
        self.color = color

    def twinx_existing_ax(self, twin_ax_return=False) -> tuple[Figure, Axes]:
        """
        Input the ax to plot twin x plot
        """
        fig, ax = self.fig_and_ax()
        ex_lines = list(ax.get_lines())
        ax_twin = ax.twinx()
        add_line = ax_twin.plot(self.x, self.y, self.line_type, color=self.color)
        ax_twin.set_ylabel(
            self.ylabel, color=self.color, fontsize=global_label_font_size
        )
        ax_twin.spines["right"].set_color(self.color)
        ax_twin.set_ylim(self.ylim)
        ax_twin.set_xlim(self.xlim)
        line_list = ex_lines + add_line
        ax_twin.tick_params(axis="y", color=self.color, labelsize=global_tick_font_size)

        if not (self.title is None):
            ax.set_title(self.title, fontsize=global_title_font_size)

        if not (self.legend is None):
            handles, labels = ax.get_legend_handles_labels()
            leg_list = labels + self.legend
            ax.legend(line_list, leg_list, fontsize=global_legend_font_size)

        for ele_label in ax_twin.get_yticklabels():
            ele_label.set_color(self.color)

        self.save_fig(fig)

        if not twin_ax_return:
            return self.post_op(fig, ax)
        else:
            return fig, ax_twin


class HtmlPlotMethod(PlotMethod):
    def __init__(
        self,
        x_in,
        y_in,
        x_label="",
        y_label="",
        title="",
        save_name="fig",
        legend=None,
        save_dir="",
        text=None,
        text_xy=None,
        existing_fig_ax=None,
    ) -> None:
        super().__init__(
            x_in,
            y_in,
            x_label,
            y_label,
            title,
            save_name,
            legend,
            save_dir,
            text,
            text_xy,
            existing_fig_ax,
        )

        self.root_dir = "PlotsHTML/"

        self.save_dir = data_file_dir + self.root_dir + save_dir

        if not os.path.exists(self.save_dir):
            os.makedirs(self.save_dir)

        self.save_html_path = self.save_dir + "/{}.html".format(self._save_name)
        pass

    # def __init__(self, x_in, y_in, x_label='', y_label='', title='', save_name='fig', legend=None, save_dir="") -> None:
    #     self.x = x_in
    #     self.y = y_in
    #     self.xlabel = x_label
    #     self.ylabel = y_label
    #     self.title = title
    #     self.legend = legend
    #     self.save_name = save_name
    #     self.root_dir = "PlotsHTML/"

    #     self.save_dir = data_file_dir + self.root_dir + save_dir

    #     if not os.path.exists(self.save_dir):
    #         os.makedirs(self.save_dir)

    #     self.save_html_path = self.save_dir + "/{}.html".format(self.save_name)
    #     pass

    def lay_out(self):
        layout = go.Layout(
            title=self.title,
            # width=width,
            # height=height,
            scene=dict(
                xaxis_title=self.xlabel,
                yaxis_title=self.ylabel,
                xaxis=dict(
                    tickfont=dict(
                        size=14,
                        family="Times New Roman",
                    )
                ),
                yaxis=dict(
                    tickfont=dict(
                        size=14,
                        family="Times New Roman",
                    )
                ),
                zaxis=dict(
                    tickfont=dict(
                        size=18,
                        family="Times New Roman",
                    )
                ),
            ),
        )
        return layout

    def line_plot(self, line_name="line"):
        """
        Use HTML to plot data
        """
        if len(array(self.x).shape) > 1:
            raise XYNotSingle(
                "X and Y should only contain one list. If you want to plot multiple lines, please use multiple_line_one_plot function"
            )
        trace1 = go.Scatter(name=line_name, x=self.x, y=self.y)
        fig = go.Figure(data=[trace1], layout=self.lay_out())
        fig.write_html(self.save_dir + "/{}.html".format(self._save_name))

        return

    def multiple_line_one_plot(self):
        trace_list = [
            go.Scatter(name=self.legend[i], x=self.x[i], y=self.y[i])
            for i in range(len(self.x))
        ]
        fig = go.Figure(data=trace_list, layout=self.lay_out())
        fig.write_html(self.save_html_path, include_mathjax="cdn")

        return


class AnimationMethod(PlotMethod):
    """
    parameter static_image_num should be the number of images you want to set to be static. Data of static image should be put ahead of the xy-list.
    """

    root_dir = "PlotsAnimations/"

    def __init__(
        self,
        x_in,
        y_in,
        static_num,
        x_label="",
        y_label="",
        x_lim=None,
        y_lim=None,
        title=None,
        save_name="animation",
        legend=None,
        save_dir="tmp",
        text=None,
        text_xy=None,
        existing_fig_ax=None,
        line_type=None,
        ax_return=False,
        fig_size=(7, 5),
        hold_on=False,
        marker_type=None,
    ) -> None:
        super().__init__(
            x_in,
            y_in,
            x_label,
            y_label,
            x_lim,
            y_lim,
            title,
            save_name,
            legend,
            save_dir,
            text,
            text_xy,
            existing_fig_ax,
            line_type,
            ax_return,
            fig_size,
            hold_on,
            marker_type,
        )

        if static_num < len(self.x):
            self.static_num = static_num
            self.save_ani_path = self.save_dir + "/{}.mp4".format(self.save_name)
        else:
            raise TypeError("Static number should be less than the length of list")

    def multi_scatter_line_core(self, ele_i, out_axes_list, marker_i, line_i):
        if isinstance(self.x[ele_i], (float, int)):
            (tmp_plot,) = self.ax.plot(
                self.x[ele_i],
                self.y[ele_i],
                marker=self.marker_type[marker_i],
                label=self.legend[ele_i],
            )
            marker_i += 1
            out_axes_list.append(tmp_plot)
        elif isinstance(self.x[ele_i], (np.ndarray, list)):
            (tmp_plot,) = self.ax.plot(
                self.x[ele_i],
                self.y[ele_i],
                self.line_type[line_i],
                label=self.legend[ele_i],
            )
            line_i += 1
            out_axes_list.append(tmp_plot)

    def one_plot_anim(self, fps=60):
        marker_i = 0
        line_i = 0
        static_axes = []
        dynamic_axes = []
        frames = len(self.x[self.static_num])
        for ele_i in range(self.static_num):
            self.multi_scatter_line_core(ele_i, static_axes, marker_i, line_i)
        for ele_i in range(self.static_num, len(self.x)):
            if isinstance(self.x[ele_i][0], (float, int)):
                (tmp_plot,) = self.ax.plot(
                    self.x[ele_i][0],
                    self.y[ele_i][0],
                    marker=self.marker_type[marker_i],
                    label=self.legend[ele_i],
                )
                marker_i += 1
                dynamic_axes.append(tmp_plot)
            elif isinstance(self.x[ele_i][0], (np.ndarray, list)):
                (tmp_plot,) = self.ax.plot(
                    self.x[ele_i][0],
                    self.y[ele_i][0],
                    self.line_type[line_i],
                    label=self.legend[ele_i],
                )
                line_i += 1
                dynamic_axes.append(tmp_plot)
        self.set_ax()
        t1 = time.perf_counter()

        def update(num_i):
            for ele_ax_i in range(len(dynamic_axes)):
                dynamic_axes[ele_ax_i].set_data(
                    self.x[self.static_num + ele_ax_i][num_i],
                    self.y[self.static_num + ele_ax_i][num_i],
                )

        ani = animation.FuncAnimation(
            fig=self.fig, func=update, frames=frames, interval=200
        )
        ani.save(self.save_ani_path, fps=fps)
        print("Complete: ", time.perf_counter() - t1)

        return


class ExcelMethod:
    def __init__(self, file_path) -> None:
        self.file_path = file_path
        pass

    @staticmethod
    def excel_col_indices(len_of_column):
        """
        Create list of column indices.
        """
        basic_list = [chr(i).capitalize() for i in range(97, 123)]
        basic_len = len(basic_list)
        if len_of_column <= basic_len:
            return basic_list[:len_of_column]
        else:
            first_letter_i = len_of_column // basic_len
            secon_letter_i = len_of_column % basic_len

            extend_list = basic_list[:]
            begin_first = 0
            while first_letter_i > 0:
                first_letter_i = first_letter_i - 1
                ele_first_letter = basic_list[begin_first]
                begin_first = begin_first + 1
                if first_letter_i == 0:
                    extend_list = extend_list + [
                        ele_first_letter + ele_secon
                        for ele_secon in basic_list[:secon_letter_i]
                    ]
                else:
                    extend_list = extend_list + [
                        ele_first_letter + ele_secon for ele_secon in basic_list
                    ]
            return extend_list

    def read_xlsx_data(self, exclude_rows_num=1, sheet_index=0):
        data = openpyxl.load_workbook(self.file_path)
        sheet = data[data.sheetnames[sheet_index]]
        column_to_read = sheet.max_column
        max_row_num = sheet.max_row
        ele_is_column_list = []
        for column_i in range(column_to_read):
            column_tag = self.excel_col_indices(column_to_read)
            ele_column_cell = sheet[
                "{0}{2}:{0}{1}".format(
                    column_tag[column_i], max_row_num, 1 + exclude_rows_num
                )
            ]
            ele_cell_dat = [
                ele_column_cell[ele_i][0].value
                for ele_i in range(max_row_num - exclude_rows_num)
            ]
            ele_is_column_list.append(ele_cell_dat)
        return ele_is_column_list


class LineFit:
    def __init__(self, x_in, y_in) -> None:
        self.x = x_in
        self.y = y_in
        pass

    def direct_fit(self):
        """
        Use curve fit function directly to fit the curve
        """

        def line_func(x, k, b):
            return k * x + b

        popt, pcov = curve_fit(line_func, self.x, self.y)

        # print("The parameters of the line are (k, b): {:.3f}, {:.3f}".format(*popt))

        return popt

    def fit_through_fixed_point(self, fixed_p):
        """
        Use least square root

        fixed_p should be (p_x, p_y)
        """

        def line_through_dot(x, k):
            return k * (x - fixed_p[0]) + fixed_p[1]

        popt, pcov = curve_fit(line_through_dot, self.x, self.y)

        # print("The parameters of the line are (k): {:.3f}".format(*popt))

        return popt


class FitMethod:
    def __init__(self) -> None:
        pass

    @staticmethod
    def std_between_two_curves(y1_m, y2):
        y1_arr = array(y1_m)
        y2_arr = array(y2)

        if len(y1_arr.shape) == 1:
            std_out = sqrt(np.sum((y1_arr - y2_arr) ** 2, axis=0))
        elif len(y1_arr.shape) == 2:
            std_out = sqrt(np.sum((y1_arr - y2_arr) ** 2, axis=1))

        return std_out


class Functions:
    def __init__(self) -> None:
        pass

    @staticmethod
    def linear_func_fixed_pars(func_type=1, *pars):
        """
        Type:
        1. k and b
        2. k and one point
        3. two points
        """
        if func_type == 1:  ##  k and b

            def func_1(x):
                return pars[0] * x + pars[1]

            return func_1
        elif func_type == 2:  ## k and one dot

            def func_2(x):
                return pars[0] * (x - pars[1][0]) + pars[1][1]

            return func_2
        elif func_type == 3:  ##  two dots
            k = (pars[0][1] - pars[1][1]) / (pars[0][0] - pars[1][0])

            def func_3(x):
                return k * (x - pars[0][0]) + pars[0][1]

            return func_3

    @staticmethod
    def linear_func(func_type=1):
        """
        Type:
        1. k and b
        2. k and one dot
        3. two dots
        """
        if func_type == 1:  ##  k and b

            def func_1(x, k, b):
                return k * x + b

            return func_1
        elif func_type == 2:  ## k and one dot

            def func_2(x, *k_and_p):
                return k_and_p[0][0] * (x - k_and_p[0][1][0]) + k_and_p[0][1][1]

            return func_2
        elif func_type == 3:  ##  two dots

            def func_3(x, *p_and_p):
                k = (p_and_p[0][0][1] - p_and_p[0][1][1]) / (
                    p_and_p[0][0][0] - p_and_p[0][1][0]
                )
                return k * (x - p_and_p[0][0][0]) + p_and_p[0][0][1]

            return func_3


class UnitsConversion:
    def __init__(self) -> None:
        pass

    @staticmethod
    def energy2omega(energy_in_meV):
        omega_arr = array(energy_in_meV) / (h_bar_eV * eV2meV)

        return omega_arr


class PubMeth:
    if "SLURM_CPUS_PER_TASK" in os.environ:
        cores_num = int(os.environ["SLURM_CPUS_PER_TASK"])
    else:
        cores_num = multiprocessing.cpu_count()
    print("Cores Num: ", cores_num)

    title_font_size = global_title_font_size
    label_font_size = global_label_font_size
    tick_font_size = global_tick_font_size

    @staticmethod  # # 寻找两个数的最大公约数
    def gcd(a, b):
        if a < b:
            return PubMeth.gcd(b, a)
        while a % b != 0:
            temp = b
            b = a % b
            a = temp
        return b

    @staticmethod
    def excel_col_indices(self, len_of_column):
        """
        Create list of column indices.
        """
        basic_list = [chr(i).capitalize() for i in range(97, 123)]
        basic_len = len(basic_list)
        if len_of_column <= basic_len:
            return basic_list[:len_of_column]
        else:
            first_letter_i = len_of_column // basic_len
            secon_letter_i = len_of_column % basic_len

            extend_list = basic_list[:]
            begin_first = 0
            while first_letter_i > 0:
                first_letter_i = first_letter_i - 1
                ele_first_letter = basic_list[begin_first]
                begin_first = begin_first + 1
                if first_letter_i == 0:
                    extend_list = extend_list + [
                        ele_first_letter + ele_secon
                        for ele_secon in basic_list[:secon_letter_i]
                    ]
                else:
                    extend_list = extend_list + [
                        ele_first_letter + ele_secon for ele_secon in basic_list
                    ]
            return extend_list

    @staticmethod  # # 寻找质数对
    def get_coprime(limit=5):
        out_coprimes = []
        for r in range(1, limit):
            for m in range(1, limit):
                if PubMeth.gcd(m, r) == 1:
                    out_coprimes.append((m, r))
        return out_coprimes

    @staticmethod  # 在多线程运算当中需要用到的重载函数
    def overdrive_func(
        core_i, input_func, list_to_count, other_args_list, trans_out_list, hint=False
    ):
        if hint:
            print("Core %s" % core_i, " is running")
        out_list = [
            input_func(ele_to_count, *other_args_list) for ele_to_count in list_to_count
        ]
        out_list.append(core_i)
        trans_out_list.append(out_list)

    @staticmethod  # # 这里是多线程运算需要用到的函数，可对计算函数进行重载并进行多进程运算（一般用于对K空间的积分），该函数可极大地利用电脑的核数进行计算
    def multi_proc_func(input_func, divided_list, args_list_f, hint=False):
        out_list_f = multiprocessing.Manager().list()
        args_list = [args_list_f, out_list_f, hint]
        path_num = len(divided_list)
        p_f = Pool(path_num)
        for i in range(path_num):
            tmp_list = args_list[:]
            tmp_list.insert(0, i)  # 进程序号
            tmp_list.insert(1, input_func)  # 输入的函数
            tmp_list.insert(2, divided_list[i])  # 输入的需要计算的（第一个）参数
            p_f.apply_async(PubMeth.overdrive_func, tuple(tmp_list))
        if hint:
            print("Waiting for all subprocesses done...")
        p_f.close()
        p_f.join()
        if hint:
            print("All subprocesses done.")
        total_list = []
        for path_i in range(path_num):
            for ele_list in out_list_f:
                if ele_list[-1] == path_i:
                    total_list.extend(ele_list[0:-1])
        return total_list

    @classmethod  # # 将K空间的k点做一个划分，以便进行多进程运算
    def divide_list(cls, input_list):
        kp_part_list_f = []
        for i in range(cls.cores_num - 1):
            kp_part_list_f.append(
                input_list[
                    int(len(input_list) / cls.cores_num)
                    * i : int(len(input_list) / cls.cores_num)
                    * (i + 1)
                ]
            )
        kp_part_list_f.append(
            input_list[int(len(input_list) / cls.cores_num) * (cls.cores_num - 1) :]
        )
        return kp_part_list_f

    @staticmethod  # # 对不同的操作系统可获得正确的保存路径
    def get_right_save_path_and_create(
        input_dir_name, data_files_or_not=True, create_or_not=True
    ):  # only take one parameter
        """
        If the directory name is like "A/B/C", then it shouldn't include the os.sep in the last position
        """
        if data_files_or_not:
            right_path = (
                expanduser("~/code")
                + os.sep
                + "Data"
                + os.sep
                + input_dir_name
                + os.sep
            )
            if create_or_not:
                if not os.path.exists(right_path):
                    os.makedirs(right_path)
            return (
                expanduser("~/code")
                + os.sep
                + "Data"
                + os.sep
                + input_dir_name
                + os.sep
            )
        else:
            right_path = os.getcwd() + os.sep + input_dir_name + os.sep
            if create_or_not:
                if not os.path.exists(right_path):
                    os.makedirs(right_path)
            return os.getcwd() + os.sep + input_dir_name + os.sep

    @staticmethod  # # 单层石墨烯的哈密顿量，已对摩尔布里渊区进行了约化
    def get_k_mat(a0, theta, Kg, kx, ky):
        return array(
            [
                [1 / Kg, sqrt(3) / 2 * a0 * (kx - 1j * ky) * exp(-1j * theta / 2)],
                [sqrt(3) / 2 * a0 * (kx + 1j * ky) * exp(1j * theta / 2), 1 / Kg],
            ]
        )

    @staticmethod  # # 将参数列表转化为矩阵形式
    def list2mat(input_list):
        return array(input_list).reshape(
            (int(sqrt(len(input_list))), int(sqrt(len(input_list))))
        )

    @staticmethod  # # 将参数列表转化为矩阵形式
    def putin2mat(args_list_in):
        total_list = []
        for ele_args in args_list_in:
            # the second parameter should be an angle(degree)
            total_list.append(ele_args[0] * exp(1j * ele_args[1] / 180 * pi))
        return PubMeth.list2mat(total_list)

    @staticmethod  # # 判断目标点是否位于选定区域内
    def isInterArea(testPoint, AreaPoint):  # testPoint为待测点[x,y]
        # AreaPoint为按顺时针顺序的4个点[[x1,y1],[x2,y2],[x3,y3],[x4,y4]]
        LBPoint = AreaPoint[0]
        LTPoint = AreaPoint[1]
        RTPoint = AreaPoint[2]
        RBPoint = AreaPoint[3]
        a_f = (LTPoint[0] - LBPoint[0]) * (testPoint[1] - LBPoint[1]) - (
            LTPoint[1] - LBPoint[1]
        ) * (testPoint[0] - LBPoint[0])
        b_f = (RTPoint[0] - LTPoint[0]) * (testPoint[1] - LTPoint[1]) - (
            RTPoint[1] - LTPoint[1]
        ) * (testPoint[0] - LTPoint[0])
        c_f = (RBPoint[0] - RTPoint[0]) * (testPoint[1] - RTPoint[1]) - (
            RBPoint[1] - RTPoint[1]
        ) * (testPoint[0] - RTPoint[0])
        d_f = (LBPoint[0] - RBPoint[0]) * (testPoint[1] - RBPoint[1]) - (
            LBPoint[1] - RBPoint[1]
        ) * (testPoint[0] - RBPoint[0])
        if (a_f > 0 and b_f > 0 and c_f > 0 and d_f > 0) or (
            a_f < 0 and b_f < 0 and c_f < 0 and d_f < 0
        ):
            return True
        else:
            return False

    @staticmethod  # # 判断点是否位于边界
    def at_corners(chosen_p, corner_list):
        test_p_f = array([chosen_p[0], chosen_p[1]])
        distance_list = []
        for ele_corner in corner_list:
            corner_arr = array([ele_corner[0], ele_corner[1]])
            diff = corner_arr - test_p_f
            distance_list.append(np.linalg.norm(diff))
        min_d = min(distance_list)
        if min_d < 0.001:
            return True

    @staticmethod  # # 判断一个晶格点在哪个超晶格基矢下的映射，其长度最短（用于紧束缚模型）
    def get_smallest_distance(original_distance_arr, input_vec_list):
        transformed_vecs_list = original_distance_arr + array(input_vec_list)
        transformed_norm_list = norm(transformed_vecs_list, axis=1)
        index_min = list(transformed_norm_list).index(min(transformed_norm_list))
        return transformed_vecs_list[index_min]

    @staticmethod  # # 判断矢量是否位于矢量列表中
    def arr_index_in_arr_list(vec, vec_list, precision=0.01):
        transformed_vecs_list = vec - array(vec_list)
        transformed_norm_list = norm(transformed_vecs_list, axis=1)
        if min(transformed_norm_list) < precision:
            return list(transformed_norm_list).index(min(transformed_norm_list))
        else:
            return "The vector is not in the list!"

    @staticmethod
    def arr_plus_arr_list(arr, arr_list):  # # returns a list
        return list(arr + arr_list)

    @staticmethod  # # 旋转矩阵，可作用于矢量得到旋转之后的矢量
    def rotation(angle):  # arc unit
        """
        2D Rotation matrix
        """
        theta = angle / 180 * pi
        rot_mat = array([[cos(theta), -sin(theta)], [sin(theta), cos(theta)]])
        return rot_mat

    @staticmethod  # # 点到点设定若干个点进行k空间路径描述
    def p2p(point1, point2, density_per_path):
        k_along = []
        if point2[0] != point1[0]:
            k_slope = (point2[1] - point1[1]) / (point2[0] - point1[0])
            number_of_points = int(
                density_per_path
                * sqrt((point2[1] - point1[1]) ** 2 + (point2[0] - point1[0]) ** 2)
            )
            for xx in linspace(point1[0], point2[0], number_of_points):
                k_along.append((xx, k_slope * (xx - point2[0]) + point2[1]))
        elif point2[0] == point1[0]:
            number_of_points = int(
                density_per_path * sqrt((point2[1] - point1[1]) ** 2)
            )
            for yy in linspace(point1[1], point2[1], number_of_points):
                k_along.append((point2[0], yy))
        return k_along

    @staticmethod  # # 三角格子划分，得到若干个shell下的所有点的矢量列表
    def tri_lattice(
        shell_i,
        a0_lattice,
        rotation_angle=0,
        shift_arr=array([0, 0]),
        dim3=False,
        z_value=0,
    ):
        base_vecs = [
            a0_lattice * array([1, 0]),
            a0_lattice * array([-1, 0]),
            a0_lattice * array([1 / 2, sqrt(3) / 2]),
            a0_lattice * array([1 / 2, -sqrt(3) / 2]),
            a0_lattice * array([-1 / 2, sqrt(3) / 2]),
            a0_lattice * array([-1 / 2, -sqrt(3) / 2]),
        ]
        all_vecs = []
        ru_list = []
        for i1 in range(1, shell_i + 1):
            all_vecs.extend(list(array(base_vecs) * i1))
        for j in range(1, shell_i):
            tmp_R6 = j * a0_lattice * array([1 / 2, sqrt(3) / 2])
            tmp_R1 = a0_lattice * array([1, 0])
            for i2 in range(1, shell_i + 1 - j):
                ru_list.append(tmp_R6 + i2 * tmp_R1)
        all_vecs.extend(ru_list)
        for i3 in range(1, 6):
            tmp_theta = 60 * i3
            tmp_new = [PubMeth.rotation(tmp_theta) @ vec for vec in ru_list]
            all_vecs.extend(tmp_new)
        all_vecs.insert(0, array([0, 0]))
        all_vecs = [
            PubMeth.rotation(rotation_angle) @ ele_vec for ele_vec in all_vecs
        ]  # Rotate all the lattices
        all_vecs = (
            array(all_vecs) + PubMeth.rotation(rotation_angle) @ shift_arr
        )  # Shift all the lattices
        if dim3:
            all_vecs = [
                array([ele_vec[0], ele_vec[1], z_value]) for ele_vec in all_vecs
            ]
        else:
            pass
        return list(all_vecs)

    @staticmethod  # # 三角格子划分，得到若干个shell下的所有点的矢量列表
    def tri_lattice_mod(
        shell_i, primitive_vec, shift_arr=array([0, 0]), dim3=False, z_value=0
    ):
        """
        shell_i: # of shells to plot
        primitive_vec: one of the primitive vec
        Plot the 2D/3D triangular lattice, if only need to plot the 2D lattices, "tri_lattice" function is preferred.
        """
        base_vecs = [PubMeth.rotation(60 * i) @ primitive_vec[:2] for i in range(0, 6)]

        all_vecs = []
        ru_list = []
        for i1 in range(1, shell_i + 1):
            all_vecs.extend(list(array(base_vecs) * i1))
        for j in range(1, shell_i):
            tmp_R6 = j * base_vecs[0]
            tmp_R1 = base_vecs[1]
            for i2 in range(1, shell_i + 1 - j):
                ru_list.append(tmp_R6 + i2 * tmp_R1)
        all_vecs.extend(ru_list)
        for i3 in range(1, 6):
            tmp_theta = 60 * i3
            tmp_new = [PubMeth.rotation(tmp_theta) @ vec for vec in ru_list]
            all_vecs.extend(tmp_new)
        all_vecs.append(array([0, 0]))
        if dim3:
            all_vecs = [
                array([ele_vec[0], ele_vec[1], z_value]) for ele_vec in all_vecs
            ]
        else:
            pass
        all_vecs = array(all_vecs) + shift_arr

        return list(all_vecs)

    @staticmethod
    def rect2diam(
        input_mat,
        file_name,
        title_name,
        save_2d_plots=True,
        rm_raw=True,
        cmap="jet",
        direction_up=True,
        save_in_case_same_name=False,
        save_mat=False,
        fig_format=global_fig_format,
        mod_pics_save_dir="Pics_mod",
        mat_save_dir="Mat_files",
        ax_mod_input=False,
        fontsize=14,
    ):
        # plot of real part
        plt.figure(dpi=330, figsize=(7, 7))
        plt.imshow(input_mat, cmap=cmap, aspect="auto")  # origin='lower')
        plt.axis("off")
        target_dir_raw = PubMeth.get_right_save_path_and_create(
            "Pics_raw", data_files_or_not=True
        )
        if save_2d_plots:
            if not os.path.exists(target_dir_raw):
                os.makedirs(target_dir_raw)
            if save_in_case_same_name:
                im_raw_save_func = functools.partial(
                    plt.savefig, bbox_inches="tight", pad_inches=0, dpi=330
                )
                PubMeth.save_if_same_name(
                    im_raw_save_func,
                    save_name=target_dir_raw + file_name + "_raw_.png",
                    target_dir="",
                )
            else:
                plt.savefig(
                    target_dir_raw + file_name + "_raw_.png",
                    bbox_inches="tight",
                    pad_inches=0,
                    dpi=330,
                )
        plt.close()

        tmp_imag = cv.imread(target_dir_raw + file_name + "_raw_.png")
        cols, rows = tmp_imag.shape[:2]  # 取长宽
        if direction_up:  # the long axis direction is from bottom to up
            point1 = np.float32([[0, 0], [rows, 0], [0, cols]])
            point2 = np.float32(
                [
                    [int(rows / 2), 0],
                    [rows, int(sqrt(3) / 2 * rows)],
                    [0, int(sqrt(3) / 2 * rows)],
                ]
            )
            M = cv.getAffineTransform(point1, point2)
            dst = cv.warpAffine(
                tmp_imag, M, (rows, int(sqrt(3) * rows)), borderValue=(255, 255, 255)
            )
        else:  # the long axis direction is from left to right
            point1 = np.float32([[0, 0], [rows, 0], [0, cols]])
            point2 = np.float32(
                [
                    [int(cols / 2 * sqrt(3)), 0],
                    [int(cols * sqrt(3)), cols // 2],
                    [0, cols // 2],
                ]
            )
            M = cv.getAffineTransform(point1, point2)
            dst = cv.warpAffine(
                tmp_imag, M, (int(cols * sqrt(3)), cols), borderValue=(255, 255, 255)
            )
        if isinstance(
            ax_mod_input, bool
        ):  # the normal mode, which doesn't related to inputing an axe
            plt.imshow(
                dst
            )  # , origin='lower'  This origin will determine the direction of image
            plt.title(title_name, fontsize=fontsize)
            plt.axis("off")
            target_dir_mod = PubMeth.get_right_save_path_and_create(
                mod_pics_save_dir, data_files_or_not=True
            )
            if not os.path.exists(target_dir_mod):
                os.makedirs(target_dir_mod)
            if save_in_case_same_name:
                im_mod_save_func = functools.partial(
                    plt.savefig, bbox_inches="tight", pad_inches=0.1, dpi=330
                )
                PubMeth.save_if_same_name(
                    im_mod_save_func,
                    target_dir_mod + file_name + "_mod_" + fig_format,
                    target_dir="",
                )
            else:
                plt.savefig(
                    target_dir_mod + file_name + "_mod_" + fig_format,
                    bbox_inches="tight",
                    pad_inches=0.1,
                    dpi=330,
                )
            plt.close()
            if rm_raw:
                shutil.rmtree(target_dir_raw)
            target_dir_mat = PubMeth.get_right_save_path_and_create(
                mat_save_dir, data_files_or_not=True
            )
            if save_mat:
                if not os.path.exists(target_dir_mat):
                    os.makedirs(target_dir_mat)
                if save_in_case_same_name:
                    mat_save_func = functools.partial(np.save, arr=input_mat)
                    PubMeth.save_if_same_name(
                        mat_save_func,
                        save_name=target_dir_mat + file_name + "_mat_.npy",
                        target_dir="",
                    )
                else:
                    np.save(target_dir_mat + file_name + "_mat_.npy", input_mat)
        else:
            ax_mod_input.imshow(
                dst
            )  # , origin='lower'  This origin will determine the direction of image
            ax_mod_input.set_title(title_name, fontsize=fontsize)
            ax_mod_input.axis("off")

    @staticmethod
    def scatter_2d_plot(
        x_list,
        y_list,
        c_mat,
        marker_size=2,
        uni_vmin=False,
        uni_vmax=False,
        marker_type="h",
        colorbar=False,
        colorbar_pad=0.01,
        colorbar_width=0.01,
        figuretitle="title",
        figure_name="figure_2d",
        figs_save_dir="scatter_2d",
        title_font_size=14,
        cbar_label="LABEL",
        cbar_font_size=15,
        cbar_label_pad=13,
        transparent_or_not=False,
        cmap="jet",
        boundary_vecs=None,
        ax_return=False,
    ):
        """
        cbar_label = {"label": , "fontsize": , "labelpad": ,}
        """
        ##  files location
        if figs_save_dir == "scatter_2d":
            figs_save_dir = data_file_dir + "scatter_2d/"
        if not os.path.exists(figs_save_dir):
            os.makedirs(figs_save_dir)
        pdf_dir = figs_save_dir + "pdf/"
        png_dir = figs_save_dir + "png/"
        if not os.path.exists(pdf_dir):
            os.makedirs(pdf_dir)
        if not os.path.exists(png_dir):
            os.makedirs(png_dir)

        ##  Plot of the 2D scatter
        if (not isinstance(uni_vmin, bool)) and (not isinstance(uni_vmax, bool)):
            vmin = uni_vmin
            vmax = uni_vmax
        else:
            vmin = c_mat.min()
            vmax = c_mat.max()
        fig, ax_2d = plt.subplots(figsize=(7, 7))
        im_var = ax_2d.scatter(
            x_list,
            y_list,
            c=c_mat.reshape((len(x_list),)),
            s=marker_size,
            vmin=vmin,
            vmax=vmax,
            marker=marker_type,
            cmap=cmap,
        )
        if not (boundary_vecs is None):
            ax_2d.plot(boundary_vecs[:, 0], boundary_vecs[:, 1], "k-")
        ax_2d.set_aspect("equal")
        ax_2d.axis("off")
        ax_2d.set_title(figuretitle, fontsize=title_font_size)
        ax_2d.set_xlim(
            [1.05 * min(boundary_vecs[:, 0]), 1.05 * max(boundary_vecs[:, 0])]
        )
        if colorbar:
            c_ax = PubMeth.add_right_cax(ax_2d, colorbar_pad, colorbar_width)
            cbar = fig.colorbar(im_var, cax=c_ax)
            cbar.set_label(
                label=cbar_label, fontsize=cbar_font_size, labelpad=cbar_label_pad
            )

        if ax_return:
            png_name: str = png_dir + figure_name + ".png"
            pdf_name: str = pdf_dir + figure_name + ".pdf"
            return fig, ax_2d, png_name, pdf_name

        fig.savefig(
            png_dir + figure_name + ".png",
            bbox_inches="tight",
            pad_inches=0.1,
            dpi=330,
            transparent=transparent_or_not,
        )
        fig.savefig(
            pdf_dir + figure_name + ".pdf",
            bbox_inches="tight",
            pad_inches=0.1,
            dpi=330,
            transparent=transparent_or_not,
        )

        plt.close()

    @staticmethod
    def single_raman_transition(sub_folder_name, target_twist):
        """
        Give the folder name of the Raman matrix files, this function can help you to read it and plot the data as 2D plot.
        """
        raman_mat_files_dir = data_file_dir + "raman_mat_files/"
        mat_subfolder = raman_mat_files_dir + sub_folder_name

        if not os.path.exists(mat_subfolder + "uni_min_max.npy"):
            file_name_list = os.listdir(mat_subfolder)
            all_mats = [
                np.load(mat_subfolder + ele_file_name)
                for ele_file_name in file_name_list
            ]
            uni_real_vmin = array(
                real(all_mats)
            ).min()  # get the min and max of all data. This is the real part
            uni_real_vmax = array(
                real(all_mats)
            ).max()  # get the min and max of all data. This is the real part
            uni_imag_vmin = array(
                imag(all_mats)
            ).min()  # get the min and max of all data. This is the imaginary part
            uni_imag_vmax = array(
                imag(all_mats)
            ).max()  # get the min and max of all data. This is the imaginary part
            ##  save the universal min and max of real and imaginary part
            real_imag_min_max_list = [
                uni_real_vmin,
                uni_real_vmax,
                uni_imag_vmin,
                uni_imag_vmax,
            ]
            np.save(mat_subfolder + "uni_min_max.npy", real_imag_min_max_list)
        else:  ##  If there is the file
            real_imag_min_max_list = np.load(mat_subfolder + "uni_min_max.npy")
            uni_real_vmin = real_imag_min_max_list[0]
            uni_real_vmax = real_imag_min_max_list[1]
            uni_imag_vmin = real_imag_min_max_list[2]
            uni_imag_vmax = real_imag_min_max_list[3]
            file_name_list = os.listdir(mat_subfolder)
            file_name_list.remove("uni_min_max.npy")
            all_mats = [
                np.load(mat_subfolder + ele_file_name)
                for ele_file_name in file_name_list
            ]
        kps_list = PubMeth.uniform_diamond_bz()
        x_list = [ele[0] for ele in kps_list]
        y_list = [ele[1] for ele in kps_list]

        ##  create a folder to save the 2d pics
        target_2d_pics_folder = data_file_dir + "raman_2d_pic/" + sub_folder_name
        if not os.path.exists(target_2d_pics_folder):
            os.makedirs(target_2d_pics_folder)

        ##  plot every mat as 2d plot
        for ele_index in range(len(file_name_list)):
            print("# of files: ", len(file_name_list))
            ##  load element matrix
            ele_mat = all_mats[ele_index]
            ele_file_name = file_name_list[ele_index].split(".")[:-1]
            ele_file_name = ".".join(ele_file_name)
            ##  plot the real part of the matrix
            PubMeth.scatter_2d_plot(
                x_list,
                y_list,
                real(ele_mat),
                marker_size=5.5,
                figuretitle=r"$\bf \theta={}\degree$".format(target_twist),
                colorbar_pad=0.02,
                colorbar_width=0.02,
                title_font_size=23,
                cbar_label="Resonance Intensity",
                figure_name="{}_real".format(ele_file_name),
                figs_save_dir=target_2d_pics_folder,
                uni_vmax=uni_real_vmax,
                uni_vmin=uni_real_vmin,
            )
            ##  plot the imag part of the matrix
            PubMeth.scatter_2d_plot(
                x_list,
                y_list,
                imag(ele_mat),
                marker_size=5.5,
                figuretitle=r"$\bf \theta={}\degree$".format(target_twist),
                colorbar_pad=0.02,
                colorbar_width=0.02,
                title_font_size=23,
                cbar_label="Resonance Intensity",
                figure_name="{}_imag".format(ele_file_name),
                figs_save_dir=target_2d_pics_folder,
                uni_vmin=uni_imag_vmin,
                uni_vmax=uni_imag_vmax,
            )
            print("Complete: ", ele_file_name)

        # tmp_mat = np.load(mat_subfolder + "Conti_raman_{}.000_ABt-TTG_mat_.npy".format(target_twist))

        # PubMeth.scatter_2d_plot(x_list, y_list, real(tmp_mat), marker_size=5.5, figuretitle=r'$\bf \theta={}\degree$'.format(target_twist), colorbar_pad=0.02, colorbar_width=0.02, title_font_size=23, cbar_label='Resonance Intensity', figure_name='{}_real'.format(target_twist), figs_save_dir=target_2d_pics_folder, uni_vmax=uni_real_vmax, uni_vmin=uni_real_vmin)
        # PubMeth.scatter_2d_plot(x_list, y_list, imag(tmp_mat), marker_size=5.5, figuretitle=r'$\bf \theta={}\degree$'.format(target_twist), colorbar_pad=0.02, colorbar_width=0.02, title_font_size=23, cbar_label='Resonance Intensity', figure_name='{}_imag'.format(target_twist), figs_save_dir=target_2d_pics_folder, uni_vmin=uni_imag_vmin, uni_vmax=uni_imag_vmax)

    @staticmethod
    def draw_box(input_vex_arr):
        x_list = array(input_vex_arr)[:, 0]
        y_list = array(input_vex_arr)[:, 1]
        plt.plot(x_list, y_list)

    @staticmethod
    def draw_dots(
        dots_arr_in,
        marker_type,
        save_or_not=False,
        hold_on=False,
        draw_width=1,
        fig_title="",
        x_label="",
        y_label="",
        leg_list=[],
        fig_format=global_fig_format,
    ):
        x_list = array(dots_arr_in)[:, 0]
        y_list = array(dots_arr_in)[:, 1]

        plt.scatter(x_list, y_list, marker=marker_type, linewidths=draw_width)
        ax = plt.gca()
        ax.set_aspect("equal")
        ax.set_title(fig_title)
        ax.set_xlabel(x_label)
        ax.set_ylabel(y_label)
        ax.legend(leg_list)  # this is usually put at the last operation
        if save_or_not:
            tmp_dir = PubMeth.get_right_save_path_and_create(
                "tmp_figs", data_files_or_not=True
            )
            if not os.path.exists(tmp_dir):
                os.makedirs(tmp_dir)
            plt.savefig(tmp_dir + "dots" + fig_format, dpi=330)
        if not hold_on:
            plt.close()

    @staticmethod
    def commensurate_angle(m0, r):
        value_cos = (3 * m0**2 + 3 * m0 * r + r**2 / 2) / (
            3 * m0**2 + 3 * m0 * r + r**2
        )
        twist_angle = np.arccos(value_cos) / pi * 180
        theta = np.arccos(value_cos)
        return twist_angle, theta

    @staticmethod
    def overlap_points(arr_list1, arr_list2):
        overlap_points = []
        for ele_arr in arr_list1:
            tmp_arr_list = arr_list2 - ele_arr
            norm_list = norm(tmp_arr_list, axis=1)
            sift_list = norm_list <= 0.01
            fit_arrs_list = array(arr_list2)[sift_list]
            overlap_points.append(fit_arrs_list)
        return list(np.unique(np.vstack(tuple(overlap_points)), axis=0))

    @staticmethod
    def plus_set(arr_list1, arr_list2, length_limit=10):
        all_overlap_list = []
        for ele_arr in arr_list1:
            tmp_arr_list = arr_list2 + ele_arr
            all_overlap_list.extend(list(tmp_arr_list))
        second_list = np.unique(array(all_overlap_list), axis=0)
        out_list = second_list[norm(second_list, axis=1) < length_limit]
        return list(out_list)

    @staticmethod
    def unit_cell_per_supercell(m0, r):
        if r % 3 != 0:
            N_unit_cell = int(3 * m0**2 + 3 * m0 * r + r**2)
        else:
            N_unit_cell = int(m0**2 + m0 * r + r**2 / 3)
        return N_unit_cell

    @staticmethod
    def moire_pattern(
        m0,
        r,
        shell_i=5,
        a0_constant=1,
        moire_figsave=False,
        k_valley_save=False,
        moire_pm_set_save=False,
        k_space=True,
        k_pm_valley_save=False,
        fig_format=global_fig_format,
    ):
        if k_space:
            N = PubMeth.unit_cell_per_supercell(m0, r)
            print("Number of unit cell in one supercell: ", N)

            comm_ang, comm_theta = PubMeth.commensurate_angle(m0, r)
            print("The twist angle is: ", comm_ang)
            lattice = PubMeth.tri_lattice(shell_i, a0_constant, shift_arr=array([0, 0]))
            lattice_p = [
                PubMeth.rotation(comm_ang / 2 + 30) @ ele_arr for ele_arr in lattice
            ]  # plus 30 degrees to rotate the dots array.
            lattice_m = [
                PubMeth.rotation(-comm_ang / 2 + 30) @ ele_arr for ele_arr in lattice
            ]
            lattice_overlap = PubMeth.overlap_points(lattice_p, lattice_m)

            K0_arr = array([a0_constant / sqrt(3), 0])
            Kp_arr = PubMeth.rotation(comm_ang / 2) @ K0_arr
            Km_arr = PubMeth.rotation(-comm_ang / 2) @ K0_arr
            Delta_K = (
                PubMeth.rotation(comm_ang / 2) @ K0_arr
                - PubMeth.rotation(-comm_ang / 2) @ K0_arr
            )
            Kp_plus_L_p = [Kp_arr + ele_arr for ele_arr in lattice_p]
            Km_plus_L_m = [Km_arr + ele_arr for ele_arr in lattice_m]
            K_overlap = PubMeth.overlap_points(Kp_plus_L_p, Km_plus_L_m)
            print("Len of K overlap is: ", len(K_overlap))
            print(cos(comm_theta / 2) * sqrt(N))
            print(sin(comm_theta / 2) * sqrt(N) * sqrt(3))

            atom_lattice_plus_set = PubMeth.plus_set(
                lattice_p, lattice_m, length_limit=4.5
            )
            tmp_plus_set = [
                array([round(ele_arr[0], 4), round(ele_arr[1], 4)])
                for ele_arr in atom_lattice_plus_set
            ]
            atom_lattice_plus_set = np.unique(array(tmp_plus_set), axis=0)

            if moire_figsave:
                PubMeth.draw_dots(
                    lattice_m, marker_type="o", save_or_not=False, hold_on=True
                )
                PubMeth.draw_dots(
                    lattice_p, marker_type="o", save_or_not=False, hold_on=True
                )
                PubMeth.draw_dots(
                    lattice_overlap,
                    marker_type="*",
                    save_or_not=True,
                    hold_on=False,
                    fig_title=r"Atomic structure: $\theta = %.2f \degree$" % comm_ang,
                    leg_list=[r"$P_{-}$", r"$P_{+}$", r"$P_{-} \cap P_{+}$"],
                )
            if moire_pm_set_save:
                PubMeth.draw_dots(
                    lattice_m, marker_type="o", save_or_not=False, hold_on=True
                )
                PubMeth.draw_dots(
                    lattice_p, marker_type="o", save_or_not=False, hold_on=True
                )
                PubMeth.draw_dots(
                    atom_lattice_plus_set,
                    marker_type=".",
                    save_or_not=True,
                    hold_on=False,
                    draw_width=0.1,
                    fig_title=r"Atomic structure: $\theta = %.2f \degree$" % comm_ang,
                    leg_list=[r"$P_{-}$", r"$P_{+}$", r"$P_{-} + P_{+}$"],
                )
            if k_valley_save:
                PubMeth.draw_dots(
                    Kp_plus_L_p,
                    marker_type="o",
                    save_or_not=False,
                    hold_on=True,
                    draw_width=0.1,
                )
                PubMeth.draw_dots(
                    Km_plus_L_m,
                    marker_type="o",
                    save_or_not=False,
                    hold_on=True,
                    draw_width=0.1,
                )
                if len(K_overlap) != 0:
                    PubMeth.draw_dots(
                        K_overlap,
                        marker_type="*",
                        save_or_not=False,
                        hold_on=True,
                        fig_title=r"$\theta = %.2f \degree$" % comm_ang,
                    )
                    legs = [
                        r"$K_{-}^{0} + P_{-}$",
                        r"$K_{+}^{0} + P_{+}$",
                        r"$\mathcal{Q}_+$",
                        r"$\sqrt{N}K$",
                        r"$-\sqrt{N}K$",
                    ]
                else:
                    legs = [
                        r"$K_{-}^{0} + P_{-}$",
                        r"$K_{+}^{0} + P_{+}$",
                        r"$\sqrt{N}K$",
                        r"$-\sqrt{N}K$",
                    ]
                plt.plot([0, sqrt(N) * K0_arr[0]], [0, sqrt(N) * K0_arr[1]], "r")
                plt.plot([0, -sqrt(N) * K0_arr[0]], [0, -sqrt(N) * K0_arr[1]], "k")
                plt.title(r"K space: $\theta = %.2f \degree$" % comm_ang)
                plt.legend(legs, fontsize=7)
                plt.savefig(
                    "./tmp_figs/k_valley_{}".format(comm_ang) + fig_format, dpi=330
                )
                plt.close()
            if k_pm_valley_save:
                PubMeth.draw_dots(
                    Kp_plus_L_p,
                    marker_type="o",
                    save_or_not=False,
                    hold_on=True,
                    draw_width=0.1,
                )
                PubMeth.draw_dots(
                    Km_plus_L_m,
                    marker_type="o",
                    save_or_not=False,
                    hold_on=True,
                    draw_width=0.1,
                )
                if len(K_overlap) != 0:
                    PubMeth.draw_dots(
                        K_overlap,
                        marker_type="*",
                        save_or_not=False,
                        hold_on=True,
                        fig_title=r"$\theta = %.2f \degree$" % comm_ang,
                    )
                    legs = [
                        r"$K_{+}^{0} + P_{+}$",
                        r"$K_{-}^{0} + P_{-}$",
                        r"$\mathcal{Q}_+$",
                        r"$P_{-} + P_{+}$(shifted)",
                        r"$K_{+}^{0} - K_{+}^{'0}$",
                        "Origin",
                    ]
                else:
                    legs = [
                        r"$K_{+}^{0} + P_{+}$",
                        r"$K_{-}^{0} + P_{-}$",
                        r"$P_{-} + P_{+}$(shifted)",
                        r"$K_{+}^{0} - K_{+}^{'0}$",
                        "Origin",
                    ]
                PubMeth.draw_dots(
                    atom_lattice_plus_set + array([norm(Delta_K), 0]) / sqrt(3),
                    marker_type=".",
                    save_or_not=False,
                    hold_on=True,
                    draw_width=0.1,
                )
                plt.plot([Kp_arr[0], -Kp_arr[0]], [Kp_arr[1], -Kp_arr[1]], "k", lw=0.4)
                plt.scatter(0, 0, marker="x", s=8)
                plt.title(r"K space: $\theta = %.2f \degree$" % comm_ang)
                plt.legend(
                    [
                        r"$K_{+}^{0} + P_{+}$",
                        r"$K_{-}^{0} + P_{-}$",
                        r"$\mathcal{Q}_+$",
                        r"$P_{-} + P_{+}$(shifted)",
                        r"$K_{+}^{0} - K_{+}^{'0}$",
                        "Origin",
                    ],
                    fontsize=6,
                )
                plt.savefig(
                    "./tmp_figs/k_valley_and_pm_sets_{}".format(comm_ang) + fig_format,
                    dpi=330,
                )
                plt.close()
        else:
            if r % 3 != 0:
                N = int(3 * m0**2 + 3 * m0 * r + r**2)
            else:
                N = int(m0**2 + m0 * r + r**2 / 3)
            print("Number of unit cell in one supercell: ", N)

            comm_ang, comm_theta = PubMeth.commensurate_angle(m0, r)
            print("The twist angle is: ", comm_ang)
            lattice = PubMeth.tri_lattice(shell_i, a0_constant, shift_arr=array([0, 0]))
            lattice_p = [
                PubMeth.rotation(comm_ang / 2) @ ele_arr for ele_arr in lattice
            ]  # plus 30 degrees to rotate the dots array.
            lattice_m = [
                PubMeth.rotation(-comm_ang / 2) @ ele_arr for ele_arr in lattice
            ]
            lattice_overlap = PubMeth.overlap_points(lattice_p, lattice_m)

            K0_arr = array([a0_constant / sqrt(3), 0])
            Kp_plus_L_p = [
                PubMeth.rotation(comm_ang / 2) @ K0_arr + ele_arr
                for ele_arr in lattice_p
            ]  # actually, K1 is the vertex of the hexagon.
            Km_plus_L_m = [
                PubMeth.rotation(-comm_ang / 2) @ K0_arr + ele_arr
                for ele_arr in lattice_m
            ]
            K_overlap = PubMeth.overlap_points(Kp_plus_L_p, Km_plus_L_m)
            print("Len of K overlap is: ", len(K_overlap))

            atom_lattice_plus_set = PubMeth.plus_set(
                lattice_p, lattice_m, length_limit=4.5
            )
            tmp_plus_set = [
                array([round(ele_arr[0], 4), round(ele_arr[1], 4)])
                for ele_arr in atom_lattice_plus_set
            ]
            atom_lattice_plus_set = np.unique(array(tmp_plus_set), axis=0)

            if moire_figsave:
                PubMeth.draw_dots(
                    lattice_m, marker_type="o", save_or_not=False, hold_on=True
                )
                PubMeth.draw_dots(
                    lattice_p, marker_type="o", save_or_not=False, hold_on=True
                )
                PubMeth.draw_dots(
                    lattice_overlap,
                    marker_type="*",
                    save_or_not=True,
                    hold_on=False,
                    fig_title=r"$\theta = %.2f \degree$" % comm_ang,
                    leg_list=[r"$L_{-}$", r"$L_{+}$", r"$L_{-} \cap L_{+}$"],
                )
            if moire_pm_set_save:
                PubMeth.draw_dots(
                    lattice_m, marker_type="o", save_or_not=False, hold_on=True
                )
                PubMeth.draw_dots(
                    lattice_p, marker_type="o", save_or_not=False, hold_on=True
                )
                PubMeth.draw_dots(
                    atom_lattice_plus_set,
                    marker_type=".",
                    save_or_not=True,
                    hold_on=False,
                    draw_width=0.1,
                    fig_title=r"$\theta = %.2f \degree$" % comm_ang,
                    leg_list=[r"$L_{-}$", r"$L_{+}$", r"$L_{-} + L_{+}$"],
                )
            if k_valley_save:
                PubMeth.draw_dots(
                    Kp_plus_L_p,
                    marker_type="o",
                    save_or_not=False,
                    hold_on=True,
                    draw_width=0.1,
                )
                PubMeth.draw_dots(
                    Km_plus_L_m,
                    marker_type="o",
                    save_or_not=False,
                    hold_on=True,
                    draw_width=0.1,
                )
                if len(K_overlap) != 0:
                    PubMeth.draw_dots(
                        K_overlap,
                        marker_type="*",
                        save_or_not=False,
                        hold_on=True,
                        fig_title=r"$\theta = %.2f \degree$" % comm_ang,
                    )
                PubMeth.draw_dots(
                    atom_lattice_plus_set,
                    marker_type=".",
                    save_or_not=False,
                    hold_on=True,
                    draw_width=0.1,
                )
                plt.plot([0, sqrt(N) * K0_arr[0]], [0, sqrt(N) * K0_arr[1]], "r")
                plt.plot([0, -sqrt(N) * K0_arr[0]], [0, -sqrt(N) * K0_arr[1]], "k")
                plt.scatter(0, 0, marker="8")
                plt.title(r"K space: $\theta = %.2f \degree$" % comm_ang)
                plt.legend(
                    [
                        r"$K_{-}^{0} + P_{-}$",
                        r"$K_{+}^{0} + P_{+}$",
                        r"$\mathcal{Q}_+$",
                        r"$\sqrt{N}K$",
                        r"$-\sqrt{N}K$",
                    ]
                )
                plt.savefig(
                    "./tmp_figs/k_valley_{}".format(comm_ang) + fig_format, dpi=330
                )

    @staticmethod
    def situate_x_labels(input_path_list, density_per_path):
        dis_list = [0]
        for i in range(0, len(input_path_list) - 1):
            p1 = array(input_path_list[i])
            p2 = array(input_path_list[i + 1])
            dis_list.append(int(norm(p1 - p2) * density_per_path) + dis_list[-1])
        return dis_list

    @staticmethod
    def operator_d_theta(delta_angle):
        delta_theta = delta_angle / 180 * pi
        return 2 * sin(delta_theta / 2) * PubMeth.rotation(90)

    @staticmethod
    def pauli_mat(mat_index):
        sigma_0 = array([[1, 0], [0, 1]])
        sigma_x = array([[0, 1], [1, 0]])
        sigma_y = array([[0, -1j], [1j, 0]])
        sigma_z = array([[1, 0], [0, -1]])
        Pauli_mat_list = [sigma_0, sigma_x, sigma_y, sigma_z]
        return Pauli_mat_list[mat_index]

    @staticmethod
    def path_between_two_vec(arr_1, arr_2, density=100, include_last=False):
        arr_list = []
        for i_add in range(int(density)):
            tmp_arr = arr_1 + (arr_2 - arr_1) / density * i_add
            arr_list.append(tmp_arr)
        if include_last:
            arr_list.append(arr_2)
        return arr_list

    @staticmethod
    def dots_around_one_point(center_vec, add_vec, density=100):
        out_arr_list = []
        add_another_dir = PubMeth.rotation(90) @ add_vec
        for i in linspace(-1, 1, density + 1):
            for j in linspace(-1, 1, density + 1):
                out_arr_list.append(center_vec + i * add_vec + j * add_another_dir)
        return out_arr_list

    @staticmethod
    def plotly_layout(
        xlabel="X", ylabel="Y", zlabel="Z", figuretitle="title", width=800, height=600
    ):
        layout = go.Layout(
            title=figuretitle,
            # width=width,
            # height=height,
            scene=dict(
                xaxis_title=xlabel,
                yaxis_title=ylabel,
                zaxis_title=zlabel,
                xaxis=dict(
                    tickfont=dict(
                        size=14,
                        family="Old Standard TT, serif",
                    )
                ),
                yaxis=dict(
                    tickfont=dict(
                        size=14,
                        family="Old Standard TT, serif",
                    )
                ),
                zaxis=dict(
                    tickfont=dict(
                        size=14,
                        family="Old Standard TT, serif",
                    )
                ),
            ),
        )
        return layout

    @staticmethod
    def find_half_filling_energy(energies_list_in):
        dim_one_list = sorted(
            np.matrix.tolist(real(array(energies_list_in).reshape(1, -1)))[0]
        )
        if len(dim_one_list) % 2 == 0:
            print("# of eigen energies is even.")
            return (
                dim_one_list[len(dim_one_list) // 2]
                + dim_one_list[len(dim_one_list) // 2 - 1]
            ) / 2
        elif len(dim_one_list) % 2 != 0:
            print("# of eigen energies is odd.")
            return dim_one_list[len(dim_one_list) // 2]

    @classmethod
    def plot_energies(
        cls,
        energies_list_in,
        fig_size=(7, 5),
        y_range=[],
        line_type="k-",
        figuretitle="title",
        x_label_pos=[],
        x_labs=[],
        save_fig_or_not=True,
        show_or_not=False,
        save_fig_npy_title="band",
        lw=2,
        hold_on=False,
        y_label="E(meV)",
        test_mode=False,
        selected_bds_indices=[],
        fig_format=global_fig_format,
        html_name="eigen_energies",
        save_npy_or_not=False,
        energy_line=[],
        save_fig_dir_name="bands",
        save_npy_dir_name="bands_data",
        ax_input=False,
        y_ticks=arange(-2000, 2001, 500),
        title_font_size=16,
        label_font_size=16,
        tick_font_size=14,
    ):
        if not isinstance(ax_input, bool):
            if save_npy_or_not:
                target_data_dir = PubMeth.get_right_save_path_and_create(
                    save_npy_dir_name, data_files_or_not=True
                )
                if not os.path.exists(target_data_dir):
                    os.makedirs(target_data_dir)
                save(target_data_dir + save_fig_npy_title + ".npy", energies_list_in)

            if len(selected_bds_indices) == 0:
                ax_input.plot(real(energies_list_in), line_type, linewidth=lw)
            else:
                selected_energies_list = [
                    real(energies_list_in)[:, bd_i] for bd_i in selected_bds_indices
                ]
                ax_input.plot(array(selected_energies_list).T, line_type, linewidth=lw)
            if len(energy_line) != 0:
                ax_input.plot(
                    np.kron(
                        ones(len(energies_list_in)).reshape(len(energies_list_in), 1),
                        energy_line,
                    ),
                    "r--",
                    linewidth=lw,
                )
            ax_input.set_ylim(y_range)
            ax_input.set_title(figuretitle, fontsize=title_font_size)
            ax_input.set_ylabel(y_label, fontsize=label_font_size)
            ax_input.set_xticks(x_label_pos, x_labs, fontsize=label_font_size)
            # ##  Tick font size

            ax_input.yaxis.set_minor_locator(tck.AutoMinorLocator())
            ax_input.tick_params(
                axis="y", which="major", length=5, labelsize=tick_font_size
            )
            ax_input.tick_params(axis="y", which="minor", length=2)

        else:
            fig, ax = plt.subplots(1, 1, figsize=fig_size, dpi=330)
            plt.figure(figsize=fig_size, dpi=330)
            if len(selected_bds_indices) == 0:
                ax.plot(real(energies_list_in), line_type, linewidth=lw)
            else:
                selected_energies_list = [
                    real(energies_list_in)[:, bd_i] for bd_i in selected_bds_indices
                ]
                ax.plot(array(selected_energies_list).T, line_type, linewidth=lw)
            if len(energy_line) != 0:
                ax.plot(
                    np.kron(
                        ones(len(energies_list_in)).reshape(len(energies_list_in), 1),
                        energy_line,
                    ),
                    "r--",
                    linewidth=lw,
                )
            if len(y_range) != 0:
                ax.set_ylim(y_range)
            ax.set_title(figuretitle, fontsize=title_font_size)
            ax.set_ylabel(y_label, fontsize=label_font_size)
            ax.set_xticks(x_label_pos, x_labs, fontsize=label_font_size)
            ax.set_xlim([0, x_label_pos[-1]])
            if len(y_ticks) != 0:
                ax.set_yticks(y_ticks, y_ticks)
                ax.yaxis.set_minor_locator(tck.AutoMinorLocator())
                ax.tick_params(
                    axis="y", which="major", length=5, labelsize=tick_font_size
                )
                ax.tick_params(axis="y", which="minor", length=2)
            if save_fig_or_not:
                target_fig_dir = PubMeth.get_right_save_path_and_create(
                    save_fig_dir_name, data_files_or_not=True
                )
                if not os.path.exists(target_fig_dir):
                    os.makedirs(target_fig_dir)
                fig.savefig(target_fig_dir + save_fig_npy_title + ".pdf", dpi=330)
                fig.savefig(target_fig_dir + save_fig_npy_title + ".png", dpi=330)
            if save_npy_or_not:
                target_data_dir = PubMeth.get_right_save_path_and_create(
                    save_npy_dir_name, data_files_or_not=True
                )
                if not os.path.exists(target_data_dir):
                    os.makedirs(target_data_dir)
                save(target_data_dir + save_fig_npy_title + ".npy", energies_list_in)
            if not hold_on:
                if show_or_not:
                    plt.show()
                plt.close()
            if test_mode:
                trace_list = []
                layout = PubMeth.plotly_layout()
                if len(selected_bds_indices) == 0:
                    for i in range(array(energies_list_in).shape[1]):
                        ele_trace = go.Scatter(
                            x=arange(array(energies_list_in).shape[0]),
                            y=array(energies_list_in)[:, i],
                        )
                        trace_list.append(ele_trace)
                    fig = go.Figure(data=trace_list, layout=layout)
                    html_dir = PubMeth.get_right_save_path_and_create(
                        "html_files", data_files_or_not=True
                    )
                    if not os.path.exists(html_dir):
                        os.makedirs(html_dir)
                    fig.write_html(html_dir + html_name + ".html")
                else:
                    for ele_i in selected_bds_indices:
                        ele_trace = go.Scatter(
                            x=arange(array(energies_list_in).shape[0]),
                            y=array(energies_list_in)[:, ele_i],
                        )
                        trace_list.append(ele_trace)
                    fig = go.Figure(data=trace_list, layout=layout)
                    html_dir = PubMeth.get_right_save_path_and_create(
                        "html_files", data_files_or_not=True
                    )
                    if not os.path.exists(html_dir):
                        os.makedirs(html_dir)
                    fig.write_html(html_dir + html_name + ".html")

    @staticmethod
    def sigma_angle_dot_p(p_arr, angle_in):
        theta_in = angle_in / 180 * pi
        p_x = p_arr[0]
        p_y = p_arr[1]
        return array(
            [
                [0, exp(-1j * theta_in) * (p_x - 1j * p_y)],
                [exp(1j * theta_in) * (p_x + 1j * p_y), 0],
            ]
        )

    @staticmethod
    def save_if_same_name(func_save, save_name, target_dir="./", splitter="_"):
        full_save_name = target_dir + save_name
        if os.path.exists(target_dir + save_name):
            flag_i = 1
            splitted_elements = full_save_name.split(splitter)
            splitted_elements[-1] = "#1" + splitted_elements[-1]
            while True:
                if os.path.exists("_".join(splitted_elements)):
                    flag_i = flag_i + 1
                    splitted_elements[-1] = (
                        "#{:.0f}".format(flag_i) + splitted_elements[-1][2:]
                    )
                else:
                    break
            func_save("_".join(splitted_elements))
        else:
            func_save(target_dir + save_name)

    @staticmethod
    def t_to_fermi_vel(t_input, a0=1.42 * sqrt(3)):
        return sqrt(3) / 2 * a0 * t_input / (h_bar_eV * eV2meV * m2A)

    @staticmethod
    def dots_in_plane(points_list, dots_density=70):
        """
        :param args_list: [[first point, second point, third point]]
        :return:
        """
        first_p = points_list[0]
        second_p = points_list[1]
        third_p = points_list[2]
        delta_x_arr = (second_p - first_p) / dots_density
        delta_y_arr = (third_p - first_p) / dots_density
        dots_arr_list = []
        for ele_m in range(dots_density):
            for ele_n in range(dots_density):
                dots_arr_list.append(delta_x_arr * ele_m + delta_y_arr * ele_n)
        return dots_arr_list

    @staticmethod
    def add_right_cax(ax, pad, width):  # ax is the axe of figure
        axpos = ax.get_position()
        caxpos = mtransforms.Bbox.from_extents(
            axpos.x1 + pad, axpos.y0, axpos.x1 + pad + width, axpos.y1
        )
        cax = ax.figure.add_axes(caxpos)

        return cax

    @staticmethod
    def add_down_cax(ax, pad, width):  # ax is the axe of figure
        axpos = ax.get_position()
        caxpos = mtransforms.Bbox.from_extents(
            axpos.x0, axpos.y0 - pad - width, axpos.x1, axpos.y0 - pad
        )
        cax = ax.figure.add_axes(caxpos)

        return cax

    @staticmethod
    def add_up_cax(ax, pad, width):  # ax is the axe of figure
        axpos = ax.get_position()
        caxpos = mtransforms.Bbox.from_extents(
            axpos.x0, axpos.y1 + pad, axpos.x1, axpos.y1 + pad + width
        )
        cax = ax.figure.add_axes(caxpos)

        return cax

    @staticmethod
    def add_color_bar(
        ax,
        im_var,
        pad,
        width,
        position,
        orientation,
        labelsize=16,
        cbar_label="LABEL",
        label_font_size=16,
        label_pad=10,
    ):
        """
        position: up, down, right
        orientation: horizontal, vertical
        """
        ##  define the ax
        if position == "right":
            c_ax = PubMeth.add_right_cax(ax, pad, width)
        elif position == "up":
            c_ax = PubMeth.add_up_cax(ax, pad, width)
        elif position == "down":
            c_ax = PubMeth.add_down_cax(ax, pad, width)

        ##  Plot the colorbar based on image variable
        cbar = plt.colorbar(im_var, cax=c_ax, orientation=orientation)
        cbar.ax.tick_params(labelsize=labelsize)
        cbar.set_label(cbar_label, fontsize=label_font_size, labelpad=label_pad)

    @staticmethod
    def list2set_within_diff(list_input, tolerance=0.001, accurate_digits=4):
        out_list = []
        out_list.append(round(list_input[0], accurate_digits))
        tmp_arr = array(list_input)
        while len(tmp_arr) != 0:
            judge = (tmp_arr - out_list[-1] * ones(len(tmp_arr))) > tolerance
            tmp_arr = tmp_arr[judge]
            if len(tmp_arr) != 0:
                out_list.append(round(tmp_arr[0], 4))
        return out_list

    @staticmethod
    def rotation_3d_to_2d(angle, angle_or_not=True):
        if angle_or_not:
            theta = angle / 180 * pi
            return array(
                [[cos(theta), -sin(theta), 0], [sin(theta), cos(theta), 0], [0, 0, 1]]
            )
        else:
            theta = angle
            return array(
                [[cos(theta), -sin(theta), 0], [sin(theta), cos(theta), 0], [0, 0, 1]]
            )

    @staticmethod
    def get_angle_between_two_vecs(arr_1, arr_2, return_angle_or_not=True):
        tmp_theta_cos = arr_1 @ arr_2 / (norm(arr_1) * norm(arr_2))
        if return_angle_or_not:
            return np.arccos(tmp_theta_cos) / pi * 180
        else:
            return np.arccos(tmp_theta_cos)

    @staticmethod
    def get_angle_to_x_axis(arr_in, return_angle_or_not=True, dim3=False):
        if dim3:
            theta_or_angle = PubMeth.get_angle_between_two_vecs(
                arr_in, array([1, 0, arr_in[-1]]), return_angle_or_not
            )
        else:
            theta_or_angle = PubMeth.get_angle_between_two_vecs(
                arr_in, array([1, 0]), return_angle_or_not
            )
        if arr_in[-1] <= 0:
            return -theta_or_angle
        else:
            return theta_or_angle

    @staticmethod
    def reflection_to_y_axis():
        return array([[1, 0, 0], [0, -1, 0], [0, 0, 1]])

    @staticmethod
    def rotate_shift_line(point1, point2, rotation_point, rotation_angle, b_shift=0):
        k = (point2[1] - point1[1]) / (point2[0] - point1[0])  # get the incline
        b = point2[1] - k * point2[0]  # get the intercept

        x0 = rotation_point[0]
        y0 = rotation_point[1]
        rotation_theta = rotation_angle / 180 * pi
        k_prime = (k * cos(rotation_theta) + sin(rotation_theta)) / (
            cos(rotation_theta) - k * sin(rotation_theta)
        )  # get the k of rotated line
        b_prime = (
            (k * x0 - y0) * (1 - cos(rotation_theta))
            - (x0 + k * y0) * sin(rotation_theta)
            + b
        ) / (
            cos(rotation_theta) - k * sin(rotation_theta)
        )  # get the intercept of the rotated line

        def rotated_func(x):  # get the function of the new line
            return k_prime * x + b_prime + b_shift

        return rotated_func  # return it

    @staticmethod
    def uniform_diamond_bz(kp_num=70):
        kp_list_f = []
        b_p_arr = array([sqrt(3) / 2, 3 / 2])
        b_n_arr = array([-sqrt(3) / 2, 3 / 2])
        for ele_m in arange(0, 1, 1 / kp_num):
            for ele_n in arange(0, 1, 1 / kp_num):
                k_p = ele_m * b_p_arr + ele_n * b_n_arr
                kp_list_f.append(k_p)
        return kp_list_f

    @staticmethod
    def read_xlsx_data(file_path, exclude_rows_num=1):
        data = openpyxl.load_workbook(file_path)
        column_tag = [chr(i).capitalize() for i in range(97, 123)]
        sheet = data[data.sheetnames[0]]
        column_to_read = sheet.max_column
        max_row_num = sheet.max_row
        ele_is_column_list = []
        for column_i in range(column_to_read):
            ele_column_cell = sheet[
                "{0}{2}:{0}{1}".format(
                    column_tag[column_i], max_row_num, 1 + exclude_rows_num
                )
            ]
            ele_cell_dat = [
                ele_column_cell[ele_i][0].value
                for ele_i in range(max_row_num - exclude_rows_num)
            ]
            ele_is_column_list.append(ele_cell_dat)
        return ele_is_column_list

    @staticmethod
    def cut_list_within_range(lst_in, min_in, max_in):
        """
        lst_in: List that we want to cut
        min: minimum of the cut
        max: maximum of the cut
        return: arr, min_index, max_index
        """
        lst_in = list(lst_in)
        arr = array(lst_in)
        arr = arr[arr >= min_in]
        arr = arr[arr <= max_in]
        min_index = lst_in.index(min(arr))
        max_index = lst_in.index(max(arr))
        return arr, min_index, max_index

    @staticmethod
    def get_interped1d_y(f_interp, input_x, target_x_range, inter_type="linear"):
        """
        # f_interp
        The fit function

        # input_x
        The input x range to get the fit function

        # target_x_range
        The range of x we want to investigate

        # inter_type:
        linear, nearest, zero, quadratic, cubic

        # return

        """
        ##  cut the useful range of the x list which fit the input x
        x_array = array(target_x_range)
        x_array = x_array[x_array < max(input_x)]
        x_array = x_array[x_array > min(input_x)]

        return f_interp(x_array)

    @staticmethod
    def get_nearest_x_index(
        input_x_data, input_target
    ):  # get the index of list which is closest to input data
        tmp_x_data_mod = abs(array(input_x_data) - input_target)
        min_data_index = list(tmp_x_data_mod).index(min(tmp_x_data_mod))
        return min_data_index

    @staticmethod
    def cut_list(lst_in, index_1, index_2):
        if index_1 > index_2:
            return lst_in[index_2 : index_1 + 1]
        elif index_2 > index_1:
            return lst_in[index_1 : index_2 + 1]

    @staticmethod
    def twisted_graphene_super_cell_area(angle_in):
        theta_in = angle_in / 180 * pi
        a0_constant = 1.42 * sqrt(3)  #   A
        super_lattice = a0_constant / (2 * sin(theta_in / 2))
        super_cell_area = sqrt(3) / 2 * super_lattice**2

        return super_cell_area

    @staticmethod
    def line_dot_track_animation(
        x_list,
        y_list,
        frames_num,
        xlabel="X",
        ylabel="Y",
        title="Title",
        file_name="demo_line_dot",
        interval_num=100,
        fps_num=60,
    ):
        fig, ax = plt.subplots()

        def update_points(num_i):
            ax.cla()
            ax.plot(x_list, y_list)
            ax.scatter([x_list[num_i]], [y_list[num_i]])
            ax.set_aspect("auto")
            ax.set_xlabel(xlabel, fontsize=12)
            ax.set_ylabel(ylabel, fontsize=12)
            ax.set_title(title, fontsize=14)
            ax.set_xlim(ax.get_xlim())
            ax.set_ylim(ax.get_ylim())

        ani = animation.FuncAnimation(
            fig, update_points, frames=frames_num, interval=interval_num
        )
        ani.save(
            "/home/aoxv/code/Data/animation/line_1d/{}.mp4".format(file_name),
            fps=fps_num,
        )

    @staticmethod
    def convert_e_to_omega(energy_in):
        """
        Convert the energy of photon to omega
        # energy_list
        the unit should be meV
        """
        omega_list = array(energy_in) / (h_bar_eV * eV2meV)

        return omega_list

    @staticmethod
    def line_evolve_animation(
        x_lists,
        y_lists,
        frames_num,
        xlabel="X",
        ylabel="Y",
        title="Title",
        file_name="demo_line_evolve",
        interval_num=10,
        fps_num=60,
    ):
        fig, ax = plt.subplots()

        def update_points(num_i):
            ax.cla()
            ax.plot(x_lists[num_i], y_lists[num_i])
            ax.set_aspect("auto")
            if isinstance(xlabel, list):
                ax.set_xlabel([num_i], fontsize=12)
            else:
                ax.set_xlabel(xlabel, fontsize=12)
            if isinstance(ylabel, list):
                ax.set_ylabel(ylabel[num_i], fontsize=12)
            else:
                ax.set_ylabel(ylabel, fontsize=12)
            if isinstance(title, list):
                ax.set_title(title[num_i], fontsize=14)
            else:
                ax.set_title(title, fontsize=14)
            ax.set_xlim(ax.get_xlim())
            ax.set_ylim(ax.get_ylim())

        ani = animation.FuncAnimation(
            fig, update_points, frames=frames_num, interval=interval_num
        )
        ani.save(
            "/home/aoxv/code/Data/animation/line_1d/{}.mp4".format(file_name),
            fps=fps_num,
        )

        print("Complete file: ", file_name)

        return

    @staticmethod
    def line_evolve_animation_in_a_row(
        xs_lists,
        ys_lists,
        frames_num,
        xlabels_list,
        ylabels_list,
        titles_list,
        file_name="demo_multi_col_line_evolve",
        interval_num=10,
        fps_num=60,
        fig_height=4,
    ):
        fig, ax_list = plt.subplots(
            nrows=1,
            ncols=len(xs_lists),
            figsize=(fig_height * len(xs_lists), fig_height),
        )

        def update_points(num_i):
            ## draw all the plots in a row
            for ele_ax_i in range(len(xs_lists)):
                ele_ax = ax_list[ele_ax_i]

                ele_x_list = xs_lists[ele_ax_i]
                ele_y_list = ys_lists[ele_ax_i]

                ele_x_label = xlabels_list[ele_ax_i]
                ele_y_label = ylabels_list[ele_ax_i]

                ele_title = titles_list[ele_ax_i]

                ele_ax.cla()
                ele_ax.plot(ele_x_list[num_i], ele_y_list[num_i])

                if isinstance(ele_x_label, list):
                    ele_ax.set_xlabel([num_i], fontsize=12)
                else:
                    ele_ax.set_xlabel(ele_x_label, fontsize=12)
                if isinstance(ele_y_label, list):
                    ele_ax.set_ylabel(ele_y_label[num_i], fontsize=12)
                else:
                    ele_ax.set_ylabel(ele_y_label, fontsize=12)
                if isinstance(ele_title, list):
                    ele_ax.set_title(ele_title[num_i], fontsize=14)
                else:
                    ele_ax.set_title(ele_title, fontsize=14)

                ele_ax.set_xlim(ele_ax.get_xlim())
                ele_ax.set_ylim(ele_ax.get_ylim())

        ani = animation.FuncAnimation(
            fig, update_points, frames=frames_num, interval=interval_num
        )
        ani.save(
            "/home/aoxv/code/Data/animation/line_1d/{}.mp4".format(file_name),
            fps=fps_num,
        )

        print("Complete file: ", file_name)

        return

    @staticmethod
    def line_3d_dot_track_animation(
        x_list,
        y_list,
        z_list,
        frames_num,
        xlabel="X",
        ylabel="Y",
        zlabel="Z",
        title="Title",
        file_name="demo_line_3d_dot",
        interval_num=100,
        fps_num=60,
    ):
        """
        Plot the line in 3D dimension
        """
        fig = plt.figure()
        ax = plt.axes(projection="3d")

        def update_points(num_i):
            ax.cla()
            ax.plot3D(x_list, y_list, z_list)
            ax.scatter([x_list[num_i]], [y_list[num_i]], [z_list[num_i]])
            ax.set_aspect("auto")
            ax.set_xlabel(xlabel, fontsize=12)
            ax.set_ylabel(ylabel, fontsize=12)
            ax.set_zlabel(zlabel, fontsize=12)
            ax.set_title(title, fontsize=14)

        ani = animation.FuncAnimation(
            fig, update_points, frames=frames_num, interval=interval_num
        )
        ani.save(
            "/home/aoxv/code/Data/animation/line_1d/{}.mp4".format(file_name),
            fps=fps_num,
        )

        print("Complete: ", file_name)

        return

    @staticmethod
    def kps_in_moire(kp_num):
        """
        return: k point list
        """
        kp_list_f = []
        b_p_arr = array([sqrt(3) / 2, 3 / 2])
        b_n_arr = array([-sqrt(3) / 2, 3 / 2])
        for ele_m in arange(0, 1, 1 / kp_num):
            for ele_n in arange(0, 1, 1 / kp_num):
                k_p = ele_m * b_p_arr + ele_n * b_n_arr
                kp_list_f.append(k_p)
        return kp_list_f

    @staticmethod
    def lorentz_real(var_list, amp_ratio, center, gamma):
        """
        return the real part of the Lorentzian peak
        ##  amp_ratio
        amp_ratio determines the maximum of the real part, range is [0, 1]
        """
        ratio = sqrt(gamma / center) * center
        amp = ratio * amp_ratio

        numerator = amp**2 * (center**2 - var_list**2)

        denominator_term1 = (center**2 - var_list**2) ** 2
        denominator_term2 = gamma**2 * var_list**2
        denominator = denominator_term1 + denominator_term2

        output_term = numerator / denominator

        return output_term

    @staticmethod
    def lorentz_imag(var_list, amp_ratio, center, gamma):
        """
        return the imaginary part of the Lorentzian peak
        ##  amp_ratio
        amp_ratio determines the maximum of the real part, range is [0, 1]
        """
        ratio = sqrt(gamma / center) * center
        amp = ratio * amp_ratio

        numerator = amp**2 * gamma * var_list

        denominator_term1 = (center**2 - var_list**2) ** 2
        denominator_term2 = gamma**2 * var_list**2
        denominator = denominator_term1 + denominator_term2

        output_term = numerator / denominator

        return output_term

    @staticmethod
    def lorentz_full(x_list, center, amp_ratio, gamma, timesi=False, pars_type="value"):
        """
        return the full Lorentzian peak (both real and imaginary part)
        #   pars_type
        "value" or "list"
        """
        if isinstance(center, (float, int)):
            real_part = PubMeth.lorentz_real(x_list, amp_ratio, center, gamma)
            imag_part = PubMeth.lorentz_imag(x_list, amp_ratio, center, gamma)

        elif isinstance(center, (list, np.ndarray)):
            real_part = [
                PubMeth.lorentz_real(
                    x_list, amp_ratio[ele_i], center[ele_i], gamma[ele_i]
                )
                for ele_i in range(len(amp_ratio))
            ]
            imag_part = [
                PubMeth.lorentz_imag(
                    x_list, amp_ratio[ele_i], center[ele_i], gamma[ele_i]
                )
                for ele_i in range(len(amp_ratio))
            ]

            real_part = sum(real_part)
            imag_part = sum(imag_part)

        output_term = real_part + 1j * imag_part

        if timesi:
            output_term = -1j * output_term

            return output_term
        else:
            return output_term

    @staticmethod
    def float_or_int(input_exp):
        """
        To see whether input expression is float or integer
        """
        output_bool = isinstance(input_exp, float) or isinstance(input_exp, int)

        return output_bool

    @staticmethod
    def array_or_list(input_exp):
        """
        To see whether the input expression is list or array
        """
        output_bool = isinstance(input_exp, list) or isinstance(input_exp, np.ndarray)

        return output_bool

    @staticmethod
    def diamond_around_o_2d(two_primitive_arrs, shell_num=3):
        """
        Create diamond shells around zero point

        Input: [primitive_vec1, primitive_vec2]

        """
        pri_vec1, pri_vec2 = two_primitive_arrs

        ##  Express the vectors along plus and minus directions
        vec_plus = pri_vec1 + pri_vec2

        ##  Define the starting point
        start_point = -(2 * (shell_num - 1) + 1) / 2 * vec_plus

        ##  Total length of one side
        side_v1 = (2 * (shell_num - 1) + 1) * pri_vec1
        side_v2 = (2 * (shell_num - 1) + 1) * pri_vec2

        ##  Two for-list
        m_list = linspace(0, 1, 2 * shell_num)
        n_list = linspace(0, 1, 2 * shell_num)

        ##  Get the array list by two for-list
        arr_list = []
        for ele_m in m_list:
            for ele_n in n_list:
                arr_list.append(start_point + ele_m * side_v1 + ele_n * side_v2)

        return arr_list

    @staticmethod
    def diamond_around_lattice_2d(two_primitive_arrs, dots_density=3):
        """
        Create diamond lattices around one lattice

        Input dots density should be odd number
        """
        ##  Output array list
        arr_list = []
        ##  Get two vectors
        vec_1 = two_primitive_arrs[0]
        vec_2 = two_primitive_arrs[1]
        ##  Get odd number
        if dots_density % 2 == 0:
            dots_density = dots_density + 1
        ##  Get multiple lattices
        vecs1_part = [ele_coeff * vec_1 for ele_coeff in arange(0, dots_density)]
        vecs2_part = [ele_coeff * vec_2 for ele_coeff in arange(0, dots_density)]

        arr_list = PubMeth.vec_lists_for_loop_plus_minus(
            vecs1_part, vecs2_part
        ).reshape((-1, 2))

        # for ele_m in arange(0, dots_density):
        #     for ele_n in arange(0, dots_density):
        #         ele_vec = ele_m * vec_1 + ele_n * vec_2
        #         arr_list.append(ele_vec)

        ##  Express the translation vector
        factor = dots_density - 1
        trans_vec = (vec_1 + vec_2) * factor / 2
        arr_list = array(arr_list) - trans_vec

        return arr_list

    @staticmethod
    def hexagon_around_o_2d(
        a0_lattice, shell_num=3, rot_angle=0, translation_arr=array([0, 0])
    ):
        """
        Create two sets of lattices which represent two lattices of graphene

        If you want to shift all the lattices, it will rotate all the lattices first and then shift all the lattices
        """
        a1 = array([1 / 2, sqrt(3) / 2]) * a0_lattice
        a2 = array([-1 / 2, sqrt(3) / 2]) * a0_lattice
        a1_a2_plus = a1 + a2  #   Plus vector of a1 and a2
        a1_a2_minu = a1 - a2  #   Minus vector of a1 and a2

        ##  Get the dots number list from up to bottom
        first_part = arange(shell_num, 2 * shell_num)  #   ascending
        second_par = arange(2 * shell_num, shell_num, -1)  #   descending
        dots_num_list = list(first_part) + list(second_par)  #   dots number list

        ##  Get the difference of dots number in each row
        row_index_arr = shell_num - arange(2 * shell_num)

        ##  Get the row index mod 2 and divided by 2 results
        row_i_div_2 = row_index_arr // 2  ##  To get the coefficient
        row_i_mod_2 = row_index_arr % 2  ##  To identify which row the dots are

        ##  Get the dots number of each row for mod 2 and divided by 2 results
        dots_num_div_2 = array(dots_num_list) // 2

        ##  Express the starting point
        lattices_1 = []
        for ele_i in range(len(dots_num_list)):
            judge_flag = row_i_mod_2[ele_i]  #   To identify which row the dots are
            ele_row_i = row_index_arr[ele_i]
            shift_i = row_i_div_2[ele_i]

            if judge_flag:  ##  Get the odd row
                ##  define the starting point
                ele_start_p = (shift_i + 1 / 3) * a1_a2_plus

                ##  get the dots on the left and right side of the starting point (including start point or not)
                row_dots_num_div_2 = dots_num_div_2[ele_i]
                shift_scale_arr = arange(-row_dots_num_div_2, row_dots_num_div_2 + 1)

                ##  Get all the points on the odd row
                points_on_row = ele_start_p + np.kron(
                    shift_scale_arr.reshape((-1, 1)), a1_a2_minu
                )

                ##  Output all the points on a row
                lattices_1.append(points_on_row)
            elif ele_row_i == 0:  ##  means it is the first row below the zero point
                ##  define the starting point
                ele_start_p = -1 / 6 * a1_a2_plus

                ##  The step of every scale factor
                row_dots_num_div_2 = dots_num_div_2[ele_i]
                shift_scale_arr = arange(
                    -(2 * row_dots_num_div_2 - 1) / 2,
                    (2 * row_dots_num_div_2 - 1) / 2 + 1,
                )

                ##  Get all the points on the odd row
                points_on_row = ele_start_p + np.kron(
                    shift_scale_arr.reshape((-1, 1)), a1_a2_minu
                )

                ##  Output all the points on a row
                lattices_1.append(points_on_row)
            elif ele_row_i != 0:  ##  means it is the even row
                ele_start_p = (shift_i - 1 / 6) * a1_a2_plus

                ##  The step of every scale factor
                row_dots_num_div_2 = dots_num_div_2[ele_i]
                shift_scale_arr = arange(
                    -(2 * row_dots_num_div_2 - 1) / 2,
                    (2 * row_dots_num_div_2 - 1) / 2 + 1,
                )

                ##  Get all the points on the odd row
                points_on_row = ele_start_p + np.kron(
                    shift_scale_arr.reshape((-1, 1)), a1_a2_minu
                )

                ##  Output all the points on a row
                lattices_1.append(points_on_row)

        lattices_1 = np.vstack(tuple(lattices_1))
        lattices_2 = np.copy(lattices_1)
        ##  Get another set of lattice
        lattices_2[:, 1] = -lattices_2[:, 1]

        all_lattices = np.vstack((lattices_1, lattices_2))

        ## rotate all the lattices
        rot_mat = PubMeth.rotation(rot_angle)
        rot_lattices = rot_mat @ all_lattices.T
        rot_lattices = rot_lattices.T

        ##  lattices after shift
        shift_lattices = rot_lattices + translation_arr

        return shift_lattices

    @staticmethod
    def vec_lists_for_loop_plus_minus(vec_list_1, vec_list_2, operation="plus"):
        """
        For-Loop to plus or minus the vectors (2D vectors) in two lists.

        Usually, as for the output matrix, the row is vector list 1 and the column is vector list 2.

        ##  Operation
        "plus" or "minus" or "minus reverse"
        """
        ##  Array-ize these vector lists
        vec_list_1 = array(vec_list_1)
        vec_list_2 = array(vec_list_2)

        ##  Length of two lists
        len_list_1 = len(vec_list_1)
        len_list_2 = len(vec_list_2)

        ##  Two one lists corresponding to two vector lists
        one_list_1 = np.ones(len_list_2)
        one_list_2 = np.ones(len_list_1).reshape((-1, 1))

        ##  Get the matrix form of two vectors
        mat_1 = np.kron(one_list_1, vec_list_1)
        mat_2 = np.kron(vec_list_2.reshape((1, -1)), one_list_2)

        ##  Get the plus or minus result of these two matrices
        if operation == "plus":
            output_mat = mat_1 + mat_2
        elif operation == "minus":
            output_mat = mat_1 - mat_2
        elif operation == "minus reverse":
            output_mat = mat_2 - mat_1

        return output_mat

    @staticmethod
    def remove_dup_2d_vec_in_a_list(vec_list, tolerance=0.01):
        """
        Remove the duplicate vectors in a vector list.
        """
        ##  Create a copy of vector list which is changing.
        changing_vec_list = list(vec_list)[:]

        ##  Define initial index
        ini_i = 0

        ##  Change the vector list and judge whether it is empty or not
        while ini_i < len(changing_vec_list[ini_i:]):
            ##  Get the chosen vector
            chosen_vec = changing_vec_list[ini_i]

            ##  Get the difference between chosen vector and the changing vector list (The chosen vector excluded)
            diff_mat = chosen_vec - changing_vec_list[ini_i + 1 :]

            ##  Get every norm of each difference vector
            diff_norm = norm(diff_mat, axis=1)

            ##  Judge whether there are difference norm equaling zero?
            judge_arr = diff_norm < tolerance

            ##  Create the indices list changing with the changing vector list. (Starting from the chosen vector index)
            indices_arr = arange(ini_i + 1, len(changing_vec_list))
            to_remove_i = indices_arr[judge_arr]

            to_remove_ele_i = list(2 * to_remove_i) + list(2 * to_remove_i + 1)
            if len(to_remove_i) > 0:
                ##  Remove the duplicate elements in the list
                changing_vec_list = np.delete(changing_vec_list, to_remove_ele_i)
            changing_vec_list = array(changing_vec_list).reshape((-1, 2))
            ini_i = ini_i + 1

        # print("all vector: \n", changing_vec_list)

        return changing_vec_list

    @staticmethod
    def remove_vecs_from_vec_list(vec_list, target_vectors, tolerance=0.05):
        """
        Remove the target vectors from vector list

        target vectors can be a vector matrix (vector list)
        """
        ##  Get the length of vector list and the indices list
        len_vec_list = len(vec_list)
        len_tar_vecs = len(target_vectors)
        indices_list = arange(len_vec_list)
        ##  The difference matrix between two vector list
        vec_diff_mat = PubMeth.vec_lists_for_loop_plus_minus(
            vec_list, target_vectors, operation="minus"
        )
        ##  Judge whether the norm is less than the tolerance or not
        diff_norm = norm(vec_diff_mat, axis=1)
        judge_arr = diff_norm < tolerance
        ##  To remove index
        remove_i_list = []
        ##  For-Loop to get the indices to remove
        for part_i in range(len(target_vectors)):
            ##  Select the part we want to judge
            ele_diff_norm = norm(vec_diff_mat[:, part_i * 2 : (part_i + 1) * 2], axis=1)
            ele_judge_arr = ele_diff_norm < tolerance

            tmp_remo_i = indices_list[ele_judge_arr]
            remo_pairs = list(2 * tmp_remo_i) + list(2 * tmp_remo_i + 1)
            # print("The removing pair is: ", remo_pairs)
            remove_i_list.append(remo_pairs)
        ##  reshape the remove indices list
        remove_i_list = [ele for ele_list in remove_i_list for ele in ele_list]

        ##  Remove the target vector in the vector list
        output_vec_list = np.delete(vec_list, remove_i_list)

        ##  Reshape the output vector list
        output_vec_list = output_vec_list.reshape((-1, 2))

        return output_vec_list

    @staticmethod
    def hexagon_around_lattice_2d(a0_lattice, shell_num=3, rot_angle=0):
        """
        Get the shell points around lattice
        """
        ##  Center point list
        center_list = [array([0, 0])]

        ##  Odd shell list and extension operation
        vector_len = a0_lattice / sqrt(3)
        vec_up = array([0, 1]) * vector_len
        vec_ld = array([-sqrt(3) / 2, -1 / 2]) * vector_len
        vec_rd = array([sqrt(3) / 2, -1 / 2]) * vector_len

        odd_shells = [[vec_up, vec_ld, vec_rd]]
        odd_ext_shell = array([vec_up, vec_ld, vec_rd])

        ##  Even shell list and extension operation
        vec_down = -vec_up
        vec_lu = -vec_rd
        vec_ru = -vec_ld

        eve_shells = []
        eve_ext_shell = array([vec_down, vec_lu, vec_ru])

        ##  Control whether we do even or odd extension
        eve_ext_flag = True

        while shell_num > 1:
            ##  control the flag
            shell_num = shell_num - 1

            ##  The shell points we have now
            lats_odd_shells = [ele for ele_list in odd_shells for ele in ele_list]
            lats_eve_shells = [ele for ele_list in eve_shells for ele in ele_list]

            existing_lats = center_list + lats_odd_shells + lats_eve_shells

            ##  Even extension or odd extension
            if eve_ext_flag:
                ##  Get the outer shell points. Shape (num_extension_points * 3(2-col-vec))
                outer_shell_ext_mat = PubMeth.vec_lists_for_loop_plus_minus(
                    odd_shells[-1], eve_ext_shell
                )

                ##  Reshape the extension points to (rows-2)
                outer_shell_ext_mat = outer_shell_ext_mat.reshape((-1, 2))

                ##  Remove the duplicate points in this vector list
                outer_shell_points = PubMeth.remove_dup_2d_vec_in_a_list(
                    outer_shell_ext_mat
                )

                ##  Remove the existing vectors
                outer_lat_points = PubMeth.remove_vecs_from_vec_list(
                    outer_shell_points, existing_lats
                )

                ##  Change the even shell lattice points
                eve_shells.append(outer_lat_points)

                ##  Swap the flag
                eve_ext_flag = False
            else:
                ##  Get the outer shell points. Shape (num_extension_points * 3(2-col-vec))
                outer_shell_ext_mat = PubMeth.vec_lists_for_loop_plus_minus(
                    eve_shells[-1], odd_ext_shell
                )

                ##  Reshape the extension points to (rows-2)
                outer_shell_ext_mat = outer_shell_ext_mat.reshape((-1, 2))

                ##  Remove the duplicate points in this vector list
                outer_shell_points = PubMeth.remove_dup_2d_vec_in_a_list(
                    outer_shell_ext_mat
                )

                ##  Remove the existing vectors
                outer_lat_points = PubMeth.remove_vecs_from_vec_list(
                    outer_shell_points, existing_lats
                )

                ##  Change the even shell lattice points
                odd_shells.append(outer_lat_points)

                ##  Swap the flag
                eve_ext_flag = True

        ##  The shell points we want to output
        lats_odd_shells = [ele for ele_list in odd_shells for ele in ele_list]
        lats_eve_shells = [ele for ele_list in eve_shells for ele in ele_list]

        existing_lats = center_list + lats_odd_shells + lats_eve_shells

        return existing_lats

    @staticmethod
    def get_center_of_curve(input_x, input_y):
        """
        Get the center of the 2D-curve numerically.

        #   Return
        xbar, ybar
        """
        diff_y = np.diff(input_y)
        diff_x = np.diff(input_x)

        der_y = diff_y / diff_x

        ds_factor = sqrt(1 + der_y**2)

        x_center_kernel = input_x[:-1] * ds_factor
        y_center_kernel = input_y[:-1] * ds_factor

        integral_s = np.trapz(ds_factor, input_x[:-1])

        integral_x = np.trapz(x_center_kernel, input_x[:-1])
        integral_y = np.trapz(y_center_kernel, input_x[:-1])

        x_bar = integral_x / integral_s
        y_bar = integral_y / integral_s

        return x_bar, y_bar

    @staticmethod
    def line_through_dot(dot_x_y, k):
        """
        Get the line function which goes through a dot
        """
        x0 = dot_x_y[0]
        y0 = dot_x_y[1]

        def line_func(x):
            y = k * (x - x0) + y0

            return y

        return line_func

    @staticmethod
    def line_incline_incept(k_incline, b_incept):
        """
        Get the line function given the incline and the incept
        """

        def line_func(x):
            y = k_incline * x + b_incept

            return y

        return line_func
