import openpyxl

import xlrd


class SheetReader:
    def range(self): return None
    def cell(self, row, column): return None
    def header(self, column): return None


class SheetReaderXls(SheetReader):
    def __init__(self, full_path):
        workbook = xlrd.open_workbook(full_path)
        self.sheet = workbook.sheet_by_index(0)

    def range(self): return range(0, self.sheet.nrows)
    def cell(self, row, column): return self.sheet.cell(row, column)
    def header(self, column): return self.sheet.cell(0, column)


class SheetReaderXlsx(SheetReader):
    def __init__(self, full_path):
        wb_obj = openpyxl.load_workbook(full_path)
        self.sheet = wb_obj.active

    def range(self): return self.sheet.iter_rows(max_row=self.sheet.max_row)
    def cell(self, row, column): return row[column]
    def header(self, column): return self.sheet.cell(1, column + 1)


def create_reader(path):
    extension = '.xlsx' if path.endswith('.xlsx') else '.xls'
    return SheetReaderXls(path) if extension == '.xls' else SheetReaderXlsx(path)
